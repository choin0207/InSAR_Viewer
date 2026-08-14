#!/usr/bin/env python3
"""insar_viewer.py — InSAR 時序檢視器: 讀取 InSAR 時序 HDF5 (/date + /timeseries),
產出單一互動 HTML 地圖 (點擊看時序 + 沿線距離-位移剖面+日期滑桿)。

使用範例:
  python3 insar_viewer.py --build InSAR_Viewer.html          # 單檔檢視器 (開檔後自選資料)
  python3 insar_viewer.py --ts timeseries.h5 --out map.html --pack   # 把資料烘進 HTML

輸入 (校正後產出, 皆同一 H×W 地理網格):
  --vel   校正後速度場 GeoTIFF (mm/yr, EPSG:4326)              [選填, 上色用;
          省略時: 若 --ts h5 root attrs 有 GEOTRANSFORM → 由 /timeseries 首尾期自算;
          否則自動找 --ts 同資料夾內唯一的 *vel*.tif, 找不到或有多個則報錯提示明確指定]
  --ts    校正後時序 h5 (/date {N}, /timeseries {N,H,W} mm)    [必要, 點時序+沿線動畫]
  --line  廊道中線 geojson (LineString/MultiLineString)         [沿線剖面用; 省略=只出地圖]
  --buffer 中線緩衝(公尺, 預設50)  只擷取此範圍內的 PS 進 HTML  → HTML 不會肥爆
  --coh   同調 GeoTIFF (同網格, 選填)   --coh-min 門檻(預設0.3)
  --gnss  GNSS 站位檔 (.csv/.txt/.geojson/.json, 選填)  自動辨識站名/經緯度欄;
          平面座標(|值|>360)視為 TWD97 EPSG:3826 自動轉經緯度; 只疊加 h5 範圍內的站
  --unit  顯示單位 mm/cm (預設: h5 attrs DISPLAY_UNIT, 否則 mm; 內部運算一律 mm)
  --vlim  色階上限(顯示單位): 給值時退回「對稱 ±vlim」; 預設為不對稱色階
          (負端=場的 p0.5、正端=p99.5, 速度場與總累積各自計算)
  --contour-int 等值線間距(顯示單位; 速度場為 顯示單位/yr), 預設 1, 0=關閉
  --max-points 上限(預設20000)
  --out   輸出 html (預設 <ts目錄>/gmtsar_timeseries_html_viewer.html)   --open 產完自動開瀏覽器

輸出: 自帶資料的 .html (Leaflet 衛星底圖 + PS 速度上色 + 點擊看時序 + 沿線距離-位移剖面+日期滑桿)。
      PS 資料內嵌(離線可看點/時序); 衛星底圖與 Leaflet/Chart.js 由 CDN 載(看底圖需連網)。
單位: 內部一律 mm; 顯示依 --unit / h5 DISPLAY_UNIT。官方分級圖配色: 綠≈穩定, 黃→橘→紅→桃紅→粉紅=下沉(負), 青→藍=抬升(正)。
"""
from __future__ import annotations
import argparse, base64, glob, gzip, json, os, subprocess, sys, webbrowser
from html import escape as _html_escape  # aliased: local vars named `html` (rendered page string) collide otherwise
import numpy as np


def _find_vel_candidates(ts_path):
    """在 --ts 同資料夾找候選速度場 GeoTIFF (檔名含 vel, 不分大小寫), 回傳排序後路徑清單."""
    d = os.path.dirname(os.path.abspath(ts_path)) or "."
    seen = set(); cands = []
    for pat in ("*vel*.tif", "*vel*.tiff", "*VEL*.tif", "*VEL*.tiff", "*Vel*.tif", "*Vel*.tiff"):
        for p in glob.glob(os.path.join(d, pat)):
            rp = os.path.abspath(p)
            if rp not in seen:
                seen.add(rp); cands.append(p)
    return sorted(cands)


def _year_span(dates_raw):
    """由 /date 首末字串算年跨距: 4 碼(YYYY)直接相減; 8 碼(YYYYMMDD)日差/365.25."""
    d0, d1 = dates_raw[0], dates_raw[-1]
    if len(d0) == 4 and len(d1) == 4:
        return float(int(d1) - int(d0))
    from datetime import datetime
    return (datetime.strptime(d1, "%Y%m%d") - datetime.strptime(d0, "%Y%m%d")).days / 365.25


def _vel_from_ts(tsd, dates_raw):
    """由 timeseries 首尾期算年化速度 (mm/yr); 任一期為 NaN 的像元結果也是 NaN."""
    first = tsd[0][...].astype(np.float64)
    last = tsd[tsd.shape[0] - 1][...].astype(np.float64)
    span = _year_span(dates_raw)
    with np.errstate(invalid="ignore"):
        vel = (last - first) / span
    return vel


# column-name keywords for GNSS files (matched case-insensitively, in order)
_GNSS_NAME_KEYS = ("station", "site", "name_c", "name_e", "name", "站名", "站碼")
_GNSS_LON_KEYS = ("lon", "longitude", "經度", "x", "e")
_GNSS_LAT_KEYS = ("lat", "latitude", "緯度", "y", "n")


def _match_col(headers, keys):
    """Return index of first header matching any key (case-insensitive), else -1."""
    low = [str(h).strip().lower() for h in headers]
    for k in keys:
        if k in low:
            return low.index(k)
    return -1


# GNSS solution archive (optional): one <STATION>_f_all.xlsx per station with
# columns Date/DOY/Year/N/E/h, where N/E are TWD97 (EPSG:3826) meters.
# Empty by default -- the public build ships no station coordinates; pass
# --gnss-dir / --gnss-ts-dir to point at your own archive.
_GNSS_ALL_DIR = ""


def _build_gnss_station_csv(gnss_dir, cache_csv):
    """Scan <STATION>_f_all.xlsx files, cache station,lon,lat as CSV.

    Station position = first data row's N/E converted to WGS84.  The cache
    is reused on later runs; delete the CSV to force a rescan (~0.5 s per
    station over CIFS).
    """
    if os.path.isfile(cache_csv):
        return cache_csv
    from openpyxl import load_workbook
    from pyproj import Transformer
    tr = Transformer.from_crs(3826, 4326, always_xy=True)
    files = sorted(glob.glob(os.path.join(gnss_dir, "*_f_all.xlsx")))
    if not files:
        raise FileNotFoundError(f"{gnss_dir} 內找不到 *_f_all.xlsx")
    rows = []
    for i, p in enumerate(files, 1):
        name = os.path.basename(p).split("_f_all")[0]
        try:
            wb = load_workbook(p, read_only=True)
            ws = wb[wb.sheetnames[0]]
            it = ws.iter_rows(min_row=1, max_row=2, values_only=True)
            hdr = [str(c).strip() for c in next(it)]
            first = next(it)
            wb.close()
            n = float(first[hdr.index("N")])
            e = float(first[hdr.index("E")])
            lo, la = tr.transform(e, n)
            rows.append((name, lo, la))
        except Exception as err:
            print(f"⚠ GNSS 站檔讀取失敗, 略過 {os.path.basename(p)}: {err}",
                  file=sys.stderr)
        if i % 50 == 0:
            print(f"  GNSS 站位掃描 {i}/{len(files)}", flush=True)
    with open(cache_csv, "w", encoding="utf-8") as f:
        f.write("station,lon,lat\n")
        for name, lo, la in rows:
            f.write(f"{name},{lo:.6f},{la:.6f}\n")
    print(f"GNSS 站位快取建立: {cache_csv} ({len(rows)} 站)", flush=True)
    return cache_csv


def _load_gnss(path):
    """讀 GNSS 站位檔 → [{'n': 站名, 'lon', 'lat'}] (WGS84 度).

    .csv/.txt: 自動偵測分隔符與 header, 欄位以關鍵字辨識 (站名/經度/緯度).
    .geojson/.json: 取 Point features, 站名從 properties 以同關鍵字找.
    座標 |值|>360 視為 TWD97 EPSG:3826 平面座標, 以 osr 轉 WGS84.
    """
    ext = os.path.splitext(path)[1].lower()
    recs = []                                     # [name, x, y]
    if ext in (".geojson", ".json"):
        gj = json.load(open(path, encoding="utf-8"))
        feats = gj["features"] if gj.get("type") == "FeatureCollection" else [gj]
        for f in feats:
            g = f.get("geometry") or {}
            if g.get("type") != "Point":
                continue
            props = f.get("properties") or {}
            keys = list(props.keys())
            ni = _match_col(keys, _GNSS_NAME_KEYS)
            name = str(props[keys[ni]]).strip() if ni >= 0 else ""
            recs.append([name, float(g["coordinates"][0]), float(g["coordinates"][1])])
    else:
        import csv
        with open(path, encoding="utf-8-sig", newline="") as fh:
            sample = fh.read(4096); fh.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t ")
            except csv.Error:
                dialect = csv.excel                 # fallback: comma
            rows = [r for r in csv.reader(fh, dialect) if r and any(c.strip() for c in r)]
        if not rows:
            raise ValueError(f"GNSS 檔無資料: {path}")
        ni = _match_col(rows[0], _GNSS_NAME_KEYS)
        xi = _match_col(rows[0], _GNSS_LON_KEYS)
        yi = _match_col(rows[0], _GNSS_LAT_KEYS)
        if xi >= 0 and yi >= 0:
            data_rows = rows[1:]                    # header row present
            if ni < 0: ni = 0
        else:
            ni, xi, yi = 0, 1, 2                    # no header: assume 站名,經度,緯度 欄序
            data_rows = rows
        for r in data_rows:
            try:
                x, y = float(r[xi]), float(r[yi])
            except (ValueError, IndexError):
                continue                            # skip non-numeric rows
            name = r[ni].strip() if 0 <= ni < len(r) else ""
            recs.append([name, x, y])
    # projected meters (TWD97 TM2) → WGS84 when values exceed degree range
    if any(abs(x) > 360 or abs(y) > 360 for _, x, y in recs):
        from osgeo import osr
        src = osr.SpatialReference(); src.ImportFromEPSG(3826)
        src.SetAxisMappingStrategy(osr.OAMS_TRADITIONAL_GIS_ORDER)
        dst = osr.SpatialReference(); dst.ImportFromEPSG(4326)
        dst.SetAxisMappingStrategy(osr.OAMS_TRADITIONAL_GIS_ORDER)
        tr = osr.CoordinateTransformation(src, dst)
        for r in recs:
            r[1], r[2], _ = tr.TransformPoint(r[1], r[2])
    return [{"n": n, "lon": round(x, 6), "lat": round(y, 6)} for n, x, y in recs]


# =============================================================================
# GNSS correction: align the InSAR velocity field / time series to GNSS station
# velocities (projected to LOS).  --gnss-correct opt-in only; see main() near the
# existing GNSS station-position block for how this wires in.
# =============================================================================

def _parse_gnss_date_str(s):
    """/date 字串 (YYYYMMDD 或 YYYY) -> datetime.date."""
    from datetime import date
    return date(int(s[:4]), int(s[4:6]), int(s[6:8])) if len(s) == 8 else date(int(s), 1, 1)


def _bilinear_sample(arr, gt, lon, lat, invalid_fn=None):
    """Bilinear-sample a (H,W) array at (lon,lat), pixel-center convention
    gt=(x0,dx,0,y0,0,dy) i.e. col=(lon-x0)/dx-0.5, row=(lat-y0)/dy-0.5 (matches how this
    file builds lon/lat from GEOTRANSFORM elsewhere, and GDAL's GetGeoTransform()).
    Falls back to the nearest valid pixel in the surrounding 3x3 window when any of the 4
    bilinear corners is invalid (NaN, or invalid_fn(value) is True); returns None if the
    3x3 window has no valid pixel either (caller should then drop/cull the station).
    """
    H, W = arr.shape
    x0, dx, _, y0, _, dy = gt

    def _valid(v):
        return np.isfinite(v) and not (invalid_fn and invalid_fn(v))

    col = (lon - x0) / dx - 0.5
    row = (lat - y0) / dy - 0.5
    c0, r0 = int(np.floor(col)), int(np.floor(row))
    fc, fr = col - c0, row - r0
    corners = []
    ok = True
    for rr, cc, w in ((r0, c0, (1 - fr) * (1 - fc)), (r0, c0 + 1, (1 - fr) * fc),
                      (r0 + 1, c0, fr * (1 - fc)), (r0 + 1, c0 + 1, fr * fc)):
        if 0 <= rr < H and 0 <= cc < W and _valid(arr[rr, cc]):
            corners.append((float(arr[rr, cc]), w))
        else:
            ok = False
    if ok:
        wsum = sum(w for _, w in corners)
        if wsum > 0:
            return sum(v * w for v, w in corners) / wsum
    rc, cc0 = int(round(row)), int(round(col))
    best = bestd = None
    for rr in range(rc - 1, rc + 2):
        for ccx in range(cc0 - 1, cc0 + 2):
            if 0 <= rr < H and 0 <= ccx < W and _valid(arr[rr, ccx]):
                d = (rr - row) ** 2 + (ccx - col) ** 2
                if best is None or d < bestd:
                    best, bestd = float(arr[rr, ccx]), d
    return best


def _enu2los(ve, vn, vu, inc_deg, az_deg):
    """ENU velocity -> LOS velocity.

    Verified against source code (not from memory), 2026-07-30:
      [1] MintPy src/mintpy/utils/utils0.py :: enu2los()
          https://github.com/insarlab/MintPy/blob/main/src/mintpy/utils/utils0.py
      [2] ISCE2 components/zerodop/topozero/Topozero.py (los.rdr band descriptions) +
          src/topozero.f90 (losang computation)
    inc_deg: incidence angle from LOCAL VERTICAL at the target, degrees, 0=zenith, always
             >=0 (== los.rdr band 1 / MintPy inc_angle; NOT look angle, NOT incLocal).
    az_deg : azimuth of the TARGET->SATELLITE vector, measured from North, ANTI-CLOCKWISE
             positive, degrees (== los.rdr band 2 raw value == MintPy az_angle).  Typical
             right-looking Sentinel-1: ascending ~102 (raw file ~-258, same angle mod 360),
             descending ~-102.
    Returns v_los with the MintPy/community sign convention: positive = toward satellite
    (uplift), negative = away from satellite (subsidence).  Verified against this viewer's
    own colour-scale convention (cbar_desc_vel: negative=下沉/subsidence, positive=抬升/
    uplift) — same sign, no flip needed for GNSS_LOS - InSAR residuals.

    v_los = ve*sin(inc)*sin(az)*(-1) + vn*sin(inc)*cos(az) + vu*cos(inc)   (inc,az in rad)

    Hand-check: ve=0,vn=0,vu=-10 mm/yr, inc=39°, descending az=-102° (heading=-168°) ->
    v_los = -10*cos(39°) = -7.7715 mm/yr (negative=away from satellite, consistent with a
    pure subsidence input).
    """
    inc = np.deg2rad(inc_deg)
    az = np.deg2rad(az_deg)
    return (np.asarray(ve) * np.sin(inc) * np.sin(az) * -1.0
            + np.asarray(vn) * np.sin(inc) * np.cos(az)
            + np.asarray(vu) * np.cos(inc))


def _heading_to_az(head_deg, look_direction="right"):
    """Along-track heading (from North, clockwise+) -> LOS target->satellite azimuth
    (from North, anti-clockwise+).  Source: MintPy utils0.heading2azimuth_angle()
    (hardcodes right-looking when not told otherwise; Sentinel-1 is right-looking)."""
    az = (head_deg - 90.0) * -1.0 if look_direction == "right" else (head_deg + 90.0) * -1.0
    return az - np.round(az / 360.0) * 360.0


_GNSS_FMT_FIELDS = ("date", "year", "doy", "n", "e", "u", "ignore")


def _parse_gnss_fmt(fmt_str):
    fields = [f.strip().lower() for f in fmt_str.split(",")]
    for f in fields:
        if f not in _GNSS_FMT_FIELDS:
            raise ValueError(f"--gnss-fmt 欄位名稱不支援: {f!r} (限 {'/'.join(_GNSS_FMT_FIELDS)})")
    if "date" not in fields and ("year" not in fields or "doy" not in fields):
        raise ValueError("--gnss-fmt 必須含 date, 或同時含 year 與 doy")
    return fields


def _read_gnss_series_default(station, ts_dir):
    """讀 <station>_f_all.xlsx (NAS 預設格式): 表頭 Date/DOY/Year/N/E/h, N/E 為 TWD97 TM2
    (EPSG:3826) 公尺絕對座標逐日解. 回傳 (dates, n_m, e_m, u_m) 或 None (檔案不存在)."""
    path = os.path.join(ts_dir, f"{station}_f_all.xlsx")
    if not os.path.isfile(path):
        return None
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()
    hdr = [str(c).strip() for c in rows[0]]
    di, ni, ei, hi = hdr.index("Date"), hdr.index("N"), hdr.index("E"), hdr.index("h")
    dates, n_m, e_m, u_m = [], [], [], []
    for r in rows[1:]:
        d = r[di]
        if d is None or r[ni] is None or r[ei] is None:
            continue
        d = d.date() if hasattr(d, "date") else _parse_gnss_date_str(str(d))
        dates.append(d)
        n_m.append(float(r[ni])); e_m.append(float(r[ei]))
        u_m.append(float(r[hi]) if r[hi] is not None else np.nan)
    return dates, np.array(n_m), np.array(e_m), np.array(u_m)


def _read_gnss_series_custom(station, ts_dir, fmt_fields, units, crs, mode, date_fmt, skip_rows):
    """自訂格式讀取 (.csv/.txt 逗號或空白分隔自動判斷, 或 .xlsx), 欄序依 --gnss-fmt.
    mode(abs/disp) 只影響語意標記: 線性回歸的斜率(速度)對絕對座標/相對位移皆等價, 不需分支.
    回傳 (dates, n_m, e_m, u_m) 皆已轉為公尺, 或 None (找不到檔案)."""
    cands = [p for p in glob.glob(os.path.join(ts_dir, f"{station}.*"))
             + glob.glob(os.path.join(ts_dir, f"{station}_*.*"))
             if os.path.splitext(p)[1].lower() in (".csv", ".txt", ".xlsx")]
    if not cands:
        return None
    cands = sorted(cands)
    path = cands[0]
    if len(cands) > 1:
        print(f"⚠ GNSS 站 {station} 時序檔比對到 {len(cands)} 個檔案, 取用 {os.path.basename(path)} "
              f"(略過 {', '.join(os.path.basename(c) for c in cands[1:])})")
    from datetime import datetime as _dt
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = [list(r) for r in ws.iter_rows(min_row=skip_rows + 1, values_only=True)]
        wb.close()
    else:
        with open(path, encoding="utf-8-sig") as fh:
            lines = fh.readlines()[skip_rows:]
        rows = []
        for ln in lines:
            ln = ln.strip()
            if ln:
                rows.append(ln.split(",") if "," in ln else ln.split())
    di = fmt_fields.index("date") if "date" in fmt_fields else None
    yi = fmt_fields.index("year") if "year" in fmt_fields else None
    doyi = fmt_fields.index("doy") if "doy" in fmt_fields else None
    ni = fmt_fields.index("n") if "n" in fmt_fields else None
    ei = fmt_fields.index("e") if "e" in fmt_fields else None
    ui = fmt_fields.index("u") if "u" in fmt_fields else None
    dates, nv, ev, uv = [], [], [], []
    for r in rows:
        if r is None or len(r) < len(fmt_fields):
            continue
        try:
            if di is not None:
                d = r[di].date() if hasattr(r[di], "date") else _dt.strptime(str(r[di]).strip(), date_fmt).date()
            else:
                from datetime import date, timedelta
                d = date(int(float(r[yi])), 1, 1) + timedelta(days=float(r[doyi]) - 1)
            n = float(r[ni]) if ni is not None else np.nan
            e = float(r[ei]) if ei is not None else np.nan
            u = float(r[ui]) if ui is not None else np.nan
        except (ValueError, TypeError, IndexError):
            continue
        dates.append(d); nv.append(n); ev.append(e); uv.append(u)
    nv, ev, uv = np.array(nv), np.array(ev), np.array(uv)
    if units == "mm":
        nv, ev, uv = nv / 1000.0, ev / 1000.0, uv / 1000.0
    if crs == "latlon":
        # n=lat, e=lon in degrees -> local equirectangular meters (slope-preserving approx)
        lat0 = np.nanmean(nv)
        ev = ev * 111320.0 * np.cos(np.deg2rad(lat0))
        nv = nv * 111320.0
    elif crs:
        # 任意 EPSG (經緯度或投影皆可): 用 pyproj 轉到 WGS84 再套同一套等距圓柱近似,
        # 取代舊版「非 latlon 就當作已經是公尺, 靜默忽略 --gnss-fmt-crs」的行為.
        try:
            epsg = int(crs)
        except ValueError:
            raise ValueError(f"--gnss-fmt-crs 不支援: {crs!r} (限 EPSG 整數 或 latlon)")
        from pyproj import Transformer
        tr = Transformer.from_crs(epsg, 4326, always_xy=True)
        lon_deg, lat_deg = tr.transform(ev, nv)
        lat0 = np.nanmean(lat_deg)
        ev = lon_deg * 111320.0 * np.cos(np.deg2rad(lat0))
        nv = lat_deg * 111320.0
    return dates, nv, ev, uv


# 若 InSAR 時窗 < 此年數, GNSS 擬合窗對稱外擴至此值 (受各站實際資料可得範圍限制,
# 外擴只是放寬查詢邊界, 實際可用樣本仍由站本身時序決定). 純常數, 不開放 CLI (見 S2).
_GNSS_SEASONAL_TARGET_YR = 2.0


def _fit_velocity_lsq(t_yr, y_m, seasonal):
    """對 (t_yr, y_m) 做最小二乘回歸 (自動剔除 NaN).
    seasonal=True : 設計矩陣 [1, t, sin(2*pi*t), cos(2*pi*t), sin(4*pi*t), cos(4*pi*t)]
                     (t=十進位年, 年 + 半年週期項, 降低短窗混疊).
    seasonal=False: 設計矩陣 [1, t] (純線性).
    回傳 (slope_mm_per_yr, resid_rms_mm, n_used); 樣本不足(rank 不足)時回傳 (nan, nan, n_used)."""
    valid = np.isfinite(y_m)
    n_used = int(valid.sum())
    ncoef = 6 if seasonal else 2
    if n_used < ncoef + 1:
        return np.nan, np.nan, n_used
    tt, yy = t_yr[valid], y_m[valid]
    if seasonal:
        A = np.column_stack([np.ones(n_used), tt, np.sin(2 * np.pi * tt), np.cos(2 * np.pi * tt),
                              np.sin(4 * np.pi * tt), np.cos(4 * np.pi * tt)])
    else:
        A = np.column_stack([np.ones(n_used), tt])
    coef, *_ = np.linalg.lstsq(A, yy, rcond=None)
    resid = yy - A @ coef
    rms_mm = float(np.sqrt(np.mean(resid ** 2))) * 1000.0
    return float(coef[1]) * 1000.0, rms_mm, n_used


def _gnss_station_velocity(station, ts_dir, default_ts_dir, fmt_fields, units, crs, mode,
                            date_fmt, skip_rows, tg0, tg1, min_days, min_span_yr, resid_max):
    """時窗 [tg0,tg1] (已依 InSAR 時窗外擴, 見 _GNSS_SEASONAL_TARGET_YR) 內回歸 GNSS 站
    N/E/h -> vn/ve/vu (mm/yr). 有效時窗 (實際資料涵蓋, 非請求窗) >=2.0 年用
    linear+annual+semiannual LSQ; 1.2<=有效時窗<2.0 年退回純線性並強烈警告季節混疊風險;
    任一分量殘差 RMS 超過 resid_max=(e_max,n_max,h_max) mm 即剔站(站況不佳, 如打點/跳點).
    回傳 (ve, vn, vu, None) 或 (None, None, None, 剔除原因字串)."""
    try:
        is_default = os.path.normpath(ts_dir) == os.path.normpath(default_ts_dir)
        res = (_read_gnss_series_default(station, ts_dir) if is_default else
               _read_gnss_series_custom(station, ts_dir, fmt_fields, units, crs, mode, date_fmt, skip_rows))
    except Exception as err:
        return None, None, None, f"讀取失敗: {err}"
    if res is None:
        return None, None, None, "找不到時序檔"
    dates, n_m, e_m, u_m = res
    if not dates:
        return None, None, None, "時序檔無有效資料列"
    order = np.argsort(dates)
    dates = [dates[i] for i in order]
    n_m, e_m, u_m = n_m[order], e_m[order], u_m[order]
    keep = np.array([tg0 <= d <= tg1 for d in dates])
    dates = [d for d, k in zip(dates, keep) if k]
    n_m, e_m, u_m = n_m[keep], e_m[keep], u_m[keep]
    n_eff = len(dates)
    if n_eff < min_days:
        return None, None, None, f"時窗內僅 {n_eff} 筆 (<{min_days})"
    span_yr = (dates[-1] - dates[0]).days / 365.25
    if span_yr < min_span_yr:
        return None, None, None, f"時窗內有效時距僅 {span_yr:.2f} 年 (<{min_span_yr})"
    t0d = dates[0]
    t_yr = np.array([(d - t0d).days / 365.25 for d in dates])
    seasonal = span_yr >= _GNSS_SEASONAL_TARGET_YR
    if not seasonal:
        print(f"  ⚠ GNSS 站 {station}: 有效時窗僅 {span_yr:.2f} 年 (<{_GNSS_SEASONAL_TARGET_YR:.1f}), "
              f"退回純線性回歸, 季節訊號可能混疊入速度估計")
    ve, rms_e, _ = _fit_velocity_lsq(t_yr, e_m, seasonal)
    vn, rms_n, _ = _fit_velocity_lsq(t_yr, n_m, seasonal)
    vu, rms_h, _ = _fit_velocity_lsq(t_yr, u_m, seasonal)
    if not (np.isfinite(ve) and np.isfinite(vn) and np.isfinite(vu)):
        return None, None, None, "N/E/h 迴歸失敗 (資料不足或全為 NaN)"
    e_max, n_max, h_max = resid_max
    bad = []
    if np.isfinite(rms_e) and rms_e > e_max:
        bad.append(f"E殘差RMS={rms_e:.1f}mm(>{e_max})")
    if np.isfinite(rms_n) and rms_n > n_max:
        bad.append(f"N殘差RMS={rms_n:.1f}mm(>{n_max})")
    if np.isfinite(rms_h) and rms_h > h_max:
        bad.append(f"h殘差RMS={rms_h:.1f}mm(>{h_max})")
    if bad:
        return None, None, None, "站況不佳剔除: " + "; ".join(bad)
    return ve, vn, vu, None


def _fit_gnss_correction(lons, lats, residuals, alpha=0.05):
    """n<6: 常數(中位數, b=c=0). n>=6: 中心化平面 a+b*(lon-lon0)+c*(lat-lat0) 最小二乘,
    F 檢定(平面 vs 常數, alpha=0.05)不顯著時降為常數(仍用中位數, 不用平面的 a).
    回傳 (model_fn(lon,lat)->mm/yr offset, mode, 係數 dict, f_info dict or None)."""
    n = len(residuals)
    resid = np.asarray(residuals, dtype=float)
    const_val = float(np.median(resid))
    f_info = None
    if n >= 6:
        lon0, lat0 = float(np.mean(lons)), float(np.mean(lats))
        dlon = np.asarray(lons, dtype=float) - lon0
        dlat = np.asarray(lats, dtype=float) - lat0
        A = np.column_stack([np.ones(n), dlon, dlat])
        coef, *_ = np.linalg.lstsq(A, resid, rcond=None)
        a, b, c = (float(x) for x in coef)
        sse_full = float(np.sum((resid - A @ coef) ** 2))
        sse_reduced = float(np.sum((resid - np.mean(resid)) ** 2))
        dfn, dfd = 2, n - 3
        if sse_full > 1e-12 and dfd > 0:
            fstat = ((sse_reduced - sse_full) / dfn) / (sse_full / dfd)
            from scipy.stats import f as _f_dist
            pval = float(_f_dist.sf(fstat, dfn, dfd))
        else:
            fstat, pval = float("inf"), 0.0
        f_info = {"f": round(fstat, 3), "p": round(pval, 4), "dfn": dfn, "dfd": dfd,
                   "sseConst": round(sse_reduced, 3), "ssePlane": round(sse_full, 3)}
        if pval < alpha:
            def model_fn(lon, lat, a=a, b=b, c=c, lon0=lon0, lat0=lat0):
                return a + b * (np.asarray(lon) - lon0) + c * (np.asarray(lat) - lat0)
            return model_fn, "plane", {"a": a, "b": b, "c": c, "lon0": lon0, "lat0": lat0}, f_info
        # F 檢定不顯著: 平面對殘差沒有顯著解釋力, 降為常數(中位數)

    def model_fn(lon, lat, v=const_val):
        return np.full(np.shape(np.asarray(lon, dtype=float)), v, dtype=float)
    return model_fn, "const", {"const": const_val}, f_info


def _gnss_correction_flow(gnss_stations, ts_dir, default_ts_dir, fmt_fields, fmt_units, fmt_crs,
                           fmt_mode, fmt_date, fmt_skip, dates_raw, vel, lon, lat, gt,
                           ensure_geometry, min_days=200, min_span_yr=1.2,
                           resid_max=(15.0, 15.0, 30.0), rms_warn=10.0):
    """共用 GNSS 校正流程: 站位篩選(呼叫方已做 extent 篩選) -> GNSS 速度(門檻剔站, 季節模型) ->
    (只有 >=1 站有可用 GNSS 速度時才呼叫 ensure_geometry() 取得/驗證幾何; 0 站或全被剔除
    時完全不要求 --los-*/--geom 參數; 幾何本身無法使用時與「0 站」分開報告) -> LOS 投影 ->
    InSAR 取樣(剔 NaN 站) -> 殘差 -> 站群共模揭露 -> 模型擬合(F 檢定). 回傳 (corr_grid
    (H,W) or None, gnss_corr_info dict)."""
    t0, t1 = _parse_gnss_date_str(dates_raw[0]), _parse_gnss_date_str(dates_raw[-1])
    span0_yr = (t1 - t0).days / 365.25
    if span0_yr < _GNSS_SEASONAL_TARGET_YR:
        pad_days = (_GNSS_SEASONAL_TARGET_YR * 365.25 - (t1 - t0).days) / 2.0
        from datetime import timedelta
        tg0, tg1 = t0 - timedelta(days=pad_days), t1 + timedelta(days=pad_days)
        print(f"GNSS 擬合窗已延伸至 {tg0.isoformat()}–{tg1.isoformat()} "
              f"(InSAR 時窗僅 {span0_yr:.2f} 年, 對稱外擴至 {_GNSS_SEASONAL_TARGET_YR:.1f} 年以降低季節混疊; "
              f"實際可用樣本仍受各站資料可得範圍限制)")
    else:
        tg0, tg1 = t0, t1
    if not gnss_stations:
        print("範圍內沒有 GNSS 測站，未做 GNSS 校正")
        return None, {"mode": "none"}
    vel_candidates = []
    for g in gnss_stations:
        ve, vn, vu, reason = _gnss_station_velocity(g["n"], ts_dir, default_ts_dir, fmt_fields,
                                                      fmt_units, fmt_crs, fmt_mode, fmt_date,
                                                      fmt_skip, tg0, tg1, min_days, min_span_yr,
                                                      resid_max)
        if ve is None:
            print(f"  GNSS 站 {g['n']} 剔除: {reason}")
            continue
        vel_candidates.append((g, ve, vn, vu))
    if not vel_candidates:
        print("範圍內沒有 GNSS 測站，未做 GNSS 校正")
        return None, {"mode": "none"}
    try:
        get_inc_az = ensure_geometry()      # 只有這裡才要求 LOS 幾何參數
    except Exception as err:
        print(f"LOS 幾何取得失敗（{err}），無法進行 GNSS 校正")
        return None, {"mode": "geom_fail"}
    used = []
    for g, ve, vn, vu in vel_candidates:
        inc, az = get_inc_az(g["lon"], g["lat"])
        if inc is None or az is None:
            print(f"  GNSS 站 {g['n']} 剔除: 無有效 LOS 幾何 (入射角/方位角取樣失敗)")
            continue
        insar_v = _bilinear_sample(vel, gt, g["lon"], g["lat"])
        if insar_v is None:
            print(f"  GNSS 站 {g['n']} 剔除: InSAR 速度場在此站無有效值")
            continue
        gnss_los = float(_enu2los(ve, vn, vu, inc, az))
        used.append({"n": g["n"], "lon": g["lon"], "lat": g["lat"], "ve": ve, "vn": vn, "vu": vu,
                      "inc": inc, "az": az, "gnss_los": gnss_los, "insar": insar_v,
                      "resid": gnss_los - insar_v})
    if not used:
        print("範圍內沒有 GNSS 測站，未做 GNSS 校正")
        return None, {"mode": "none"}
    lons = np.array([u["lon"] for u in used]); lats = np.array([u["lat"] for u in used])
    resid = np.array([u["resid"] for u in used])
    names = [u["n"] for u in used]

    # S4: 站群共模(平均 ENU)揭露 -- 這是 GNSS 參考框架本身的訊息(如 ITRF 板塊運動), 不是
    # 程式錯誤; 不自動 demean, 只印出讓使用者自行判斷是否與 InSAR 成品的框架一致.
    cm_ve, cm_vn, cm_vu = float(np.mean([u["ve"] for u in used])), \
        float(np.mean([u["vn"] for u in used])), float(np.mean([u["vu"] for u in used]))
    cm_inc, cm_az = float(np.mean([u["inc"] for u in used])), float(np.mean([u["az"] for u in used]))
    cm_los = float(_enu2los(cm_ve, cm_vn, cm_vu, cm_inc, cm_az))
    print(f"  站群共模 ve={cm_ve:.1f}/vn={cm_vn:.1f}/vu={cm_vu:.1f} mm/yr, "
          f"LOS 常數貢獻約 {cm_los:.1f} mm/yr —— 此值反映 GNSS 參考框架(如 ITRF 板塊運動), "
          f"對已在地方框架/已校正的 InSAR 成品套用會整體平移，請確認框架一致")

    if len(used) == 1:
        print(f"  單站對齊，殘差無法評估")
    model_fn, mode, coef, f_info = _fit_gnss_correction(lons, lats, resid)
    rms_before = float(np.sqrt(np.mean(resid ** 2)))
    rms_after = (None if len(used) == 1 else
                 float(np.sqrt(np.mean((resid - model_fn(lons, lats)) ** 2))))
    jj, ii = np.meshgrid(np.arange(vel.shape[1]), np.arange(vel.shape[0]))
    corr_grid = model_fn(lon[jj], lat[ii])
    print(f"GNSS 校正: {len(used)} 站 ({mode}), 用站={','.join(names)}")
    print(f"  校正前 RMS={rms_before:.3f} mm/yr, 校正後 RMS="
          + (f"{rms_after:.3f} mm/yr" if rms_after is not None else "無法評估(僅 1 站)"))
    print(f"  係數: {coef}")
    if f_info is not None:
        print(f"  F 檢定(平面 vs 常數): F={f_info['f']}, p={f_info['p']} "
              f"(df={f_info['dfn']},{f_info['dfd']}), SSE_const={f_info['sseConst']}, "
              f"SSE_plane={f_info['ssePlane']} -> {'採用平面' if mode == 'plane' else '不顯著, 降為常數'}")
    if rms_after is not None and rms_after > rms_warn:
        print(f"  ⚠⚠ 校正後 RMS={rms_after:.3f} mm/yr 超過警告門檻 {rms_warn} mm/yr —— "
              f"校正品質不佳，請檢查站況/幾何/模型選擇")
    if mode == "plane":
        lo_span = float(lon[-1]) - float(lon[0]) if lon[-1] >= lon[0] else float(lon[0]) - float(lon[-1])
        la_span = float(lat[-1]) - float(lat[0]) if lat[-1] >= lat[0] else float(lat[0]) - float(lat[-1])
        lo_cov = (float(lons.max()) - float(lons.min())) / lo_span if lo_span > 0 else 1.0
        la_cov = (float(lats.max()) - float(lats.min())) / la_span if la_span > 0 else 1.0
        print(f"  站群涵蓋範圍佔格網: 經度 {lo_cov*100:.0f}%, 緯度 {la_cov*100:.0f}%")
        if lo_cov < 0.6 or la_cov < 0.6:
            print(f"  ⚠ 平面外插範圍大 (站群涵蓋 <60%)，格網邊緣的校正值為外插結果，可信度較低")
    info = {"mode": mode, "nsta": len(used), "stations": names,
            "rmsBefore": round(rms_before, 3),
            "rmsAfter": (round(rms_after, 3) if rms_after is not None else None),
            "rmsWarn": bool(rms_after is not None and rms_after > rms_warn),
            "coef": coef, "commonMode": {"ve": round(cm_ve, 2), "vn": round(cm_vn, 2),
                                          "vu": round(cm_vu, 2), "losConst": round(cm_los, 2)}}
    return corr_grid, info


def _make_los_rdr_sampler(path):
    """開 ISCE2 geocoded 2-band LOS 檔 (.rdr/.vrt): band1=入射角(deg,自天頂,0=nodata),
    band2=方位角(deg,自北逆時針,目標->衛星; Topozero.py 慣例). 回傳 get_inc_az(lon,lat)
    取樣函式 (雙線性 + 3x3 最近有效像元後援, incidence==0 視為 nodata).

    效能: 全解析度 los.rdr 可達 GB 級, 只有少量站要取樣 -> 逐站在目標像元周圍讀一個小視窗
    (band.ReadAsArray(xoff,yoff,...)), 不整檔載入; 視窗附帶的 geotransform 相應平移
    (x0+xoff*dx, y0+yoff*dy), 讓 _bilinear_sample 的 col/row 換算對得上這個小視窗而非
    原始全圖 (直接沿用原始 gt 會讓每一站的入射角/方位角悄悄取樣到錯誤像元).

    幾何失效偵測 (與「範圍內 0 站」區分, 見 _gnss_correction_flow 的 try/except):
      - GeoTransform 為 GDAL 未設定時的預設 identity 值 -> 無地理參考, 直接 raise.
      - band2 不存在時 get_inc_az.has_az=False, 呼叫方需自行決定要不要用 --los-heading
        後援 (未給後援時應在呼叫端 raise, 不要讓每一站各自默默剔除)."""
    from osgeo import gdal
    ds = gdal.Open(path)
    if ds is None:
        raise FileNotFoundError(f"無法開啟 --los-rdr: {path}")
    gt = ds.GetGeoTransform()
    if gt == (0.0, 1.0, 0.0, 0.0, 0.0, 1.0):
        raise ValueError(f"{path} 無有效地理參考(GeoTransform)")
    W, H = ds.RasterXSize, ds.RasterYSize
    band1 = ds.GetRasterBand(1)
    band2 = ds.GetRasterBand(2) if ds.RasterCount >= 2 else None
    x0, dx, _, y0, _, dy = gt

    def _invalid_inc(v):
        return v == 0.0

    def get_inc_az(lon, lat):
        col = (lon - x0) / dx - 0.5
        row = (lat - y0) / dy - 0.5
        c0, r0 = int(np.floor(col)), int(np.floor(row))
        pad = 3   # >=3px 邊界, 讓 _bilinear_sample 既有的 3x3 後援仍有空間可搜尋
        xoff, yoff = max(0, c0 - pad), max(0, r0 - pad)
        xend, yend = min(W, c0 + pad + 2), min(H, r0 + pad + 2)
        xsize, ysize = xend - xoff, yend - yoff
        if xsize <= 0 or ysize <= 0:
            return None, None
        gt_win = (x0 + xoff * dx, dx, 0.0, y0 + yoff * dy, 0.0, dy)
        inc_win = band1.ReadAsArray(xoff, yoff, xsize, ysize).astype(np.float64)
        inc = _bilinear_sample(inc_win, gt_win, lon, lat, invalid_fn=_invalid_inc)
        if inc is None:
            return None, None
        az = None
        if band2 is not None:
            az_win = band2.ReadAsArray(xoff, yoff, xsize, ysize).astype(np.float64)
            az = _bilinear_sample(az_win, gt_win, lon, lat, invalid_fn=_invalid_inc)
        return inc, az

    get_inc_az.has_az = band2 is not None
    # GDAL gotcha: Band objects keep a *weak* link to their parent Dataset in this binding;
    # once _make_los_rdr_sampler() returns, the local `ds` would be garbage-collected and any
    # later band1/band2.ReadAsArray() call segfaults (use-after-free). Anchor `ds` to the
    # closure's lifetime so it stays alive exactly as long as get_inc_az does.
    get_inc_az._ds_keepalive = ds
    return get_inc_az


# sentinel wrapped around the "gnss" key when json.dumps()-ing the data dict, so the gnss
# array can be located/replaced later by regex (in-browser CSV reload + re-export; see
# gnssCsvFile handler in _HTML_TEMPLATE). Must match the JS-side GNSS_PLACEHOLDER used by
# buildShareHTML() in _APP_JS.
_GNSS_PLACEHOLDER = "__LEVELING_GNSS_PLACEHOLDER__"
_GNSS_MARK_S, _GNSS_MARK_E = "/*__GNSS_S__*/", "/*__GNSS_E__*/"


def _embed_gnss_json(data_json, gnss_list):
    """Replace the `"gnss": "__LEVELING_GNSS_PLACEHOLDER__"` placeholder (inserted into the
    dict before json.dumps) with the real gnss array wrapped in /*__GNSS_S/E__*/ comment
    markers, so a browser can later regex-locate and swap just this segment (used by the
    in-browser "載入 GNSS CSV / 匯出更新 HTML" feature to produce a chainable, re-loadable
    export)."""
    needle = json.dumps(_GNSS_PLACEHOLDER)
    replacement = _GNSS_MARK_S + json.dumps(gnss_list, ensure_ascii=False) + _GNSS_MARK_E
    out, n = data_json.replace(needle, replacement, 1), data_json.count(needle)
    if n != 1:
        raise RuntimeError(f"_embed_gnss_json: expected 1 placeholder, found {n}")
    return out


# asymmetric colormap stops — MUST mirror the JS NEG_STOPS/POS_STOPS in _HTML_TEMPLATE
# Classified-style palette matching an official 1 cm/yr classified legend (
# classes).  17 equidistant stops per side: with --vlim 80 each step is
# 5 mm/yr, class-center colors sit at odd stops and class boundaries blend.
# negative (subsidence): green → yellow-green → yellow → orange → red →
# magenta → pink → pale pink at -negLim
_NEG_STOPS = np.array([[128, 255, 64], [128, 255, 64], [160, 255, 64],
                       [192, 255, 64], [224, 240, 64], [255, 224, 64],
                       [255, 192, 64], [255, 160, 64], [255, 128, 64],
                       [255, 96, 64], [255, 96, 128], [255, 96, 192],
                       [255, 128, 224], [255, 160, 255], [255, 192, 255],
                       [255, 224, 255], [255, 240, 255]], dtype=np.float64)
# positive (uplift): green → cyan → light blue → blue at +posLim
_POS_STOPS = np.array([[64, 255, 64], [64, 255, 64], [64, 255, 160],
                       [64, 255, 255], [64, 240, 255], [64, 224, 255],
                       [64, 208, 255], [64, 192, 255], [32, 160, 255],
                       [0, 128, 255], [0, 96, 255], [0, 64, 255],
                       [0, 56, 255], [0, 48, 255], [0, 40, 255],
                       [0, 32, 255], [0, 24, 255]], dtype=np.float64)


def _interp_stops(stops, s):
    """Piecewise-linear interpolation over equidistant stops; s in [0,1] → (...,3) rounded."""
    x = s * (len(stops) - 1)
    i = np.clip(np.floor(x).astype(np.int64), 0, len(stops) - 2)
    f = (x - i)[..., None]
    return np.round(stops[i] + (stops[i + 1] - stops[i]) * f)


def _colormap_rgba(arr, neg_lim, pos_lim):
    """Asymmetric two-sided colormap → (H,W,4) uint8 RGBA; NaN → alpha 0.

    Negative side normalized by neg_lim (clamped to purple beyond), positive
    side by pos_lim (clamped to blue). Mirrors the JS col(v, lims) exactly.
    """
    finite = np.isfinite(arr)
    v = np.where(finite, arr, 0.0)
    rgb_neg = _interp_stops(_NEG_STOPS, np.clip(-v / neg_lim, 0.0, 1.0))
    rgb_pos = _interp_stops(_POS_STOPS, np.clip(v / pos_lim, 0.0, 1.0))
    rgb = np.where((v >= 0)[..., None], rgb_pos, rgb_neg)
    out = np.zeros(arr.shape + (4,), dtype=np.uint8)
    out[..., :3] = rgb.astype(np.uint8)
    out[..., 3] = np.where(finite, 255, 0)
    return out


def _asym_lims(field):
    """Asymmetric color limits for a display-unit field: neg=|p0.5|, pos=p99.5.

    Guards: a missing side falls back to 5% of the other (never 0 or negative).
    """
    fin = field[np.isfinite(field)]
    if fin.size == 0:
        return 1.0, 1.0
    lo = float(np.percentile(fin, 0.5))
    hi = float(np.percentile(fin, 99.5))
    neg = -lo if lo < 0 else 0.0
    pos = hi if hi > 0 else 0.0
    if neg <= 0:
        neg = max(pos * 0.05, 1e-6)
    if pos <= 0:
        pos = max(neg * 0.05, 1e-6)
    return neg, pos


def _round_coords(obj, ndig=5):
    """Recursively round GeoJSON coordinate floats to ndig decimals (size reduction)."""
    if isinstance(obj, list):
        return [_round_coords(x, ndig) for x in obj]
    if isinstance(obj, float):
        return round(obj, ndig)
    return obj


def _make_contours(arr, gt, interval, tag):
    """gdal.ContourGenerate on a display-unit grid → GeoJSON FeatureCollection.

    Lines simplified with ~half-pixel tolerance; coordinates rounded to 5 decimals.
    Each feature carries properties.lev (contour level, display units).
    """
    from osgeo import gdal, ogr, osr
    gdal.UseExceptions()
    H, W = arr.shape
    nodata = -99999.0
    mem = gdal.GetDriverByName("MEM").Create("", W, H, 1, gdal.GDT_Float32)
    mem.SetGeoTransform(gt)
    band = mem.GetRasterBand(1)
    band.WriteArray(np.where(np.isfinite(arr), arr, nodata).astype(np.float32))
    band.SetNoDataValue(nodata)
    vds = ogr.GetDriverByName("Memory").CreateDataSource(f"contour_{tag}")
    srs = osr.SpatialReference(); srs.ImportFromEPSG(4326)
    lyr = vds.CreateLayer("contour", srs, ogr.wkbLineString)
    lyr.CreateField(ogr.FieldDefn("ID", ogr.OFTInteger))
    lyr.CreateField(ogr.FieldDefn("elev", ogr.OFTReal))
    gdal.ContourGenerate(band, interval, 0.0, [], 1, nodata, lyr, 0, 1)
    tol = abs(gt[1]) * 0.5                      # ~half pixel simplify tolerance
    feats = []
    for f in lyr:
        geom = f.GetGeometryRef().Simplify(tol)
        gj = json.loads(geom.ExportToJson())
        gj["coordinates"] = _round_coords(gj["coordinates"])
        feats.append({"type": "Feature",
                      "properties": {"lev": round(f.GetField("elev"), 3)},
                      "geometry": gj})
    return {"type": "FeatureCollection", "features": feats}


def _rgba_to_png_b64(rgba, tag):
    """RGBA (H,W,4) uint8 → PNG bytes via gdal /vsimem/ → (base64 str, byte size)."""
    from osgeo import gdal
    gdal.UseExceptions()
    H, W = rgba.shape[:2]
    mem = gdal.GetDriverByName("MEM").Create("", W, H, 4, gdal.GDT_Byte)
    for b in range(4):
        mem.GetRasterBand(b + 1).WriteArray(rgba[..., b])
    vpath = f"/vsimem/overlay_{tag}.png"
    gdal.GetDriverByName("PNG").CreateCopy(vpath, mem)
    fh = gdal.VSIFOpenL(vpath, "rb")
    gdal.VSIFSeekL(fh, 0, 2)
    size = gdal.VSIFTellL(fh)
    gdal.VSIFSeekL(fh, 0, 0)
    data = gdal.VSIFReadL(1, size, fh)
    gdal.VSIFCloseL(fh)
    gdal.Unlink(vpath)
    return base64.b64encode(data).decode(), int(size)


def _rasterize_display_mask(path, gt, H, W):
    """GeoJSON Polygon/MultiPolygon → bool mask (H,W) on the display grid.

    |x|>360 coords are treated as TWD97 EPSG:3826 and reprojected first.
    Vector source is opened with gdal.OpenEx (older GDAL's Rasterize rejects
    ogr.Open handles — see rules/50-known-pitfalls.md).
    """
    from osgeo import gdal, osr
    gdal.UseExceptions()
    gj = json.load(open(path, encoding="utf-8"))
    feats = gj["features"] if gj.get("type") == "FeatureCollection" else [gj]

    def _maxabs(o):
        if isinstance(o, list):
            return max((_maxabs(x) for x in o), default=0.0)
        return abs(o) if isinstance(o, (int, float)) else 0.0

    src_path = path
    if max(_maxabs((f.get("geometry") or f).get("coordinates", [])) for f in feats) > 360:
        src = osr.SpatialReference(); src.ImportFromEPSG(3826)
        src.SetAxisMappingStrategy(osr.OAMS_TRADITIONAL_GIS_ORDER)
        dst = osr.SpatialReference(); dst.ImportFromEPSG(4326)
        dst.SetAxisMappingStrategy(osr.OAMS_TRADITIONAL_GIS_ORDER)
        tr = osr.CoordinateTransformation(src, dst)

        def _tx(o):
            if isinstance(o[0], (int, float)):
                x, y, _ = tr.TransformPoint(float(o[0]), float(o[1]))
                return [round(x, 8), round(y, 8)]
            return [_tx(x) for x in o]

        for f in feats:
            g = f.get("geometry") or f
            g["coordinates"] = _tx(g["coordinates"])
        buf = json.dumps({"type": "FeatureCollection",
                          "features": [{"type": "Feature", "properties": {},
                                        "geometry": (f.get("geometry") or f)} for f in feats]})
        src_path = "/vsimem/mask_wgs84.json"
        gdal.FileFromMemBuffer(src_path, buf.encode("utf-8"))
    vec = gdal.OpenEx(src_path, gdal.OF_VECTOR)       # NOT ogr.Open (pitfall)
    mem = gdal.GetDriverByName("MEM").Create("", W, H, 1, gdal.GDT_Byte)
    mem.SetGeoTransform(gt)
    srs = osr.SpatialReference(); srs.ImportFromEPSG(4326)
    mem.SetProjection(srs.ExportToWkt())
    gdal.Rasterize(mem, vec, burnValues=[1])
    arr = mem.ReadAsArray().astype(bool)
    vec = None
    if src_path.startswith("/vsimem/"):
        gdal.Unlink(src_path)
    return arr


def _read_tif(path):
    from osgeo import gdal
    gdal.UseExceptions()
    ds = gdal.Open(path)
    a = ds.GetRasterBand(1).ReadAsArray().astype("float64")
    gt = ds.GetGeoTransform()
    H, W = a.shape
    # 像素中心 lon/lat (EPSG:4326 假設)
    lon = gt[0] + (np.arange(W) + 0.5) * gt[1]
    lat = gt[3] + (np.arange(H) + 0.5) * gt[5]
    return a, lon, lat, gt


def _load_line_utm(geojson_path, epsg_utm=32651):
    """讀中線 geojson → 合併成 shapely 線(UTM). 回 (line_utm, transformer_fwd)."""
    import shapely
    from shapely.geometry import shape
    from shapely.ops import linemerge, transform as shp_transform
    from pyproj import Transformer
    gj = json.load(open(geojson_path))
    lines = []                                    # 攤平: feature 可能是 MultiLineString
    for f in gj["features"]:
        g = shape(f["geometry"])
        if g.geom_type == "MultiLineString":
            lines.extend(list(g.geoms))
        elif g.geom_type == "LineString":
            lines.append(g)
    if not lines:
        raise ValueError("中線 geojson 無 LineString")
    merged = linemerge(lines) if len(lines) > 1 else lines[0]
    tr = Transformer.from_crs(4326, epsg_utm, always_xy=True)
    line_utm = shp_transform(lambda x, y, z=None: tr.transform(x, y), merged)
    return line_utm, tr


_L_CSS = "https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"
_L_JS = "https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"
_L_ROT = "https://unpkg.com/leaflet-rotate@0.2.8/dist/leaflet-rotate-src.js"
_L_CHART = "https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"
_TILE = "https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}"
_EXPORT = "https://server.arcgisonline.com/arcgis/rest/services/World_Imagery/MapServer/export"
# h5wasm: 瀏覽器端直讀 HDF5 (「載入 HDF5」按鈕). iife 單檔已把 hdf5_util.wasm 以
# base64 data URL 內含, 不需另外載入 .wasm 檔; 版本鎖定避免 CDN 改版行為不一致.
_H5WASM_VER = "0.7.5"
_H5WASM_JS = f"https://cdn.jsdelivr.net/npm/h5wasm@{_H5WASM_VER}/dist/iife/h5wasm.js"


def _online_assets():
    """線上版: CDN 函式庫 + 底圖切換(衛星影像/OpenStreetMap) (檔案小, 但看底圖/圖表需連網)."""
    libs = (f'<link rel="stylesheet" href="{_L_CSS}"/>\n<script src="{_L_JS}"></script>\n'
            f'<script src="{_L_ROT}"></script>\n<script src="{_L_CHART}"></script>\n'
            # defer: 4 MB 的 h5wasm 不擋首頁渲染; 頁面 script 執行時 h5wasm 可能尚未就緒,
            # 故按鈕在「點擊當下」才檢查 typeof h5wasm (見 template 的 loadH5File)
            f'<script defer src="{_H5WASM_JS}"></script>')
    basemap = (
        f"const _sat=L.tileLayer('{_TILE}',{{maxZoom:19,crossOrigin:'anonymous',attribution:'Esri World Imagery'}});\n"
        "const _osm=L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png',"
        "{maxZoom:19,crossOrigin:'anonymous',attribution:'© OpenStreetMap contributors'});\n"
        "_sat.addTo(map);\n"
        "window._lyrCtrl=L.control.layers({[t('base_satellite')]:_sat,'OpenStreetMap':_osm},null,"
        "{position:'topleft',collapsed:false}).addTo(map);\n"
        "window.__i18nRegisterLayer(()=>_sat,'base_satellite');")
    # 圖層控制器改放右上並預設收合: 左上是主控制面板 (InSAR_Viewer 多了 AOI 與 GNSS 兩區),
    # 留在 topleft 會被面板整個蓋住
    # 維持 collapsed:false: 收合模式需要 Leaflet 的 layers.png 圖示, file:// 開檔時
    # 相對 CDN 的圖片路徑抓不到 (ERR_FILE_NOT_FOUND)
    basemap = basemap.replace("{position:'topleft',collapsed:false}",
                              "{position:'topright',collapsed:false}")
    return libs, basemap


def _pack_assets(with_h5=True):
    """打包版: 內嵌函式庫(不依賴 CDN) + 線上底圖(僅需可連 Esri/OSM 圖磚). 適合把單一 HTML 分享給其他電腦.

    with_h5=False 時不內嵌 h5wasm: --app 版的 libs 會被嵌入兩次 (head 原文 + 分享用模板的
    gzip+base64), 而 grd 直讀 app 用不到 HDF5 直讀, 內嵌會平白多出約 6 MB.
    """
    import urllib.request

    def get(u):
        return urllib.request.urlopen(u, timeout=60).read().decode("utf-8")
    print("  打包: 下載內嵌函式庫...")
    css, js1, js2, js3 = get(_L_CSS), get(_L_JS), get(_L_ROT), get(_L_CHART)
    libs = (f"<style>{css}</style>\n<script>{js1}</script>\n"
            f"<script>{js2}</script>\n<script>{js3}</script>")
    if with_h5:
        print("  打包: 下載內嵌 h5wasm (瀏覽器端直讀 HDF5, 約 4 MB, 含 base64 wasm)...")
        h5js = get(_H5WASM_JS)
        # a literal "</script" anywhere in the bundle would end the tag early and
        # silently truncate the page (currently none, kept as a guard)
        h5js = h5js.replace("</script", "<\\/script")
        libs += f"\n<script>{h5js}</script>"
    _, basemap = _online_assets()
    return libs, basemap


_SHEETJS_URL = "https://cdn.jsdelivr.net/npm/xlsx@0.18.5/dist/xlsx.full.min.js"


def _sheetjs_asset():
    """內嵌 SheetJS (約 0.9 MB): 瀏覽器端解析 GNSS 時序 .xlsx (<站碼>_f_all.xlsx)."""
    import urllib.request
    print("  打包: 下載內嵌 SheetJS (瀏覽器端讀 .xlsx, 約 0.9 MB)...")
    js = urllib.request.urlopen(_SHEETJS_URL, timeout=60).read().decode("utf-8")
    js = js.replace("</script", "<\\/script")     # 同 h5wasm: 避免提早關閉 <script>
    return f"<script>{js}</script>"


def _auto_gnss_stations(a):
    """--build 時的站位清單: 有 NAS 就重建快取, 否則用腳本目錄既有的 gnss_stations_all.csv."""
    cache = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                         "gnss_stations_all.csv")
    gdir = getattr(a, "gnss_dir", "") or ""
    if gdir and os.path.isdir(gdir):
        try:
            return _load_gnss(_build_gnss_station_csv(gdir, cache))
        except Exception as err:
            print(f"⚠ GNSS 站位重建失敗, 改用既有快取: {err}", file=sys.stderr)
    if os.path.isfile(cache):
        try:
            return _load_gnss(cache)
        except Exception as err:
            print(f"⚠ GNSS 站位快取讀取失敗, 本次不內嵌站位: {err}", file=sys.stderr)
    return []


def _offline_assets(pts_json):
    """離線版: 內嵌 Leaflet/Chart.js + 產生時抓該區一張衛星影像嵌入(imageOverlay). 完全離線可開."""
    import urllib.request, base64
    def get(u, txt=True):
        r = urllib.request.urlopen(u, timeout=60).read()
        return r.decode("utf-8") if txt else r
    print("  離線: 下載內嵌函式庫...")
    css, js1, js2, js3 = get(_L_CSS), get(_L_JS), get(_L_ROT), get(_L_CHART)
    libs = (f"<style>{css}</style>\n<script>{js1}</script>\n"
            f"<script>{js2}</script>\n<script>{js3}</script>")
    lons = [p["lon"] for p in pts_json]; lats = [p["lat"] for p in pts_json]
    pad = 0.02
    x0, x1 = min(lons) - pad, max(lons) + pad
    y0, y1 = min(lats) - pad, max(lats) + pad
    W = 1600; H = max(400, min(int(W * (y1 - y0) / (x1 - x0)), 2000))
    print(f"  離線: 抓 AOI 衛星影像 {W}x{H}...")
    img = base64.b64encode(get(f"{_EXPORT}?bbox={x0},{y0},{x1},{y1}"
                               f"&bboxSR=4326&imageSR=4326&size={W},{H}&format=jpg&f=image", txt=False)).decode()
    basemap = (f"L.imageOverlay('data:image/jpeg;base64,{img}',[[{y0},{x0}],[{y1},{x1}]]).addTo(map);")
    return libs, basemap


def _build_app(a):
    """產出 InSAR_Viewer 單檔 app: 開啟後選時序 HDF5 或 GMTSAR 資料夾, 全部計算在 JS 完成.

    重用 _HTML_TEMPLATE 的呈現層: 把 viewer script 包成 startViewer(D),
    前面加上 檔案解析/內插/速度回歸模組 (_APP_JS), 函式庫以 _pack_assets 內嵌
    (含 h5wasm 直讀 HDF5 與 SheetJS 讀 GNSS .xlsx).
    """
    gnss_all = _load_gnss(a.gnss) if a.gnss else _auto_gnss_stations(a)
    libs, basemap = _pack_assets(with_h5=True)    # 需要 h5wasm: 兩種來源都可能是 HDF5
    libs += "\n" + _sheetjs_asset()
    # split at the main viewer script (unique marker — the minimap has its own earlier <script>)
    head, script = _HTML_TEMPLATE.split("<script>\nconst D = /*__DATA__*/;", 1)
    viewer_js = ("// D comes in as the startViewer argument"
                 + script.rsplit("</script></body></html>", 1)[0]).replace("/*__BASEMAP__*/", basemap)
    head = (head.replace("<!--__LIBS__-->", libs)
                .replace("</head><body>", "</head><body class=\"preload\">")
                .replace("</style>",
                         " body.preload .panel:not(#apppanel):not(#miniwin){display:none}\n</style>")
                .replace("<!--__TITLE__-->", _html_escape(a.title)))
    # 匯出分享版用模板: 與 --ts --pack 產出完全同構 (含內嵌函式庫/小地圖/__test),
    # gzip+base64 內嵌 (同時避免內文 </script> 打斷 app 頁)
    tpl_share = (_HTML_TEMPLATE.replace("<!--__LIBS__-->", libs)
                 .replace("/*__BASEMAP__*/", basemap)
                 .replace("<!--__TITLE__-->", _html_escape(a.title)))
    tpl_b64 = base64.b64encode(gzip.compress(tpl_share.encode("utf-8"), 6)).decode()
    html = (head + _APP_PANEL
            + "<script>const APP_SHARE_TPL_B64=\"" + tpl_b64 + "\";</script>\n"
            + "<script>\nwindow.GNSS_ALL = " + json.dumps(gnss_all, ensure_ascii=False) + ";\n"
            + "function startViewer(D){\n" + viewer_js + "\n}\n"
            + _APP_JS + "</script></body></html>")
    with open(a.app, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"✓ 已產出 InSAR_Viewer {a.app} "
          f"({os.path.getsize(a.app)/1e6:.1f} MB, 內嵌 GNSS 站位 {len(gnss_all)} 站)")
    return a.app


def _run_one_key(a):
    """--ts-dir 一鍵模式: output_data/<name>TS.h5 與 <name>_vel.tif 不存在(或 --rebuild)時
    以 subprocess 呼叫 gmtsar2h5.py 轉檔, 再把 a.ts/a.vel/a.out 指到轉檔產物,
    交由既有 main() 流程接手 (--title/--vlim/--pack/--gnss 等全部照舊可用)."""
    name = a.name or os.path.basename(os.path.normpath(a.ts_dir))
    outdir = os.path.join(os.getcwd(), "output_data")
    os.makedirs(outdir, exist_ok=True)
    h5_path = os.path.join(outdir, f"{name}TS.h5")
    vel_path = os.path.join(outdir, f"{name}_vel.tif")

    need_build = a.rebuild or not (os.path.isfile(h5_path) and os.path.isfile(vel_path))
    if need_build:
        script = os.path.join(os.path.dirname(os.path.abspath(__file__)), "gmtsar2h5.py")
        cmd = [sys.executable, script,
               "--ts-dir", a.ts_dir,
               "--out-h5", h5_path,
               "--out-vel", vel_path]
        if a.boundary:
            cmd += ["--boundary", a.boundary]
        if a.vel_geojson:
            cmd += ["--vel-geojson", a.vel_geojson]
        if a.deramp:
            cmd += ["--deramp", a.deramp]
        if a.grid_step is not None:
            cmd += ["--grid-step", str(a.grid_step)]
        if a.max_ps_dist_km is not None:
            cmd += ["--max-ps-dist-km", str(a.max_ps_dist_km)]
        if a.k_neigh is not None:
            cmd += ["--k-neigh", str(a.k_neigh)]
        print(f"一鍵模式: 轉檔 {' '.join(cmd)}", flush=True)
        rc = subprocess.call(cmd)
        if rc != 0:
            sys.exit(f"一鍵模式: 轉檔失敗 (gmtsar2h5.py 結束碼 {rc})")
    else:
        print(f"一鍵模式: 重用既有 {h5_path} 與 {vel_path}(加 --rebuild 可強制重跑)", flush=True)

    if not a.ts:
        a.ts = h5_path
    if not a.vel:
        a.vel = vel_path
    if not a.out:
        a.out = os.path.join(outdir, f"{name}_map_share.html")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--vel", required=False, default="")
    ap.add_argument("--ts", required=False, default="")
    ap.add_argument("--ts-dir", default="",
                    help="一鍵模式: GMTSAR 時序資料夾 (內有 disp_NNN_ll.xy 與 data_date.txt), "
                         "與 --ts 互斥; 未給 --ts/--vel/--out 時自動轉檔並接手既有流程")
    ap.add_argument("--name", default="",
                    help="一鍵模式輸出檔名前綴 (預設 = --ts-dir 的 basename)")
    ap.add_argument("--boundary", default="",
                    help="[一鍵模式 passthrough] AOI 邊界 geojson, 轉給 gmtsar2h5.py")
    ap.add_argument("--vel-geojson", default="",
                    help="[一鍵模式 passthrough] 轉給 gmtsar2h5.py")
    ap.add_argument("--deramp", default="",
                    help="[一鍵模式 passthrough] 格式 C0,C1,C2,LON0,LAT0, 轉給 gmtsar2h5.py")
    ap.add_argument("--grid-step", type=float, default=None,
                    help="[一鍵模式 passthrough] 轉給 gmtsar2h5.py")
    ap.add_argument("--max-ps-dist-km", type=float, default=None,
                    help="[一鍵模式 passthrough] 轉給 gmtsar2h5.py")
    ap.add_argument("--k-neigh", type=int, default=None,
                    help="[一鍵模式 passthrough] 轉給 gmtsar2h5.py")
    ap.add_argument("--rebuild", action="store_true",
                    help="一鍵模式: 強制重跑轉檔 (預設: 產物已存在就重用)")
    ap.add_argument("--build", "--app", dest="app", default="",
                    help="產出 InSAR_Viewer 單檔 HTML (不需 --ts; 使用者開檔後自行選時序 HDF5 "
                         "或 GMTSAR 資料夾, 讀檔/內插/速度回歸/繪圖全在瀏覽器內完成)")
    ap.add_argument("--line", default="")
    ap.add_argument("--buffer", type=float, default=50.0)
    ap.add_argument("--coh", default="")
    ap.add_argument("--gnss", default="",
                    help="GNSS 站位檔 (.csv/.txt/.geojson/.json), 只疊加 h5 範圍內的站")
    ap.add_argument("--gnss-dir", default=_GNSS_ALL_DIR,
                    help="未給 --gnss 時自動掃此資料夾的 <站碼>_f_all.xlsx 建站位快取, "
                         "自動疊加範圍內的站 (設空字串停用自動搜尋)")
    ap.add_argument("--gnss-correct", action="store_true",
                    help="啟用 GNSS 校正: 用站速度(投影 LOS)對齊 InSAR 速度場與時序 "
                         "(預設關閉, 不影響既有流程)")
    ap.add_argument("--gnss-ts-dir", default=_GNSS_ALL_DIR,
                    help="GNSS 時序資料夾 (預設同 NAS 站位路徑); 指向非預設資料夾時必須另給 --gnss-fmt")
    ap.add_argument("--gnss-fmt", default="",
                    help="自訂 GNSS 時序格式: 逗號列出欄位順序, 欄名限 date/year/doy/n/e/u/ignore "
                         "(例 \"date,n,e,u\"); 只在 --gnss-ts-dir 指向非預設資料夾時使用/必要")
    ap.add_argument("--gnss-fmt-units", choices=["m", "mm"], default="m",
                    help="自訂格式 n/e/u 欄位單位 (預設 m)")
    ap.add_argument("--gnss-fmt-crs", default="",
                    help="自訂格式座標系: EPSG 整數 (任意, 經緯度或投影皆可, 用 pyproj 轉 WGS84) "
                         "或 latlon (等同 EPSG:4326, n=lat,e=lon, 度); "
                         "預設(空字串)視為已是公尺平面座標 (同 NAS TM2 格式, 不需轉換)")
    ap.add_argument("--gnss-fmt-mode", choices=["abs", "disp"], default="abs",
                    help="自訂格式 n/e/u 為絕對座標(abs)或相對位移(disp); 站位座標一律取自站位快取/檔。"
                         "註: 目前僅語意標記, 兩者計算完全相同 (斜率對絕對座標/相對位移等價)")
    ap.add_argument("--gnss-fmt-date", default="%Y%m%d",
                    help="自訂格式 date 欄的 strptime 格式 (預設 %%Y%%m%%d)")
    ap.add_argument("--gnss-fmt-skip", type=int, default=1,
                    help="自訂格式表頭列數 (預設 1)")
    ap.add_argument("--gnss-resid-max", default="15,15,30",
                    help="GNSS 站速度回歸殘差 RMS 剔站門檻, 逗號分隔 \"E,N,h\" (mm, 預設 15,15,30); "
                         "任一分量超過即剔站(站況不佳, 如打點/跳點)")
    ap.add_argument("--gnss-min-days", type=int, default=200,
                    help="GNSS 站有效時窗內最少樣本數門檻 (預設 200), 不足即剔站")
    ap.add_argument("--gnss-min-span", type=float, default=1.2,
                    help="GNSS 站有效時窗最短年數門檻 (預設 1.2), 不足即剔站; "
                         ">=2.0 年用 linear+annual+semiannual, 介於門檻與 2.0 年間退回純線性並警告")
    ap.add_argument("--gnss-rms-warn", type=float, default=10.0,
                    help="GNSS 校正後 RMS 超過此值 (mm/yr, 預設 10.0) 印出品質警告 (終端 + HTML badge)")
    ap.add_argument("--los-rdr", default="",
                    help="ISCE2 geocoded LOS 幾何檔 (.rdr/.vrt, band1=入射角[deg,自天頂,0=nodata], "
                         "band2=方位角[deg,自北逆時針,目標→衛星], 選填); --gnss-correct 時可取代 "
                         "--los-inc/--los-heading (band2 缺時仍需 --los-heading 補方位角)")
    ap.add_argument("--los-inc", type=float, default=None,
                    help="LOS 入射角(度, 自天頂); --gnss-correct 且無 --los-rdr 時必填")
    ap.add_argument("--los-heading", type=float, default=None,
                    help="衛星飛行方向 heading(度, 自北順時針); --gnss-correct 且無 --los-rdr(含band2) 時必填 "
                         "(Sentinel-1 升軌 heading≈-13°、降軌≈193°，僅供參考請以軌道資料為準)")
    ap.add_argument("--mask", default="",
                    help="GeoJSON Polygon/MultiPolygon 遮罩: 只顯示多邊形內的資料 "
                         "(座標 |x|>360 視為 TWD97 EPSG:3826 自動轉 WGS84)")
    ap.add_argument("--coh-min", type=float, default=0.3)
    ap.add_argument("--unit", choices=["mm", "cm"], default="",
                    help="顯示單位 (預設: h5 attrs DISPLAY_UNIT, 否則 mm); 內部運算一律 mm")
    ap.add_argument("--vlim", type=float, default=0.0,
                    help="給值時色階退回對稱 ±vlim (顯示單位); 預設不對稱 (p0.5/p99.5)")
    ap.add_argument("--contour-int", type=float, default=0.1,
                    help="等值線基準間距 (顯示單位; 速度場為 顯示單位/yr), 0=關閉")
    ap.add_argument("--max-points", type=int, default=20000)
    ap.add_argument("--utm-epsg", type=int, default=32651)
    ap.add_argument("--title", default="InSAR Viewer",
                    help="頁面標題前綴 (預設 InSAR Viewer)")
    ap.add_argument("--out", default="")
    ap.add_argument("--open", action="store_true")
    ap.add_argument("--offline", action="store_true",
                    help="離線版: 內嵌 Leaflet/Chart.js + 產生時抓該區衛星影像嵌入(imageOverlay), 完全離線可開")
    ap.add_argument("--pack", action="store_true",
                    help="打包分享版: 內嵌函式庫(不依賴 CDN) + 線上底圖(對方電腦只需一般網路即可開)")
    ap.add_argument("--no-grid", action="store_true",
                    help="不內嵌全解析度時序格網(關閉 HTML 內互動剖面畫線/匯入取樣功能, 檔案較小)")

    # argparse 只認得 "-18.1" 是負數 (非 option); "--deramp -18.1,...,-74.9,..."
    # 因為含逗號, 不符合負數判斷, 會被誤認成一個不存在的 option (同 gmtsar2h5.py
    # parse_args 的既有 workaround). 把 "--deramp VALUE" 改寫成 "--deramp=VALUE"
    # 讓原本 space-separated 的 CLI 用法照樣可行。
    argv = sys.argv[1:]
    fixed = []
    i = 0
    while i < len(argv):
        if argv[i] == "--deramp" and i + 1 < len(argv):
            fixed.append(f"--deramp={argv[i + 1]}")
            i += 2
        else:
            fixed.append(argv[i])
            i += 1
    a = ap.parse_args(fixed)

    if a.ts and a.ts_dir:
        ap.error("--ts 與 --ts-dir 互斥, 請擇一使用")

    if a.app and a.gnss_correct:
        ap.error("--app 模式不支援 --gnss-correct (瀏覽器端 grd 直讀 app 尚未實作 GNSS 校正管線)")

    if a.app:
        return _build_app(a)

    if a.ts_dir:
        _run_one_key(a)

    if not a.ts:
        ap.error("--ts 為必要參數 (或用 --ts-dir 一鍵模式 / --app 產出 grd 直讀 app)")

    import h5py, shapely

    ff = h5py.File(a.ts, "r")
    dates_raw = [d.decode() if isinstance(d, bytes) else str(d) for d in ff["date"][:]]
    dates = [f"{d[:4]}-{d[4:6]}-{d[6:8]}" if len(d) == 8 else d for d in dates_raw]
    tsd = ff["timeseries"]
    N = tsd.shape[0]

    # vel_algo 標記速度場來源演算法, 隨 data 內嵌供 JS GRID 後援 (computeFieldFromGridT)
    # 與畫面實際上色的場同步 (同源缺陷修復): "ols"=逐格最小二乘回歸 (--vel 外部 GeoTIFF,
    # 由 gmtsar2h5 產製, 與 h5OlsVel 全期回歸誤差 <1e-6 mm/yr); "endpoint"=首尾期簡單差分.
    if a.vel:
        # (a) 明確指定 --vel: 完全走既有流程 (回溯相容)
        vel, lon, lat, gt = _read_tif(a.vel)
        vel_algo = "ols"
    elif "GEOTRANSFORM" in ff.attrs:
        # (b) 未指定 --vel, 但 h5 自帶地理參考: 由 timeseries 首尾期自算速度場
        H, W = tsd.shape[1:]
        gt = tuple(float(x) for x in ff.attrs["GEOTRANSFORM"])
        lon = gt[0] + (np.arange(W) + 0.5) * gt[1]
        lat = gt[3] + (np.arange(H) + 0.5) * gt[5]
        vel = _vel_from_ts(tsd, dates_raw)
        vel_algo = "endpoint"
        unit = ff.attrs.get("UNIT", b"mm")
        unit = unit.decode() if isinstance(unit, (bytes, bytearray)) else unit
        print(f"未指定 --vel: 由 h5 GEOTRANSFORM attrs + /timeseries 首尾期自算速度場 (單位={unit}/yr)")
    else:
        # (c) 未指定 --vel 且 h5 無地理參考 attrs: 到 --ts 同資料夾找唯一的 *vel*.tif
        cands = _find_vel_candidates(a.ts)
        if len(cands) == 1:
            vel, lon, lat, gt = _read_tif(cands[0])
            vel_algo = "ols"
            print(f"未指定 --vel: 自動使用 {cands[0]}")
        elif len(cands) == 0:
            sys.exit(f"未指定 --vel, h5 亦無 GEOTRANSFORM attrs, 且 {os.path.dirname(os.path.abspath(a.ts)) or '.'} "
                     f"找不到 *vel*.tif, 請明確指定 --vel")
        else:
            sys.exit("未指定 --vel, h5 亦無 GEOTRANSFORM attrs, 同資料夾有多個候選速度場 GeoTIFF: "
                     + ", ".join(cands) + " — 請明確指定 --vel")

    H, W = vel.shape
    coh = None
    if a.coh and os.path.isfile(a.coh):
        coh, _, _, _ = _read_tif(a.coh)
        if coh.shape != vel.shape:
            print(f"⚠ 同調維度 {coh.shape} != 速度 {vel.shape}, 忽略同調", file=sys.stderr); coh = None

    if tsd.shape[1:] != (H, W):
        sys.exit(f"時序網格 {tsd.shape[1:]} != 速度 {(H, W)}, 兩者需同網格")

    # GeoJSON 遮罩: rasterize 到顯示格網, 遮罩外設 NaN (下游 vel/lims/rasters/等值線/grid/points 全部生效)
    dmask = None
    if a.mask:
        dmask = _rasterize_display_mask(a.mask, gt, H, W)
        nb = int(np.isfinite(vel).sum())
        vel = np.where(dmask, vel, np.nan)
        print(f"遮罩 {os.path.basename(a.mask)}: 有效像元 {nb} -> {int(np.isfinite(vel).sum())}")

    # GNSS 站位: 未指定 --gnss 時自動掃 --gnss-dir 建站位快取
    if not a.gnss and a.gnss_dir:
        if os.path.isdir(a.gnss_dir):
            cache = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                 "gnss_stations_all.csv")
            try:
                a.gnss = _build_gnss_station_csv(a.gnss_dir, cache)
                print(f"未指定 --gnss: 自動使用站位快取 {os.path.basename(cache)}"
                      f" (來源 {a.gnss_dir})")
            except Exception as err:
                print(f"⚠ GNSS 自動站位失敗, 本次不疊 GNSS: {err}",
                      file=sys.stderr)
        else:
            print(f"⚠ GNSS 資料夾不存在, 本次不疊 GNSS: {a.gnss_dir}",
                  file=sys.stderr)

    # GNSS 站位: 讀檔 + 只留 h5 格網 bbox 內的站
    gnss = []
    if a.gnss:
        gnss_all = _load_gnss(a.gnss)
        lo0, lo1 = float(min(lon[0], lon[-1])), float(max(lon[0], lon[-1]))
        la0, la1 = float(min(lat[0], lat[-1])), float(max(lat[0], lat[-1]))
        gnss = [g for g in gnss_all
                if lo0 <= g["lon"] <= lo1 and la0 <= g["lat"] <= la1]
        print(f"GNSS: 檔內 {len(gnss_all)} 站, 範圍內 {len(gnss)} 站")

    # GNSS 校正 (選用, 預設關閉): 用站速度(投影 LOS)對齊 vel/時序. 只有 --gnss-correct 時才跑;
    # 校正後 vel 直接覆寫, 下游所有單期消費者(rasters/contours/pts_json/lims)自動吃到校正值,
    # 只有隨時間變化的量(series/cum_full/GRID full)需要另外乘上 (t_e - t0) 年再加(見下方三處).
    gnss_corr_grid = None
    gnss_corr_years = None
    gnss_corr_info = {"mode": "off"}
    if a.gnss_correct:
        if os.path.normpath(a.gnss_ts_dir) != os.path.normpath(_GNSS_ALL_DIR) and not a.gnss_fmt:
            ap.error("--gnss-ts-dir 指向非預設資料夾時必須提供 --gnss-fmt 說明每欄格式 (見 --help)")
        try:
            fmt_fields = _parse_gnss_fmt(a.gnss_fmt) if a.gnss_fmt else None
        except ValueError as err:
            ap.error(str(err))

        def _ensure_geometry():
            # 注意: 這裡 raise 的例外(非 ap.error)會被 _gnss_correction_flow 接住, 印出
            # 「LOS 幾何取得失敗」訊息並回傳 mode='geom_fail' -- 與「範圍內真的 0 站」分開
            # 報告(S12), 不再讓 band2 缺失這類資料問題偽裝成「0 站」靜默結束.
            if a.los_rdr:
                sampler = _make_los_rdr_sampler(a.los_rdr)
                if not sampler.has_az and a.los_heading is None:
                    raise RuntimeError(f"{a.los_rdr} 只有 1 band(無方位角) 且未給 --los-heading 補方位角")
                if a.los_heading is not None:
                    fallback_az = _heading_to_az(a.los_heading)

                    def get_inc_az(lo, la, _s=sampler, _fb=fallback_az):
                        inc, az = _s(lo, la)
                        return (inc, az if az is not None else (_fb if inc is not None else None))
                    return get_inc_az
                return sampler
            if a.los_inc is not None and a.los_heading is not None:
                const_inc, const_az = a.los_inc, _heading_to_az(a.los_heading)
                return lambda lo, la, _i=const_inc, _a=const_az: (_i, _a)
            ap.error("--gnss-correct 需要 LOS 幾何: 提供 --los-rdr，或同時提供 --los-inc 與 --los-heading "
                     "(Sentinel-1 升軌 heading≈-13°、降軌≈193°，僅供參考請以軌道資料為準)")

        try:
            resid_max = tuple(float(x) for x in a.gnss_resid_max.split(","))
            if len(resid_max) != 3:
                raise ValueError
        except ValueError:
            ap.error(f"--gnss-resid-max 格式錯誤: {a.gnss_resid_max!r} (須為 \"E,N,h\" 三個數字)")

        gnss_corr_grid, gnss_corr_info = _gnss_correction_flow(
            gnss, a.gnss_ts_dir, _GNSS_ALL_DIR, fmt_fields, a.gnss_fmt_units, a.gnss_fmt_crs,
            a.gnss_fmt_mode, a.gnss_fmt_date, a.gnss_fmt_skip, dates_raw, vel, lon, lat, gt,
            _ensure_geometry, min_days=a.gnss_min_days, min_span_yr=a.gnss_min_span,
            resid_max=resid_max, rms_warn=a.gnss_rms_warn)
        if gnss_corr_grid is not None:
            vel = vel + gnss_corr_grid
            _t0 = _parse_gnss_date_str(dates_raw[0])
            gnss_corr_years = np.array([(_parse_gnss_date_str(d) - _t0).days / 365.25 for d in dates_raw])

    # 有效點: 速度有限 (+同調門檻)
    m = np.isfinite(vel)
    if coh is not None:
        m &= np.isfinite(coh) & (coh >= a.coh_min)
    jj, ii = np.meshgrid(np.arange(W), np.arange(H))
    LON = lon[jj]; LAT = lat[ii]

    line_utm = tr = None
    if a.line and os.path.isfile(a.line):
        line_utm, tr = _load_line_utm(a.line, a.utm_epsg)
        # 先用中線 bbox(+緩衝)在 lon/lat 粗篩, 避免對全圖轉投影
        minx, miny, maxx, maxy = line_utm.bounds
        from pyproj import Transformer
        inv = Transformer.from_crs(a.utm_epsg, 4326, always_xy=True)
        c = [inv.transform(x, y) for x, y in [(minx, miny), (maxx, maxy)]]
        lo0, la0 = c[0]; lo1, la1 = c[1]
        pad = a.buffer / 1e5 * 1.5   # ~m→deg 粗放
        m &= (LON >= min(lo0, lo1) - pad) & (LON <= max(lo0, lo1) + pad)
        m &= (LAT >= min(la0, la1) - pad) & (LAT <= max(la0, la1) + pad)

    idx = np.where(m)
    plon = LON[idx]; plat = LAT[idx]; pvel = vel[idx]
    pcoh = coh[idx] if coh is not None else np.full(plon.shape, np.nan)
    pi = idx[0]; pj = idx[1]

    station = None
    if line_utm is not None and plon.size:
        from scipy.spatial import cKDTree
        px, py = tr.transform(plon, plat)          # vectorized → UTM
        # 中線加密成點建 KDTree, 對全部候選查最近距離做快速粗篩 (避免對百萬點做 shapely 距離)
        segs = list(line_utm.geoms) if line_utm.geom_type == "MultiLineString" else [line_utm]
        step = max(a.buffer / 2.0, 5.0)
        lpx = []
        for s in segs:
            n = max(2, int(s.length / step) + 1)
            for dd in np.linspace(0, s.length, n):
                p = s.interpolate(dd); lpx.append((p.x, p.y))
        tree = cKDTree(np.asarray(lpx))
        dist, _ = tree.query(np.c_[px, py])        # 近似最近距離(m)
        c = dist <= a.buffer + step                # 粗篩(加密殘差, 稍放寬)
        plon, plat, pvel, pcoh, pi, pj, kx, ky = (plon[c], plat[c], pvel[c], pcoh[c],
                                                  pi[c], pj[c], px[c], py[c])
        # 對粗篩後少量點精算: 精確點到線距離 + 沿線距離(station)
        dex = shapely.distance(shapely.points(kx, ky), line_utm)
        k = dex <= a.buffer
        plon, plat, pvel, pcoh, pi, pj = plon[k], plat[k], pvel[k], pcoh[k], pi[k], pj[k]
        station = shapely.line_locate_point(line_utm, shapely.points(kx[k], ky[k]))
        print(f"廊道 buffer {a.buffer:.0f}m 內 PS: {plon.size}")
    if plon.size == 0:
        sys.exit("範圍內無 PS (檢查中線是否與速度場重疊 / 放大 buffer / 降 coh-min)")

    # 上限: 超過就依同調(或隨機)抽樣
    if plon.size > a.max_points:
        order = np.argsort(-(pcoh if coh is not None else pvel))[:a.max_points] \
            if coh is not None else np.random.default_rng(0).choice(plon.size, a.max_points, replace=False)
        plon, plat, pvel, pcoh, pi, pj = plon[order], plat[order], pvel[order], pcoh[order], pi[order], pj[order]
        if station is not None: station = station[order]

    # 里程歸零(起點=0) + 廊道中線折線(給地圖畫線+標起訖) + 廊道方位角(初始地圖旋轉對齊)
    line_lonlat = []; bearing = 0.0
    if station is not None and line_utm is not None:
        amin = float(station.min()); amax = float(station.max())
        station = station - amin
        from pyproj import Transformer
        inv = Transformer.from_crs(a.utm_epsg, 4326, always_xy=True)
        ns = min(400, max(2, int((amax - amin) / 100) + 1))
        for s in np.linspace(amin, amax, ns):
            p = line_utm.interpolate(float(s)); lo, la = inv.transform(p.x, p.y)
            line_lonlat.append([round(lo, 6), round(la, 6)])
        p0 = line_utm.interpolate(amin); p1 = line_utm.interpolate(amax)
        bearing = float(np.degrees(np.arctan2(p1.x - p0.x, p1.y - p0.y)) % 360)

    # 取每點時序 (mm). h5 fancy-index 需排序; 逐點讀太慢→整段讀該 bbox 再取
    i0, i1, j0, j1 = pi.min(), pi.max() + 1, pj.min(), pj.max() + 1
    block = tsd[:, i0:i1, j0:j1]                       # N × h × w
    series = block[:, pi - i0, pj - j0].T              # P × N (mm)
    if gnss_corr_grid is not None:
        series = series + gnss_corr_grid[pi, pj][:, None] * gnss_corr_years[None, :]

    # 顯示單位: CLI --unit > h5 attrs DISPLAY_UNIT > 預設 mm (內部運算保持 mm)
    unit = a.unit
    if not unit:
        u = ff.attrs.get("DISPLAY_UNIT", b"")
        unit = u.decode() if isinstance(u, (bytes, bytearray)) else str(u)
    if unit not in ("mm", "cm"):
        unit = "mm"
    uscale = 0.1 if unit == "cm" else 1.0     # display value = mm value * uscale
    ndig = 2 if unit == "cm" else 1           # keep 0.1 mm precision in either unit
    print(f"顯示單位: {unit}")

    # 不對稱色階上限 (對全格網顯示值算): 負端=|p0.5|, 正端=p99.5; --vlim 給值時退回對稱
    cum_full = tsd[N - 1][...].astype(np.float64)
    if gnss_corr_grid is not None:
        cum_full = cum_full + gnss_corr_grid * gnss_corr_years[-1]
    if dmask is not None:
        cum_full = np.where(dmask, cum_full, np.nan)
    vel_d = vel * uscale
    cum_d = cum_full * uscale
    neg_v, pos_v = _asym_lims(vel_d)
    if a.vlim:
        neg_v = pos_v = a.vlim                # explicit --vlim: symmetric limits, new colormap
    neg_c, pos_c = _asym_lims(cum_d)
    lims = {"vel": {"neg": round(neg_v, 3), "pos": round(pos_v, 3)},
            "cum": {"neg": round(neg_c, 3), "pos": round(pos_c, 3)}}
    print(f"色階: 速度 -{neg_v:.2f}/+{pos_v:.2f} {unit}/yr, 總累積 -{neg_c:.2f}/+{pos_c:.2f} {unit}")

    # 半透明 raster 疊加: 速度場 + 總累積 ts[-1] → RGBA PNG 內嵌 (不對稱色階, 顯示單位)
    png_vel, sz_vel = _rgba_to_png_b64(_colormap_rgba(vel_d, neg_v, pos_v), "vel")
    png_cum, sz_cum = _rgba_to_png_b64(_colormap_rgba(cum_d, neg_c, pos_c), "cum")
    # imageOverlay corners = grid outer edges (gt origin is the outer corner, not pixel center)
    lat_edges = (gt[3], gt[3] + gt[5] * H)
    lon_edges = (gt[0], gt[0] + gt[1] * W)
    rb = [[min(lat_edges), min(lon_edges)], [max(lat_edges), max(lon_edges)]]
    rasters = {"vel": "data:image/png;base64," + png_vel,
               "cum": "data:image/png;base64," + png_cum,
               "bounds": [[round(rb[0][0], 6), round(rb[0][1], 6)],
                          [round(rb[1][0], 6), round(rb[1][1], 6)]]}
    print(f"  raster 疊加 PNG: 速度場 {sz_vel/1e6:.1f} MB, 總累積 {sz_cum/1e6:.1f} MB")

    # 等值線 (顯示單位; 速度場間距=contour-int 顯示單位/yr)
    contours = None
    if a.contour_int > 0:
        fc_vel = _make_contours(vel_d, gt, a.contour_int, "vel")
        fc_cum = _make_contours(cum_d, gt, a.contour_int, "cum")
        contours = {"vel": fc_vel, "cum": fc_cum, "base": a.contour_int}
        csz = len(json.dumps(contours))
        print(f"  等值線 (間距 {a.contour_int:g} {unit}): 速度場 {len(fc_vel['features'])} 條, "
              f"總累積 {len(fc_cum['features'])} 條, 內嵌 {csz/1e6:.2f} MB")

    # 點資料一律以顯示單位嵌入 (HTML 端不再換算)
    pts_json = []
    for k in range(plon.size):
        rec = {"lon": round(float(plon[k]), 6), "lat": round(float(plat[k]), 6),
               "v": round(float(pvel[k]) * uscale, ndig),
               "t": [round(float(x) * uscale, ndig) if np.isfinite(x) else None for x in series[k]]}
        if coh is not None: rec["c"] = round(float(pcoh[k]), 2)
        if station is not None: rec["s"] = round(float(station[k]), 0)
        pts_json.append(rec)

    # 全解析度時序格網內嵌 (供 HTML 端互動剖面雙線性取樣; 量化 int16 + gzip 壓縮縮小體積)
    grid = None
    if not a.no_grid:
        raw_bytes = N * H * W * 4                      # 原始 float32 (N,H,W) 大小
        if raw_bytes > 200 * 1024 * 1024:
            print(f"⚠ 略過全解析度格網內嵌: 原始資料 {raw_bytes/1e6:.0f} MB 超過 200MB 上限 "
                  f"(互動剖面功能將停用; 可忽略此警告或加 --no-grid 明確停用)", file=sys.stderr)
        else:
            print(f"  量化並壓縮全解析度時序格網 ({N}x{H}x{W})...")
            full = tsd[:].astype(np.float64)            # N,H,W (mm)
            if gnss_corr_grid is not None:
                full = full + gnss_corr_grid[None, :, :] * gnss_corr_years[:, None, None]
            if dmask is not None:
                full = np.where(dmask[None, ...], full, np.nan)
            finite = np.isfinite(full)
            GRID_SCALE = 0.1
            GRID_NODATA = -32768
            scaled = np.where(finite, np.clip(np.round(full / GRID_SCALE), -32767, 32767), GRID_NODATA)
            q = scaled.astype(np.int16)
            raw = q.tobytes()
            comp = gzip.compress(raw, compresslevel=6, mtime=0)
            b64 = base64.b64encode(comp).decode()
            # embed scale in display units so JS gridSample returns display values
            grid = {"b64": b64, "shape": [N, H, W], "gt": [float(x) for x in gt],
                    "scale": GRID_SCALE * uscale, "nodata": GRID_NODATA}
            print(f"  格網: raw {len(raw)/1e6:.1f} MB → gzip {len(comp)/1e6:.1f} MB → b64 {len(b64)/1e6:.1f} MB")

    data = {"dates": dates, "unit": unit, "lims": lims, "hasLine": station is not None,
            "n": len(pts_json), "buffer": a.buffer, "line": line_lonlat,
            "bearing": round(bearing, 1), "points": pts_json, "grid": grid,
            "gnss": _GNSS_PLACEHOLDER, "velAlgo": vel_algo,
            "rasters": rasters, "contours": contours}
    if a.gnss_correct:
        data["gnssCorr"] = gnss_corr_info
    out = a.out or os.path.join(os.path.dirname(os.path.abspath(a.ts)), "mtinsar_viewer.html")
    libs, basemap = _online_assets()
    if a.pack:
        libs, basemap = _pack_assets()              # 內嵌函式庫 + 線上底圖 (分享用)
    if a.offline:
        libs, basemap = _offline_assets(pts_json)   # 內嵌函式庫 + 抓該區衛星影像嵌入
    data_json = _embed_gnss_json(json.dumps(data, ensure_ascii=False), gnss)
    html = (_HTML_TEMPLATE.replace("/*__DATA__*/", data_json)
            .replace("<!--__LIBS__-->", libs).replace("/*__BASEMAP__*/", basemap)
            .replace("<!--__TITLE__-->", _html_escape(a.title)))
    if a.offline:
        # offline build has no online tiles: hide the overview minimap
        html = html.replace("/*__MINIOFF__*/false", "/*__MINIOFF__*/true")
    with open(out, "w", encoding="utf-8") as f:
        f.write(html)
    sz = os.path.getsize(out) / 1e6
    print(f"✓ 已產出 {out}  ({len(pts_json)} 點, {N} 期, {sz:.1f} MB)")
    if a.open:
        webbrowser.open("file://" + os.path.abspath(out))
    return out


_HTML_TEMPLATE = r"""<!doctype html><html lang="zh-Hant"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title><!--__TITLE__--> 時序檢視器</title>
<!--__LIBS__-->
<style>
 :root{--fs:12px;--fs-sm:11px}
 html,body{margin:0;height:100%;font-family:system-ui,"Noto Sans TC",sans-serif;overflow:hidden}
 #map{position:absolute;inset:0}
 /* 面板統一白底 (半透明), 文字改深色對比; 小螢幕安全上限見下方 @media */
 .panel{position:absolute;background:rgba(255,255,255,.92);color:#1e2933;border-radius:8px;
   box-shadow:0 3px 14px rgba(0,0,0,.35);z-index:1000;font-size:var(--fs);
   max-width:calc(100vw - 16px);max-height:calc(100vh - 16px);box-sizing:border-box}
 /* cbar 為可拖曳+可縮放小視窗 (.win 的 min 尺寸對它太大, 個別覆寫) */
 #cbar{top:auto;bottom:12px;left:12px;width:clamp(170px,15vw,240px);height:auto;min-width:170px;min-height:0}
 #cbar .body{gap:2px}
 #ctrl{top:10px;left:56px;padding:6px 10px;max-width:min(58vw,760px)}
 .win{display:flex;flex-direction:column;resize:both;overflow:hidden;min-width:240px;min-height:140px}
 /* 右上讓出約 96px 給 Leaflet 圖層控制器 (InSAR_Viewer 把它從 topleft 移到 topright,
    因為左上已被主控制面板佔滿) */
 #tswin{top:122px;right:12px;width:clamp(280px,28vw,430px);height:clamp(160px,30vh,250px)}
 #linewin{top:calc(136px + clamp(160px,30vh,250px));right:12px;width:clamp(230px,20vw,300px);height:auto}
 #linewin .body{gap:8px}
 select,input[type=number]{background:#fff;color:#1e2933;border:1px solid #b6c2cd;border-radius:3px;padding:2px 4px}
 #linewin input[type=file]{color:#34424e;font-size:var(--fs-sm)}
 #linewin label{white-space:nowrap}
 #profwin{bottom:12px;left:calc(24px + clamp(170px,15vw,240px));width:clamp(420px,44vw,640px);height:clamp(190px,32vh,300px);background:rgba(255,255,255,.97);color:#222}
 #profwin .muted{color:#667}
 #profwin .cw{background:#fff;border:1px solid #cfd8dc;border-top:3px solid #43a047;border-radius:3px}
 button:disabled{opacity:.45;cursor:not-allowed}
 .hdr{cursor:move;padding:7px 10px;background:rgba(43,108,176,.75);color:#fff;border-radius:8px 8px 0 0;
   font-weight:600;display:flex;justify-content:space-between;align-items:center;user-select:none}
 .body{flex:1;display:flex;flex-direction:column;padding:7px 9px;gap:5px;min-height:0}
 .cw{flex:1;position:relative;min-height:60px}
 .ctrls{display:flex;align-items:center;gap:7px;flex-wrap:wrap}
 input[type=range]{flex:1;min-width:80px}
 /* 漸層由 applyLayer() 依實際 negLim/posLim 比例動態設定 (與 col() 單一色階來源同步) */
 .cbargrad{height:12px;border-radius:3px;margin:5px 0;background:#888}
 /* 總覽小地圖放大: 點放大鈕切到約視窗 40% 寬高 (還原鈕點回原尺寸) */
 #miniwin.mini-expanded{width:40vw !important}
 #miniwin.mini-expanded #minimap{height:40vh !important}
 .muted{color:#5a6b78}.big{font-size:calc(var(--fs) + 1px);font-weight:600}
 /* GNSS 站: CSS border-trick 三角形(亮黃填充)+drop-shadow 深色描邊, 衛星影像上醒目 */
 .gnss-tri{width:0;height:0;border-left:7px solid transparent;border-right:7px solid transparent;
   border-bottom:12px solid #ffd21e;filter:drop-shadow(0 0 1.5px #000) drop-shadow(0 0 1px #000)}
 .gnss-name{background:rgba(255,255,255,.78);border:0;border-radius:3px;box-shadow:none;
   color:#111;font-size:10px;font-weight:600;padding:0 3px}
 .gnss-name::before{display:none}
 .ctr-lab{background:none !important;border:none !important;box-shadow:none !important}
 .ctr-lab span{font:10px/1 sans-serif;color:var(--ctrcol,#222);white-space:nowrap;
  text-shadow:-1px 0 #fff,1px 0 #fff,0 -1px #fff,0 1px #fff}
 /* 畫線模式: marker/tooltip/popup 與互動向量(含 PS 點 svg path)全部穿透點擊, 全圖維持十字游標 */
 .drawing .leaflet-marker-pane,.drawing .leaflet-tooltip-pane,.drawing .leaflet-popup-pane,
 .drawing .leaflet-interactive{pointer-events:none !important}
 .drawing,.drawing *{cursor:crosshair !important}
 button{background:#2b6cb0;color:#fff;border:0;border-radius:4px;padding:3px 9px;cursor:pointer;font-size:var(--fs)}
 .sb{background:#3a4553}
 /* 小螢幕自適應: 縮字級/padding, 面板寬高上限見 .panel/.win/#tswin/#profwin 等的 min()/vw/vh */
 @media (max-width:1440px),(max-height:820px){
  :root{--fs:11px;--fs-sm:10px}
  .hdr{padding:5px 8px}
  .body{padding:5px 7px;gap:4px}
  #ctrl{padding:5px 8px}
 }
 @media (max-width:1152px),(max-height:720px){
  :root{--fs:10px;--fs-sm:9px}
  .hdr{padding:4px 6px}
  .body{padding:4px 6px;gap:3px}
  #ctrl{padding:4px 6px}
 }
</style></head><body>
<div id="map"></div>
<button id="langToggle" title="Switch language / 切換語言"
 style="position:absolute;top:10px;left:50%;transform:translateX(-50%);z-index:1200;padding:3px 9px">中</button>
<div id="ctrl" class="panel"><span class="big" id="titleTxt"><!--__TITLE__--></span> <span class="muted" id="meta"></span><span class="muted" id="gnssCorrBadge"></span>
 <div class="ctrls" style="margin-top:6px">
  <label class="muted" data-i18n="ctrl_layer">圖層</label>
  <select id="layerMode"><option value="vel">速度場 (mm/yr)</option><option value="cum">總累積變形量 (mm)</option></select>
  <button id="exportTif" disabled data-i18n="btn_export_tif" data-i18n-title="title_export_tif">匯出數值 GeoTIFF</button>
  <button id="exportViewTif" data-i18n="btn_export_view_tif" data-i18n-title="title_export_view_tif">匯出畫面 GeoTIFF</button>
 </div>
 <div class="ctrls" style="margin-top:4px">
  <label class="muted" data-i18n="aoi_label">AOI 匯出</label>
  <button id="aoiDrawBtn" data-i18n="aoi_draw_btn">▭ 框選 AOI</button>
  <button class="sb" id="aoiClearBtn" data-i18n="aoi_clear_btn">清除 AOI</button>
  <button id="aoiPngBtn" data-i18n="aoi_export_png">⬇ AOI PNG</button>
  <button id="aoiTifBtn" data-i18n="aoi_export_tif">⬇ AOI GeoTIFF</button>
 </div>
 <div class="muted" id="aoiStatus" style="min-height:14px"></div>
 <div class="ctrls" style="margin-top:4px">
  <label class="muted" data-i18n="gnss_dir_label">GNSS 時序資料夾</label>
  <input type="file" id="gnssTsDir" webkitdirectory multiple style="flex:1;min-width:0">
 </div>
 <div class="ctrls" style="margin-top:2px">
  <label class="muted" data-i18n="gnss_crs_label">座標系</label>
  <select id="gnssCrs" style="flex:none">
   <option value="3826" data-i18n="gnss_crs_twd97">TWD97 TM2 (EPSG:3826)</option>
   <option value="4326" data-i18n="gnss_crs_lonlat">經緯度 (EPSG:4326)</option>
  </select>
  <label class="muted" data-i18n="gnss_unit_label">單位</label>
  <select id="gnssUnit" style="flex:none"><option value="m">m</option><option value="mm">mm</option></select>
  <label class="muted" data-i18n="gnss_fmt_label">欄位</label>
  <input type="text" id="gnssFmt" value="date,doy,year,n,e,u" style="width:150px;flex:none"
   data-i18n-title="gnss_fmt_label">
  <button id="gnssDirLoad" data-i18n="gnss_dir_load">讀取</button>
 </div>
 <div class="muted" id="gnssTsStatus" style="min-height:14px"></div>
 <div class="ctrls" style="margin-top:4px">
  <label class="muted" data-i18n="ctrl_opacity">透明度</label>
  <input type="range" id="rasterOpacity" min="0" max="100" value="60" style="width:100px">
  <span class="muted" id="rasterOpacityVal">60%</span>
 </div>
 <div class="ctrls" style="margin-top:4px">
  <label class="muted"><input type="checkbox" id="contourEnable"> <span data-i18n="ctrl_contour_enable">等值線</span></label>
  <label class="muted" data-i18n="ctrl_contour_interval">間距</label>
  <input type="number" id="contourInt" min="0" step="0.1" style="width:64px;flex:none" disabled data-i18n-title="title_contour_interval">
  <span class="muted" id="contourIntUnit">mm/yr</span>
  <label class="muted" data-i18n="ctrl_contour_color">顏色</label>
  <input type="color" id="contourColor" value="#222222" style="width:28px;height:18px;padding:0;border:1px solid #999;vertical-align:middle;cursor:pointer">
  <label class="muted"><input type="checkbox" id="contourLabels" checked> <span data-i18n="ctrl_contour_labels">數值</span></label>
  <span class="muted" id="ctrStatus"></span>
 </div>
 <div class="ctrls" style="margin-top:4px">
  <label class="muted" data-i18n="ctrl_threshold_zone">閾值區域: 值 &lt;</label>
  <input type="number" id="thVal" step="0.1" style="width:60px;flex:none" data-i18n-title="title_threshold_val" title="門檻值 (單位隨目前圖層)">
  <span class="muted" id="thUnit"></span>
  <label class="muted" data-i18n="ctrl_border_color">框線</label>
  <input type="color" id="thColor" value="#ff00ff" style="width:28px;height:18px;padding:0;border:1px solid #999;vertical-align:middle;cursor:pointer">
  <button id="thApply" data-i18n="btn_apply">套用</button>
  <button id="thClear" class="sb" data-i18n="btn_clear">清除</button>
 </div>
 <div class="muted" id="thArea" style="min-height:14px"></div>
 <div class="ctrls" style="margin-top:4px">
  <label class="muted" data-i18n="ctrl_load_gnss_csv">載入 GNSS CSV</label>
  <input type="file" id="gnssCsvFile" accept=".csv,.txt" style="flex:1;min-width:0">
  <button id="exportGnssHtml" disabled data-i18n="btn_export_gnss_html" data-i18n-title="title_export_gnss_html">⬇ 匯出更新 HTML</button>
 </div>
 <div class="muted" id="gnssLoadStatus" style="min-height:14px"></div>
 <div class="ctrls" style="margin-top:4px">
  <label class="muted" data-i18n="ctrl_load_h5">載入 HDF5</label>
  <input type="file" id="h5File" accept=".h5,.hdf5" style="flex:1;min-width:0">
  <input type="text" id="h5Title" data-i18n-placeholder="ph_h5_title" placeholder="新標題 (選填)" style="width:104px">
  <button id="h5Load" data-i18n="ctrl_load_h5" data-i18n-title="title_h5_load" title="用瀏覽器直讀 .h5 時序檔 (make_pt_ts_h5.py 格式) 並就地更新整個檢視器, 不需重跑 Python">載入 HDF5</button>
 </div>
 <div class="muted" id="h5Status" style="min-height:14px"></div>
 </div>
<div id="cbar" class="win panel"><div class="hdr" id="cbarhdr"><span id="cbarTitle" data-i18n="cbar_title_vel">速度</span></div>
 <div class="body">
  <div class="ctrls" style="margin-bottom:2px">
   <select id="cbarLayerSel" style="font-size:var(--fs-sm)" data-i18n-title="cbar_switch_title" title="切換圖層 (與左上「圖層」下拉同步)">
    <option value="vel" data-i18n="cbar_opt_vel">速度場</option><option value="cum" data-i18n="cbar_opt_cum">總累積</option>
   </select>
  </div>
  <div class="cbargrad" id="cbarGrad"></div>
  <div id="cbarTicks" style="position:relative;height:13px;font-size:var(--fs-sm)"></div>
  <div class="ctrls" style="font-size:var(--fs-sm)">
   <label class="muted" data-i18n="btn_min">最小</label><input type="number" id="cbMin" step="0.1" style="width:56px;flex:none" disabled>
   <label class="muted" data-i18n="btn_max">最大</label><input type="number" id="cbMax" step="0.1" style="width:56px;flex:none" disabled>
   <button id="cbApply" style="padding:1px 7px" disabled data-i18n="btn_apply">套用</button>
   <button id="cbAuto" class="sb" style="padding:1px 7px" disabled data-i18n="btn_auto">自動</button>
  </div>
  <div class="ctrls" style="justify-content:center"><span class="muted" id="cbarDesc" data-i18n="cbar_desc_vel">綠=穩定 黃→紅→粉=下沉 青→藍=抬升</span></div>
 </div></div>

<div id="tswin" class="win panel"><div class="hdr" id="tshdr"><span id="tstitle" data-i18n="ts_default_title">點地圖看座標點時間序列</span></div>
 <div class="body"><div class="cw"><canvas id="tschart"></canvas></div>
  <div id="tsinfo" style="font-size:var(--fs);color:#000;min-height:15px"></div>
  <div class="ctrls"><button id="tsPngExport" data-i18n="ts_png_btn">⬇ 時序 PNG</button>
   <label class="muted" data-i18n="rate_png_width_label">圖寬</label>
   <input type="number" id="tsPngW" value="1200" min="400" max="4000" step="100" style="width:60px"></div>
 </div></div>

<div id="linewin" class="win panel"><div class="hdr" id="linehdr"><span data-i18n="line_title">互動剖面</span></div>
 <div class="body">
  <div class="ctrls"><button id="drawBtn" data-i18n="line_draw_btn">✎ 手動畫線</button><button class="sb" id="clearLineBtn" data-i18n="btn_clear">清除</button>
   <label class="muted" data-i18n="line_color_label">線色</label>
   <input type="color" id="profLineColor" value="#ffcc00" data-i18n-title="title_line_color"
    style="width:28px;height:18px;padding:0;border:1px solid #999;vertical-align:middle;cursor:pointer"></div>
  <div class="ctrls"><label class="muted" data-i18n="line_import_label">匯入線檔</label>
   <input type="file" id="lineFile" accept=".geojson,.json,.shp" style="flex:1;min-width:0"></div>
  <div class="ctrls"><label class="muted" data-i18n="mask_label">遮罩</label>
   <input type="file" id="maskFile" accept=".geojson,.json" style="flex:1;min-width:0" disabled>
   <button id="drawMaskBtn" disabled data-i18n="mask_draw_btn">✎ 畫遮罩</button>
   <button class="sb" id="maskClear" disabled data-i18n="mask_clear_btn">清除遮罩</button></div>
  <div class="ctrls"><label class="muted" data-i18n="sample_step_label">取樣間距(m)</label>
   <input type="number" id="sampleStep" value="50" min="1" step="1" style="width:64px">
   <label class="muted" data-i18n="curve_label">曲線</label>
   <select id="curveMode"><option value="rate">變形速率 (mm/yr)</option><option value="cum">總累積變形量 (mm)</option></select></div>
  <div class="muted" id="lineStatus" style="min-height:14px"></div>
 </div></div>

<div id="profwin" class="win panel"><div class="hdr" id="profhdr">
 <span><span data-i18n="prof_title">沿線位移剖面</span> <span class="muted" id="dlabel"></span>
  <span class="muted" id="profcur" style="margin-left:8px"></span></span></div>
 <div class="body">
  <div class="ctrls"><button id="play" data-i18n="prof_play">▶ 播放</button>
   <input type="range" id="slider" min="0" value="0"><span class="muted" id="dval"></span></div>
  <div class="cw"><canvas id="profchart"></canvas></div>
  <div class="ctrls"><button class="sb" id="gifRangeBtn" data-i18n="gif_range_btn" data-i18n-title="title_gif_range">▭ 框選範圍</button>
   <button class="sb" id="gifFullBtn" data-i18n="gif_full_btn" data-i18n-title="title_gif_full">全段</button>
   <label class="muted">fps</label>
   <input type="number" id="gifFps" value="4" min="1" max="20" step="1" style="width:44px">
   <button id="gifExport" data-i18n="gif_export_btn" data-i18n-title="title_gif_export">⬇ 輸出 GIF</button></div>
  <div class="ctrls"><button id="ratePngExport" data-i18n="rate_png_btn" data-i18n-title="title_rate_png">⬇ 速率剖面圖 PNG</button>
   <label class="muted" data-i18n="rate_png_width_label">圖寬</label>
   <input type="number" id="ratePngW" value="1400" min="400" max="4000" step="100" style="width:60px"></div>
  <div class="muted" id="gifStatus" style="min-height:14px"></div>
  <div class="ctrls muted"><span data-i18n="prof_basemap_rotation">底圖旋轉</span><input type="range" id="brg" min="0" max="359" value="0">
   <button class="sb" id="align" data-i18n="prof_align">對齊廊道</button><button class="sb" id="north" data-i18n="prof_north">北↑</button></div>
 </div></div>
<div id="miniwin" class="panel" style="right:12px;bottom:12px;width:clamp(140px,12vw,180px);overflow:hidden;z-index:1100">
 <div id="minihdr" style="display:flex;justify-content:space-between;align-items:center;padding:3px 8px;cursor:move">
  <span class="muted" style="font-size:var(--fs-sm)" data-i18n="mini_overview">總覽</span>
  <span>
   <button id="miniExpand" class="sb" style="padding:0 6px;font-size:var(--fs-sm);line-height:16px" data-i18n-title="mini_expand_title" title="放大/還原總覽">⤢</button>
   <button id="miniToggle" class="sb" style="padding:0 6px;font-size:var(--fs-sm);line-height:16px">—</button>
  </span>
 </div>
 <div id="minimap" style="width:100%;height:clamp(140px,22vh,200px)"></div>
</div>
<script>
// ===== i18n bootstrap (top-level so it's shared across all <script> tags incl. --app's
// startViewer closure and _APP_JS; window.* so it survives being re-declared if this
// template is ever concatenated twice; DOM apply is deferred, see below) =====
window.I18N = {
"zh":{
"ctrl_layer":"圖層",
"ctrl_opacity":"透明度",
"ctrl_contour_enable":"等值線",
"ctrl_contour_interval":"間距",
"title_contour_interval":"等值線間距 (mm; 速度場圖層為 mm/yr), 任意正數",
"ctrl_contour_color":"顏色",
"ctrl_contour_labels":"數值",
"ctrl_threshold_zone":"閾值區域: 值 <",
"ctrl_border_color":"框線",
"ctrl_load_gnss_csv":"載入 GNSS CSV",
"ctrl_load_h5":"載入 HDF5",
"btn_apply":"套用",
"btn_clear":"清除",
"btn_auto":"自動",
"btn_min":"最小",
"btn_max":"最大",
"btn_export_tif":"匯出數值 GeoTIFF",
"btn_export_view_tif":"匯出畫面 GeoTIFF",
"btn_export_gnss_html":"⬇ 匯出更新 HTML",
"title_export_tif":"格網載入後可匯出",
"title_export_view_tif":"把目前視野合成 RGB GeoTIFF (EPSG:3857), 可疊進 GIS",
"title_threshold_val":"門檻值 (單位隨目前圖層)",
"title_export_gnss_html":"載入 GNSS CSV 後可匯出帶新測站的 HTML",
"title_h5_load":"用瀏覽器直讀 .h5 時序檔 (make_pt_ts_h5.py 格式) 並就地更新整個檢視器, 不需重跑 Python",
"ph_h5_title":"新標題 (選填)",
"cbar_switch_title":"切換圖層 (與左上「圖層」下拉同步)",
"cbar_opt_vel":"速度場",
"cbar_opt_cum":"總累積",
"cbar_desc_vel":"綠=穩定 黃→紅→粉=下沉 青→藍=抬升",
"cbar_desc_cum":"綠=穩定 黃→紅→粉=累積下沉 青→藍=累積抬升",
"cbar_title_vel":"速度 {U}/yr",
"cbar_title_cum":"總累積變形量 {U}",
"ts_default_title":"點地圖看座標點時間序列",
"ts_coord":"座標 ({lon}, {lat})",
"ts_ps_title":"PS #{i}",
"ts_coh":"  coh={c}",
"ts_along_line":"  沿線 {km} km",
"common_subsidence":"下沉",
"common_uplift":"抬升",
"ts_velocity":"速度 {sign}{v} {U}/yr ({dir})",
"ts_cum_prefix":"總累積變形量 ",
"ts_reg_vel":"回歸速度 {sign}{v} {U}/yr",
"ts_reg_vel_na":"回歸速度 —",
"ts_cum_val":"總累積變形量 {sign}{v} {U} ({dir})",
"ts_cum_na":"總累積變形量 —",
"chart_ts_noise":"時序(雜訊)",
"chart_vel_trend":"速度趨勢",
"chart_ts_grid":"時序(格網取樣)",
"chart_reg_trend":"回歸趨勢",
"chart_current_epoch":"當期",
"axis_displacement":"位移 {U}",
"axis_along_dist":"沿線距離 km",
"prof_title":"沿線位移剖面",
"prof_play":"▶ 播放",
"prof_pause":"⏸ 暫停",
"prof_basemap_rotation":"底圖旋轉",
"prof_align":"對齊廊道",
"prof_north":"北↑",
"prof_mileage":"里程 {km} km",
"prof_start":"起點 0 km",
"prof_end":"終點 {km} km",
"line_title":"互動剖面",
"line_draw_btn":"✎ 手動畫線",
"line_drawing_btn":"✎ 畫線中(雙擊結束)",
"line_import_label":"匯入線檔",
"line_color_label":"線色",
"title_line_color":"剖面線顏色 (中線/互動線/畫線中暫時線同步變色)",
"gif_range_btn":"▭ 框選範圍",
"gif_range_btn_active":"▭ 拖曳中(放開結束)",
"title_gif_range":"在地圖上拖曳出矩形, 只輸出框內那一段剖面",
"gif_full_btn":"全段",
"title_gif_full":"取消框選, 輸出整條剖面",
"gif_export_btn":"⬇ 輸出 GIF",
"title_gif_export":"把第一期到最後一期的剖面播放輸出成 GIF 動畫檔",
"gif_msg_drag":"在地圖上按住拖曳出矩形",
"gif_msg_cancelled":"已取消框選",
"gif_msg_full":"輸出範圍: 整條剖面",
"gif_msg_range":"輸出範圍: {a}–{b} km ({n} 取樣點)",
"gif_msg_noline":"目前沒有剖面線可框選",
"gif_msg_toofew":"框內剖面線點數不足, 已改為整條剖面",
"gif_msg_nochart":"目前沒有剖面圖",
"gif_msg_rotate":"請先將地圖轉回正北再框選",
"gif_msg_aborted":"剖面圖已變更, 已中止輸出",
"gif_msg_shrunk":"期數多, 幀寬自動縮到 {w}px 以控制記憶體",
"rate_png_btn":"⬇ 速率剖面圖 PNG",
"title_rate_png":"把目前剖面線的沿線變化速率輸出成一張 PNG (QGIS 風格: 單條綠線+方形標記)",
"rate_png_width_label":"圖寬",
"rate_axis_y":"變化速率 ({U}/yr)",
"rate_axis_x":"距離 (公尺)",
"rate_msg_noline":"請先畫一條剖面線",
"rate_msg_nodata":"剖面線上沒有可用的速率資料",
"rate_msg_done":"已輸出速率剖面圖 {w}×{h}px ({n} 點)",
"gif_msg_capture":"擷取畫面 {i}/{n}",
"gif_msg_encoding":"編碼 GIF 中…",
"gif_msg_done":"已輸出 {n} 幀 {w}×{h}px, {mb} MB",
"gif_msg_fail":"輸出失敗: {err}",
"mask_label":"遮罩",
"mask_draw_btn":"✎ 畫遮罩",
"mask_drawing_btn":"✎ 畫遮罩中(雙擊閉合)",
"mask_clear_btn":"清除遮罩",
"sample_step_label":"取樣間距(m)",
"curve_label":"曲線",
"curve_opt_rate":"變形速率 ({U}/yr)",
"curve_opt_cum":"總累積變形量 {U}",
"opt_vel_label":"速度場 ({U}/yr)",
"opt_cum_label":"總累積變形量 ({U})",
"mini_overview":"總覽",
"mini_expand_title":"放大/還原總覽",
"msg_computing":"計算中…",
"msg_reading_file":"讀取檔案中…",
"msg_not_enough_line_pts":"線段點數不足",
"msg_sampling":"取樣中…",
"msg_sampled_n":"已取樣 {n} 點",
"msg_click_add_point":"地圖上單擊加點, 雙擊結束",
"msg_click_add_vertex":"地圖上單擊加頂點, 雙擊閉合多邊形 (ESC 取消)",
"msg_mask_not_enough_vertex":"遮罩頂點不足 (至少 3 點), 未套用",
"msg_mask_failed":"遮罩失敗: ",
"msg_mask_applied":"遮罩已套用 (內含有效像元 {n})",
"msg_mask_cleared":"已清除遮罩",
"err_no_polygon":"找不到 Polygon/MultiPolygon",
"msg_import_failed":"匯入失敗: ",
"msg_line_import_failed":"線檔匯入失敗: ",
"err_no_linestring":"找不到有效的 LineString/MultiLineString",
"err_unsupported_shp":"不支援的 shape type: {t} (僅支援 3/13/23 PolyLine)",
"err_shp_no_coords":"shp 無有效座標",
"msg_no_grid_disabled":"此檔未內嵌全解析度格網(以 --no-grid 產生, 或資料過大自動略過)",
"msg_no_decompression":"瀏覽器不支援 DecompressionStream, 互動剖面停用",
"alert_no_decompression":"此瀏覽器不支援 DecompressionStream, 互動剖面功能已停用(建議使用新版 Chrome/Edge/Firefox)",
"msg_decompressing":"格網解壓中…",
"msg_decompress_failed":"格網解壓失敗: ",
"msg_no_grid_short":"此檔未內嵌全解析度格網",
"title_export_tif_ok":"把目前顯示圖層輸出成 GeoTIFF (EPSG:4326, float32, NaN=無值)",
"title_export_tif_bad":"需要內嵌格網才能匯出 (--no-grid 版不支援): ",
"title_needs_grid":"需要內嵌格網 (--no-grid 版不支援)",
"err_grid_not_ready":"GRID 未就緒",
"alert_export_tif_failed":"匯出 GeoTIFF 失敗: ",
"err_rotate_north_first":"請先將地圖轉回正北再匯出",
"alert_basemap_capture_failed":"底圖無法擷取 (CORS), 僅輸出資料圖層",
"alert_export_view_failed":"匯出畫面失敗: ",
"thr_invalid":"請輸入有效數值",
"thr_no_grid":"此檔未內嵌全解析度格網, 無法計算閾值面積",
"thr_result":"< {thresh} {unit} 面積: {area} km² ({count} 像元)",
"thr_failed":"閾值計算失敗: ",
"lyr_contour":"等值線",
"lyr_raster":"變形場 (raster)",
"lyr_gnss":"GNSS 測站",
"lyr_threshold":"閾值區域邊界",
"ctr_computing":"等值線計算中…",
"ctr_failed":"等值線計算失敗",
"ctr_invalid_interval":"請輸入正數間距 (mm)",
"ctr_no_grid_interval":"此間距非內嵌等值線的整數倍, 且本頁未內嵌格網可動態計算",
"gnss_popup_title":"GNSS 測站 {n}",
"gnss_popup_lon":"經度 {v}",
"gnss_popup_lat":"緯度 {v}",
"gnss_no_stations_skipped":"未解析出任何測站 (略過 {n} 列)",
"gnss_loaded_n":"已載入 {n} 站",
"gnss_skipped_suffix":" (略過 {n} 列)",
"gnss_csv_parse_failed":"CSV 解析失敗: ",
"gnss_read_failed":"讀取檔案失敗",
"err_gnss_marker_missing":"原始頁面找不到 GNSS 標記, 無法匯出 (此頁可能非本工具產生)",
"gnss_export_suffix_filename":"_gnss更新.html",
"gnss_corr_none":"範圍內沒有 GNSS 測站，未做 GNSS 校正",
"gnss_corr_geom_fail":"LOS 幾何取得失敗，無法進行 GNSS 校正",
"gnss_corr_badge":"GNSS 校正：{n} 站（{model}）",
"gnss_corr_rms_warn":"⚠ 校正後 RMS={rms} mm/yr 偏高",
"gnss_corr_model_const":"常數",
"gnss_corr_model_plane":"一次平面",
"gnss_corr_stale":"已載入新資料，未經 GNSS 校正",
"alert_export_failed":"匯出失敗: ",
"title_gnss_export_stale":"已載入其他 HDF5 資料, 此匯出僅適用本頁原始資料 (請重跑 Python 產生新 HTML)",
"h5_not_loaded":"h5wasm 未載入 (離線版未內嵌; 線上版請確認可連 CDN)",
"h5_initializing":"h5wasm 初始化中…",
"h5_open_failed":"無法以 HDF5 開啟 {name} (格式不符或檔案損毀)",
"h5_missing_datasets":"檔案缺少 /date 或 /timeseries (需 make_pt_ts_h5.py 產生的格式)",
"h5_missing_geotransform":"檔案缺少 GEOTRANSFORM 屬性 (無地理參考, 無法定位)",
"h5_bad_geotransform":"GEOTRANSFORM 需為 6 個有限數值且像素間距不得為 0",
"h5_bad_shape":"timeseries 維度需為 (期,列,行), 實際為 ({shp})",
"h5_too_few_epochs":"期數 {n} 少於 2 期, 無法計算速度",
"h5_date_mismatch":"date 期數 {a} 與 timeseries 期數 {b} 不一致",
"h5_bad_dates":"date 需為 YYYYMMDD 且時間需遞增",
"h5_bad_unit":"不支援的 UNIT=\"{u}\" (需 mm 或 cm)",
"h5_parsing":"解析中… {n} 期 {h}x{w}",
"h5_size_mismatch":"timeseries 資料量 {a} 與宣告維度 {b} 不符",
"h5_computing_vel":"回歸速度計算中…",
"h5_no_valid_px":"檔案內沒有任何有效像元 (整份都是 NaN?)",
"h5_quantizing":"格網量化中…",
"h5_redrawing":"重繪中…",
"h5_select_first":"請先選擇 .h5 檔案",
"h5_load_failed":"載入失敗: ",
"alert_h5_load_failed":"載入 HDF5 失敗: ",
"h5_loaded_summary":"已載入 {fname}: {n} 期 {h}x{w}, {pts} 點, ",
"h5_loaded_scale":"色階 -{neg}/+{pos} {U}/yr",
"meta_summary":"{n} 點 · {nd} 期 · {d0}~{d1}",
"meta_corridor":" · 廊道±{buf}m",
"base_satellite":"Esri 衛星影像",
"ui_title_suffix":"時序檢視器",
"app_title":"InSAR Viewer — 時序檢視器",
"app_desc":"選擇一種資料來源。讀檔、內插、速度回歸與繪圖<b>全部在瀏覽器內完成</b>, 資料不會上傳。",
"app_preparing":"色階/等值線/渲染準備中…",
"err_no_valid_px_after":"鑲嵌/投影後無有效像元 (檢查 grd 座標系與涵蓋範圍)",
"app_rendering":"渲染中…",
"app_error_prefix":"錯誤: ",
"err_calc_not_done":"尚未完成計算",
"btn_exporting":"匯出中…",
"btn_export_share":"⬇ 匯出分享版 HTML",
"title_export_share":"把計算結果烘進單一 HTML, 分享對象開啟不需 grd 檔",
"src_h5_title":"① 時序 HDF5 (MintPy / dolphin / gmtsar2h5)",
"src_coh_label":"同調性遮罩檔 (選填, 兩種來源皆適用)",
"src_coh_min":"門檻",
"src_h5_hint":"支援 MintPy 屬性 (X_FIRST/Y_FIRST/X_STEP/Y_STEP, 單位 m) 與 gmtsar2h5 的 GEOTRANSFORM (單位 mm/cm)。",
"src_gmt_title":"② GMTSAR 時序資料夾",
"src_gmt_step":"格網間距 (m)",
"src_gmt_buffer":"資料緩衝 (m)",
"src_gmt_buffer_title":"距離最近 PS 點超過此距離的網格留白 (不外插)。GMTSAR 沒有同調性遮罩, 有效範圍由 PS 點分布決定; 等同 gmtsar2h5.py 的 --max-ps-dist-km",
"src_gmt_hint":"需含 disp_NNN_ll.xy (lon lat 位移mm) 與 data_date.txt。散點以格網平均聚合 (非克利金), 屬快視結果; 有效範圍由 PS 點分布＋資料緩衝決定, 不需外部遮罩。",
"src_opt_title":"共同選項",
"src_vel_label":"速度場校正檔 (選填)",
"src_vel_title":"用外部速度場逐格改寫時序趨勢 (D += (v_ref − v_ols)·t), 校正後每格的回歸速度就等於參考速度。支援 GeoJSON 點檔與二維 HDF5 速度場; 取不到參考值的網格保持原值",
"src_vel_field":"屬性欄",
"src_vel_deramp":"坡面扣除 C0,C1,C2,LON0,LAT0",
"src_vel_deramp_title":"先從參考速度扣掉平面 C0+C1·(lon−LON0)+C2·(lat−LAT0) 再拿來校正 (等同 gmtsar2h5.py 的 --deramp), 留空則不扣",
"vel_loading":"讀取速度場校正檔…",
"vel_retrending":"逐格改寫時序趨勢 (參考點 {n})…",
"vel_note":"速度場校正: 參考 {n} 點, 已校正 {fix} 格, 未涵蓋 {skip} 格",
"vel_note_deramp":" (參考速度已扣坡面)",
"vel_no_features":"速度場檔沒有 features",
"vel_no_field":"點檔內找不到數值欄「{f}」",
"vel_bad_shape":"速度場資料集不是二維 ({a})",
"vel_no_geo":"速度場 HDF5 缺地理屬性 (X_FIRST/Y_STEP 或 GEOTRANSFORM)",
"vel_bad_deramp":"坡面係數格式須為 C0,C1,C2,LON0,LAT0 五個數字",
"src_opt_deramp":"去除線性坡面 (deramp)",
"src_opt_unit":"顯示單位",
"src_opt_vlim":"色階上限 (0=自動)",
"src_opt_vlim_title":"速度場與累積變形的色階對稱上限 (顯示單位)。0=依資料自動取 p0.5/p99.5——資料若「大部分穩定、少數極端」, 自動色階會被壓得很窄, 少數極端值全部飽和。要與官方 1 cm/yr 分級圖對齊請填 80 (mm) 或 8 (cm)",
"lims_auto":"色階 (自動): {a} ~ {b} {U}／yr — 極端值會飽和; 要對齊官方分級請填色階上限或在左下色階視窗改",
"lims_fixed":"色階 (指定): {a} ~ {b} {U}／yr",
"src_run":"載入並繪圖",
"src_gnss_note":"GNSS 時序資料夾在載入後於左上面板選取",
"src_pick_first":"請先選擇時序 HDF5 或 GMTSAR 資料夾",
"coh_applying":"套用同調性遮罩 (門檻 {thr})…",
"coh_note":"同調性遮罩: 濾除 {n} 個像元 (門檻 {thr})",
"coh_no_dataset":"遮罩檔內找不到可用的資料集",
"coh_shape_mismatch":"遮罩尺寸 {a} 與時序 {b} 不符, 且遮罩檔沒有地理屬性可供重取樣",
"coh_bad_shape":"遮罩資料集不是二維 ({a})",
"coh_resampled":" (依經緯度重取樣, {n} 格落在遮罩範圍外)",
"gmt_no_disp":"資料夾內找不到 disp_NNN_ll.xy",
"gmt_no_datefile":"資料夾內找不到 data_date.txt",
"gmt_date_count_mismatch":"data_date.txt 有 {a} 列, 但 disp 檔有 {b} 個",
"gmt_reading":"讀取散點… {i}/{n} ({name})",
"gmt_gridding":"格網聚合… {i}/{n}",
"gmt_no_points":"沒有可用的散點資料",
"gmt_grid_too_big":"格網過大 ({w}×{h}), 請調大格網間距",
"gmt_note":"GMTSAR 快視模式: {step} m 格網平均, 資料緩衝 {buf} m ({rad} 格), 非克利金",
"deramp_running":"擬合並去除線性坡面…",
"deramp_note":"已對 {n} 期去除線性坡面 (deramp)",
"aoi_draw_btn":"▭ 框選 AOI",
"aoi_clear_btn":"清除 AOI",
"aoi_export_png":"⬇ AOI PNG",
"aoi_export_tif":"⬇ AOI GeoTIFF",
"aoi_hint":"在地圖上拖曳出矩形範圍",
"aoi_none":"尚未框選 AOI",
"aoi_set":"AOI: {w}×{h} 像元",
"aoi_empty":"AOI 範圍內沒有有效資料",
"aoi_label":"AOI 匯出",
"ts_png_btn":"⬇ 時序 PNG",
"ts_png_none":"請先點地圖選一個位置",
"gnss_dir_label":"GNSS 時序資料夾",
"gnss_dir_load":"讀取",
"err_share_anchor":"分享版 HTML 找不到插入位置",
"gnss_crs_label":"座標系",
"gnss_crs_twd97":"TWD97 TM2 (EPSG:3826)",
"gnss_crs_lonlat":"經緯度 (EPSG:4326)",
"gnss_unit_label":"單位",
"gnss_fmt_label":"欄位",
"gnss_dir_none":"資料夾內找不到 <站碼>_f_all.xlsx 或 .csv/.txt",
"gnss_dir_loaded":"已讀取 {n} 站 GNSS 時序 (點站名看時序)",
"gnss_dir_reading":"讀取 GNSS 時序… {i}/{n} ({name})",
"gnss_sheetjs_missing":"SheetJS 未載入, 無法讀 .xlsx",
"gnss_ts_title":"GNSS {sta} ({comp})",
"gnss_comp_label":"分量",
"src_unused_zh":""
},
"en":{
"ctrl_layer":"Layer",
"ctrl_opacity":"Opacity",
"ctrl_contour_enable":"Contours",
"ctrl_contour_interval":"Interval",
"title_contour_interval":"Contour interval (mm; mm/yr for the velocity layer), any positive number",
"ctrl_contour_color":"Color",
"ctrl_contour_labels":"Labels",
"ctrl_threshold_zone":"Threshold zone: value <",
"ctrl_border_color":"Border color",
"ctrl_load_gnss_csv":"Load GNSS CSV",
"ctrl_load_h5":"Load HDF5",
"btn_apply":"Apply",
"btn_clear":"Clear",
"btn_auto":"Auto",
"btn_min":"Min",
"btn_max":"Max",
"btn_export_tif":"Export value GeoTIFF",
"btn_export_view_tif":"Export view GeoTIFF",
"btn_export_gnss_html":"⬇ Export updated HTML",
"title_export_tif":"Available after the grid loads",
"title_export_view_tif":"Composite the current view as an RGB GeoTIFF (EPSG:3857) for GIS overlay",
"title_threshold_val":"Threshold value (unit follows the current layer)",
"title_export_gnss_html":"Available after loading a GNSS CSV: export HTML with the new stations",
"title_h5_load":"Read a .h5 time-series file (make_pt_ts_h5.py format) directly in the browser and update the whole viewer in place, no Python re-run needed",
"ph_h5_title":"New title (optional)",
"cbar_switch_title":"Switch layer (synced with the Layer dropdown above)",
"cbar_opt_vel":"Velocity field",
"cbar_opt_cum":"Cumulative",
"cbar_desc_vel":"Green=stable Yellow→Red→Pink=subsidence Cyan→Blue=uplift",
"cbar_desc_cum":"Green=stable Yellow→Red→Pink=cumulative subsidence Cyan→Blue=cumulative uplift",
"cbar_title_vel":"Velocity {U}/yr",
"cbar_title_cum":"Cumulative displacement {U}",
"ts_default_title":"Click the map to view a point's time series",
"ts_coord":"Coord ({lon}, {lat})",
"ts_ps_title":"PS #{i}",
"ts_coh":"  coh={c}",
"ts_along_line":"  along-line {km} km",
"common_subsidence":"Subsidence",
"common_uplift":"Uplift",
"ts_velocity":"Velocity {sign}{v} {U}/yr ({dir})",
"ts_cum_prefix":"Cumulative displacement ",
"ts_reg_vel":"Regression velocity {sign}{v} {U}/yr",
"ts_reg_vel_na":"Regression velocity —",
"ts_cum_val":"Cumulative displacement {sign}{v} {U} ({dir})",
"ts_cum_na":"Cumulative displacement —",
"chart_ts_noise":"Time series (raw)",
"chart_vel_trend":"Velocity trend",
"chart_ts_grid":"Time series (grid sample)",
"chart_reg_trend":"Regression trend",
"chart_current_epoch":"Current epoch",
"axis_displacement":"Displacement {U}",
"axis_along_dist":"Along-line distance km",
"prof_title":"Along-line displacement profile",
"prof_play":"▶ Play",
"prof_pause":"⏸ Pause",
"prof_basemap_rotation":"Basemap rotation",
"prof_align":"Align corridor",
"prof_north":"North ↑",
"prof_mileage":"Distance {km} km",
"prof_start":"Start 0 km",
"prof_end":"End {km} km",
"line_title":"Interactive profile",
"line_draw_btn":"✎ Draw line",
"line_drawing_btn":"✎ Drawing (double-click to finish)",
"line_import_label":"Import line file",
"line_color_label":"Line colour",
"title_line_color":"Profile line colour (corridor line, interactive line and the in-progress line all follow)",
"gif_range_btn":"▭ Select range",
"gif_range_btn_active":"▭ Dragging (release to finish)",
"title_gif_range":"Drag a rectangle on the map to export only that part of the profile",
"gif_full_btn":"Whole line",
"title_gif_full":"Clear the selection and export the whole profile",
"gif_export_btn":"⬇ Export GIF",
"title_gif_export":"Export the profile playback (first to last epoch) as an animated GIF",
"gif_msg_drag":"Press and drag on the map to draw a rectangle",
"gif_msg_cancelled":"Selection cancelled",
"gif_msg_full":"Export range: whole profile",
"gif_msg_range":"Export range: {a}–{b} km ({n} samples)",
"gif_msg_noline":"No profile line to select from",
"gif_msg_toofew":"Too few profile points inside the rectangle; using the whole profile",
"gif_msg_nochart":"No profile chart yet",
"gif_msg_rotate":"Please rotate the map back to north before selecting a range",
"gif_msg_aborted":"Profile chart changed; export aborted",
"gif_msg_shrunk":"Many epochs: frame width reduced to {w}px to limit memory",
"rate_png_btn":"⬇ Rate profile PNG",
"title_rate_png":"Export the along-line rate of change as a PNG (QGIS style: single green line with square markers)",
"rate_png_width_label":"Width",
"rate_axis_y":"Rate of change ({U}/yr)",
"rate_axis_x":"Distance (m)",
"rate_msg_noline":"Draw a profile line first",
"rate_msg_nodata":"No usable rate data along the profile line",
"rate_msg_done":"Rate profile exported, {w}×{h}px ({n} points)",
"gif_msg_capture":"Capturing frame {i}/{n}",
"gif_msg_encoding":"Encoding GIF…",
"gif_msg_done":"Exported {n} frames, {w}×{h}px, {mb} MB",
"gif_msg_fail":"Export failed: {err}",
"mask_label":"Mask",
"mask_draw_btn":"✎ Draw mask",
"mask_drawing_btn":"✎ Drawing mask (double-click to close)",
"mask_clear_btn":"Clear mask",
"sample_step_label":"Sample spacing (m)",
"curve_label":"Curve",
"curve_opt_rate":"Deformation rate ({U}/yr)",
"curve_opt_cum":"Cumulative displacement {U}",
"opt_vel_label":"Velocity field ({U}/yr)",
"opt_cum_label":"Cumulative displacement ({U})",
"mini_overview":"Overview",
"mini_expand_title":"Expand/restore overview",
"msg_computing":"Computing…",
"msg_reading_file":"Reading file…",
"msg_not_enough_line_pts":"Not enough points on the line",
"msg_sampling":"Sampling…",
"msg_sampled_n":"Sampled {n} points",
"msg_click_add_point":"Click on the map to add points, double-click to finish",
"msg_click_add_vertex":"Click on the map to add vertices, double-click to close the polygon (ESC to cancel)",
"msg_mask_not_enough_vertex":"Not enough mask vertices (at least 3 needed); not applied",
"msg_mask_failed":"Mask failed: ",
"msg_mask_applied":"Mask applied ({n} valid pixels inside)",
"msg_mask_cleared":"Mask cleared",
"err_no_polygon":"No Polygon/MultiPolygon found",
"msg_import_failed":"Import failed: ",
"msg_line_import_failed":"Line file import failed: ",
"err_no_linestring":"No valid LineString/MultiLineString found",
"err_unsupported_shp":"Unsupported shape type: {t} (only 3/13/23 PolyLine supported)",
"err_shp_no_coords":"shp has no valid coordinates",
"msg_no_grid_disabled":"This file has no embedded full-resolution grid (generated with --no-grid, or auto-skipped due to size)",
"msg_no_decompression":"Browser doesn't support DecompressionStream; interactive profile disabled",
"alert_no_decompression":"This browser doesn't support DecompressionStream; the interactive profile is disabled (use a recent Chrome/Edge/Firefox)",
"msg_decompressing":"Decompressing grid…",
"msg_decompress_failed":"Grid decompression failed: ",
"msg_no_grid_short":"This file has no embedded full-resolution grid",
"title_export_tif_ok":"Export the currently displayed layer as GeoTIFF (EPSG:4326, float32, NaN=no data)",
"title_export_tif_bad":"Requires an embedded grid to export (not supported in --no-grid builds): ",
"title_needs_grid":"Requires an embedded grid (not supported in --no-grid builds)",
"err_grid_not_ready":"GRID not ready",
"alert_export_tif_failed":"Export GeoTIFF failed: ",
"err_rotate_north_first":"Please rotate the map back to north before exporting",
"alert_basemap_capture_failed":"Basemap could not be captured (CORS); exporting the data layer only",
"alert_export_view_failed":"Export view failed: ",
"thr_invalid":"Enter a valid number",
"thr_no_grid":"This file has no embedded full-resolution grid; cannot compute threshold area",
"thr_result":"< {thresh} {unit} area: {area} km² ({count} px)",
"thr_failed":"Threshold computation failed: ",
"lyr_contour":"Contours",
"lyr_raster":"Deformation field (raster)",
"lyr_gnss":"GNSS stations",
"lyr_threshold":"Threshold zone boundary",
"ctr_computing":"Computing contours…",
"ctr_failed":"Contour computation failed",
"ctr_invalid_interval":"Enter a positive interval (mm)",
"ctr_no_grid_interval":"This interval is not a multiple of the embedded contour base, and this page has no embedded grid for dynamic computation",
"gnss_popup_title":"GNSS station {n}",
"gnss_popup_lon":"Longitude {v}",
"gnss_popup_lat":"Latitude {v}",
"gnss_no_stations_skipped":"No stations parsed ({n} rows skipped)",
"gnss_loaded_n":"Loaded {n} stations",
"gnss_skipped_suffix":" ({n} rows skipped)",
"gnss_csv_parse_failed":"CSV parse failed: ",
"gnss_read_failed":"Failed to read file",
"err_gnss_marker_missing":"GNSS marker not found in the source page; cannot export (this page may not have been generated by this tool)",
"gnss_export_suffix_filename":"_gnss_updated.html",
"gnss_corr_none":"No GNSS station within extent; GNSS calibration not applied",
"gnss_corr_geom_fail":"LOS geometry acquisition failed; GNSS calibration not possible",
"gnss_corr_badge":"GNSS calibration: {n} station(s) ({model})",
"gnss_corr_rms_warn":"⚠ post-calibration RMS={rms} mm/yr is high",
"gnss_corr_model_const":"constant",
"gnss_corr_model_plane":"plane",
"gnss_corr_stale":"New data loaded; not GNSS-calibrated",
"alert_export_failed":"Export failed: ",
"title_gnss_export_stale":"Other HDF5 data has been loaded; this export only applies to the page's original data (re-run Python to generate a new HTML)",
"h5_not_loaded":"h5wasm not loaded (not bundled in the offline build; the online build needs CDN access)",
"h5_initializing":"Initializing h5wasm…",
"h5_open_failed":"Could not open {name} as HDF5 (invalid format or corrupted file)",
"h5_missing_datasets":"File is missing /date or /timeseries (requires the make_pt_ts_h5.py format)",
"h5_missing_geotransform":"File is missing the GEOTRANSFORM attribute (no georeference, cannot locate)",
"h5_bad_geotransform":"GEOTRANSFORM must be 6 finite numbers with nonzero pixel spacing",
"h5_bad_shape":"timeseries dimensions must be (epoch,row,col), got ({shp})",
"h5_too_few_epochs":"{n} epochs is fewer than 2; cannot compute velocity",
"h5_date_mismatch":"date has {a} epochs but timeseries has {b}; mismatch",
"h5_bad_dates":"date must be YYYYMMDD and strictly increasing",
"h5_bad_unit":"Unsupported UNIT=\"{u}\" (must be mm or cm)",
"h5_parsing":"Parsing… {n} epochs {h}x{w}",
"h5_size_mismatch":"timeseries data length {a} doesn't match the declared dimensions {b}",
"h5_computing_vel":"Computing regression velocity…",
"h5_no_valid_px":"File has no valid pixels (all NaN?)",
"h5_quantizing":"Quantizing grid…",
"h5_redrawing":"Redrawing…",
"h5_select_first":"Select a .h5 file first",
"h5_load_failed":"Load failed: ",
"alert_h5_load_failed":"Load HDF5 failed: ",
"h5_loaded_summary":"Loaded {fname}: {n} epochs {h}x{w}, {pts} points, ",
"h5_loaded_scale":"scale -{neg}/+{pos} {U}/yr",
"meta_summary":"{n} points · {nd} epochs · {d0}~{d1}",
"meta_corridor":" · corridor ±{buf}m",
"base_satellite":"Esri Satellite Imagery",
"ui_title_suffix":"Viewer",
"app_title":"InSAR Viewer — time-series viewer",
"app_desc":"Pick one data source. Reading, gridding, velocity regression and rendering <b>all happen in your browser</b>; no data is uploaded.",
"app_preparing":"Preparing color scale/contours/render…",
"err_no_valid_px_after":"No valid pixels after mosaicking/projection (check the grd coordinate system and coverage)",
"app_rendering":"Rendering…",
"app_error_prefix":"Error: ",
"err_calc_not_done":"Computation not finished yet",
"btn_exporting":"Exporting…",
"btn_export_share":"⬇ Export share HTML",
"title_export_share":"Bake the computed result into a single HTML; recipients can open it without grd files",
"src_h5_title":"\u2460 Time-series HDF5 (MintPy / dolphin / gmtsar2h5)",
"src_coh_label":"Coherence mask file (optional, applies to both sources)",
"src_coh_min":"Threshold",
"src_h5_hint":"Supports MintPy attributes (X_FIRST/Y_FIRST/X_STEP/Y_STEP, unit m) and the gmtsar2h5 GEOTRANSFORM (unit mm/cm).",
"src_gmt_title":"\u2461 GMTSAR time-series folder",
"src_gmt_step":"Grid spacing (m)",
"src_gmt_buffer":"Data buffer (m)",
"src_gmt_buffer_title":"Cells farther than this from the nearest PS point are left blank (no extrapolation). GMTSAR has no coherence mask, so the valid extent comes from the PS distribution itself — same idea as --max-ps-dist-km in gmtsar2h5.py",
"src_gmt_hint":"Needs disp_NNN_ll.xy (lon lat displacement in mm) and data_date.txt. Scattered points are averaged per cell (not kriging) — a quick-look result; the valid extent comes from the PS distribution plus the data buffer, no external mask needed.",
"src_opt_title":"Common options",
"src_vel_label":"Reference velocity file (optional)",
"src_vel_title":"Re-trend each cell to an external velocity field (D += (v_ref − v_ols)·t); after this the regression velocity of every cell equals the reference. Accepts a GeoJSON point file or a 2-D HDF5 velocity grid; cells with no reference value keep their original series",
"src_vel_field":"Field",
"src_vel_deramp":"Ramp removal C0,C1,C2,LON0,LAT0",
"src_vel_deramp_title":"Subtract the plane C0+C1·(lon−LON0)+C2·(lat−LAT0) from the reference velocity before using it (same as --deramp in gmtsar2h5.py); leave empty to skip",
"vel_loading":"Reading the reference velocity file…",
"vel_retrending":"Re-trending cells ({n} reference points)…",
"vel_note":"Velocity calibration: {n} reference points, {fix} cells re-trended, {skip} cells not covered",
"vel_note_deramp":" (ramp removed from the reference)",
"vel_no_features":"The velocity file has no features",
"vel_no_field":"Value field \"{f}\" not found in the point file",
"vel_bad_shape":"Velocity dataset is not 2-D ({a})",
"vel_no_geo":"The velocity HDF5 has no geo attributes (X_FIRST/Y_STEP or GEOTRANSFORM)",
"vel_bad_deramp":"Ramp coefficients must be five numbers: C0,C1,C2,LON0,LAT0",
"src_opt_deramp":"Remove linear ramp (deramp)",
"src_opt_unit":"Display unit",
"src_opt_vlim":"Color limit (0 = auto)",
"src_opt_vlim_title":"Symmetric color limit for the velocity and cumulative layers (display unit). 0 = auto from the data (p0.5/p99.5) — for data that is mostly stable with a few extremes, the auto range gets squeezed and the extremes all saturate. Use 80 (mm) or 8 (cm) to match a 1 cm/yr classified legend",
"lims_auto":"Color scale (auto): {a} ~ {b} {U}/yr — extremes saturate; set a color limit or adjust it in the colorbar window",
"lims_fixed":"Color scale (fixed): {a} ~ {b} {U}/yr",
"src_run":"Load and render",
"src_gnss_note":"Pick the GNSS time-series folder from the top-left panel after loading",
"src_pick_first":"Select a time-series HDF5 or a GMTSAR folder first",
"coh_applying":"Applying coherence mask (threshold {thr})…",
"coh_note":"Coherence mask: {n} pixels removed (threshold {thr})",
"coh_no_dataset":"No usable dataset found in the mask file",
"coh_shape_mismatch":"Mask size {a} does not match the time series {b}, and the mask file has no geo attributes to resample from",
"coh_bad_shape":"Mask dataset is not 2-D ({a})",
"coh_resampled":" (resampled by lon/lat, {n} cells fell outside the mask)",
"gmt_no_disp":"No disp_NNN_ll.xy found in the folder",
"gmt_no_datefile":"No data_date.txt found in the folder",
"gmt_date_count_mismatch":"data_date.txt has {a} rows but there are {b} disp files",
"gmt_reading":"Reading points… {i}/{n} ({name})",
"gmt_gridding":"Gridding… {i}/{n}",
"gmt_no_points":"No usable scattered points",
"gmt_grid_too_big":"Grid too large ({w}×{h}); increase the grid spacing",
"gmt_note":"GMTSAR quick-look: {step} m cell averaging, {buf} m data buffer ({rad} cells), not kriging",
"deramp_running":"Fitting and removing the linear ramp…",
"deramp_note":"Linear ramp removed from {n} epochs (deramp)",
"aoi_draw_btn":"▭ Draw AOI",
"aoi_clear_btn":"Clear AOI",
"aoi_export_png":"⬇ AOI PNG",
"aoi_export_tif":"⬇ AOI GeoTIFF",
"aoi_hint":"Drag a rectangle on the map",
"aoi_none":"No AOI selected yet",
"aoi_set":"AOI: {w}×{h} pixels",
"aoi_empty":"No valid data inside the AOI",
"aoi_label":"AOI export",
"ts_png_btn":"⬇ Time-series PNG",
"ts_png_none":"Click a map location first",
"gnss_dir_label":"GNSS time-series folder",
"gnss_dir_load":"Load",
"err_share_anchor":"Cannot find the injection point in the share HTML",
"gnss_crs_label":"CRS",
"gnss_crs_twd97":"TWD97 TM2 (EPSG:3826)",
"gnss_crs_lonlat":"Lon/Lat (EPSG:4326)",
"gnss_unit_label":"Unit",
"gnss_fmt_label":"Columns",
"gnss_dir_none":"No <STATION>_f_all.xlsx or .csv/.txt found in the folder",
"gnss_dir_loaded":"Loaded GNSS series for {n} stations (click a station to see it)",
"gnss_dir_reading":"Reading GNSS series… {i}/{n} ({name})",
"gnss_sheetjs_missing":"SheetJS is not loaded; cannot read .xlsx",
"gnss_ts_title":"GNSS {sta} ({comp})",
"gnss_comp_label":"Component",
"src_unused_en":""
}
};
window.__i18nHooks = window.__i18nHooks || [];         // functions to re-render live dynamic text
window.__i18nOverlayEntries = window.__i18nOverlayEntries || [];   // {get:()=>layer, key} for L.control.layers renames
window.currentLang = (function(){
 try{return localStorage.getItem('viewerLang')||'zh';}catch(e){return 'zh';}
})();
window.t = function(key, params){
 const dict = window.I18N[window.currentLang] || window.I18N.zh;
 let s = (dict && dict[key] != null) ? dict[key] : ((window.I18N.zh && window.I18N.zh[key]) != null ? window.I18N.zh[key] : key);
 if(!params) return s;
 return s.replace(/\{(\w+)\}/g, (m, k) => (params[k] !== undefined ? params[k] : m));
};
window.__i18nRegisterLayer = function(getter, key){
 window.__i18nOverlayEntries.push({get: getter, key: key});
};
function __i18nRefreshLayerControl(){
 const ctrl = window._lyrCtrl;
 if(!ctrl || !ctrl._layers) return;
 ctrl._layers.forEach(entry=>{
  const found = window.__i18nOverlayEntries.find(e=>e.get()===entry.layer);
  if(found) entry.name = window.t(found.key);
 });
 if(ctrl._update) ctrl._update();
}
window.__i18nHooks.push(__i18nRefreshLayerControl);
window.applyLanguage = function(lang){
 window.currentLang = lang;
 try{localStorage.setItem('viewerLang', lang);}catch(e){}
 document.documentElement.lang = lang==='en' ? 'en' : 'zh-Hant';
 document.querySelectorAll('[data-i18n]').forEach(el=>{el.textContent = window.t(el.getAttribute('data-i18n'));});
 document.querySelectorAll('[data-i18n-title]').forEach(el=>{el.title = window.t(el.getAttribute('data-i18n-title'));});
 document.querySelectorAll('[data-i18n-placeholder]').forEach(el=>{el.placeholder = window.t(el.getAttribute('data-i18n-placeholder'));});
 // data-i18n-html: like data-i18n but for the one static string that carries markup
 // (the --app file-picker description's <b>/<br>); the dictionary text is trusted,
 // author-controlled markup, not user input, so innerHTML is safe here.
 document.querySelectorAll('[data-i18n-html]').forEach(el=>{el.innerHTML = window.t(el.getAttribute('data-i18n-html'));});
 const btn = document.getElementById('langToggle');
 if(btn) btn.textContent = lang==='en' ? '中' : 'EN';
 window.__i18nHooks.forEach(fn=>{try{fn();}catch(e){console.error(e);}});
};
document.getElementById('langToggle').onclick = function(){
 window.applyLanguage(window.currentLang==='en' ? 'zh' : 'en');
};
// deferred: run after ALL synchronous inline scripts on this page (incl. the __SRC_HTML
// snapshot further down) have finished, so the GNSS-export snapshot still captures the
// template's pristine zh markup regardless of the visitor's saved language preference.
setTimeout(function(){window.applyLanguage(window.currentLang);}, 0);
</script>
<script>
// ===== 總覽小地圖 (獨立 script: --app 模式下選檔前就顯示) =====
(function(){
 const MINI_OFF=/*__MINIOFF__*/false;                // --offline 版無網路 tile → 隱藏
 const win=document.getElementById('miniwin');
 if(MINI_OFF||typeof L==='undefined'){win.style.display='none';return;}
 const mini=L.map('minimap',{zoomControl:false,dragging:false,scrollWheelZoom:false,
  doubleClickZoom:false,boxZoom:false,keyboard:false,touchZoom:false,attributionControl:false});
 L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png',{maxZoom:10}).addTo(mini);
 const MINI_BOUNDS=[[21.5,119.5],[25.5,122.5]];      // 固定全台灣, 不再隨主圖縮放
 mini.fitBounds(MINI_BOUNDS);
 let rect=null,mainRef=null;
 function upd(){
  if(!mainRef)return;
  const b=mainRef.getBounds();                       // 旋轉時=外接框
  if(rect)rect.setBounds(b);
  else rect=L.rectangle(b,{color:'#e11',weight:2,fill:true,fillOpacity:.05,interactive:false}).addTo(mini);
 }
 window.__miniAttach=function(m){mainRef=m;m.on('moveend zoom',upd);upd();};
 mini.on('click',e=>{if(mainRef)mainRef.setView(e.latlng,mainRef.getZoom());});
 // 收合/展開
 const body=document.getElementById('minimap'),btn=document.getElementById('miniToggle');
 btn.onclick=()=>{
  const hidden=body.style.display==='none';
  body.style.display=hidden?'':'none';
  btn.textContent=hidden?'—':'▣';
  if(hidden)setTimeout(()=>mini.invalidateSize(),50);
 };
 // 放大/還原: 切 CSS class 改變面板與地圖容器尺寸, Leaflet 容器尺寸變動後
 // 必須呼叫 invalidateSize() 重算內部量測, 並重新 fitBounds 維持涵蓋全資料範圍
 const expandBtn=document.getElementById('miniExpand');
 expandBtn.onclick=()=>{
  const expanded=win.classList.toggle('mini-expanded');
  expandBtn.textContent=expanded?'⤡':'⤢';
  setTimeout(()=>{mini.invalidateSize();mini.fitBounds(MINI_BOUNDS);},50);
 };
 // 面板拖曳 (標題列)
 const hdr=document.getElementById('minihdr');
 let ox,oy,on=false;
 hdr.addEventListener('mousedown',e=>{
  if(e.target===btn)return;
  on=true;ox=e.clientX-win.offsetLeft;oy=e.clientY-win.offsetTop;e.preventDefault();});
 document.addEventListener('mousemove',e=>{
  if(!on)return;
  win.style.left=(e.clientX-ox)+'px';win.style.top=(e.clientY-oy)+'px';
  win.style.right='auto';win.style.bottom='auto';
  if(window.clampToViewport)clampToViewport(win);});   // main script 定義的視窗邊界收回 (若已載入)
 document.addEventListener('mouseup',()=>on=false);
 window.addEventListener('resize',()=>{if(window.clampToViewport)clampToViewport(win);});
 window.__miniExpose={mini,getRect:()=>rect};        // E2E 掛鉤
})();
</script>
<script>
const D = /*__DATA__*/;
// 匯出「更新 GNSS」HTML 用: 儘早捕捉原始頁面原始碼 (在本 script 自身的資料指定之後,
// 但仍早於任何 map/Leaflet 初始化); 總覽小地圖(獨立 script, 先於本 script 執行)已對
// #minimap 做過 DOM 掛載, 複製一份 documentElement 並把 #minimap 內容清空, 使擷取結果
// 與原始靜態檔案一致 (自包含性/CDN/pack/offline 屬性完全跟隨原檔; #minimap 以外的內容
// 在本行之前未被任何 script 更動過, 故逐位元對應原檔).
const __srcClone = document.documentElement.cloneNode(true);
const __miniClone = __srcClone.querySelector('#minimap');
if (__miniClone) __miniClone.innerHTML = '';
const __SRC_HTML = "<!DOCTYPE html>\n" + __srcClone.outerHTML;
const km = s => s/1000;
const U = D.unit||'mm';                       // display unit (values are already embedded in it)
// 不對稱色階 (與 Python _NEG_STOPS/_POS_STOPS 單一來源同步):
// 官方分級圖圖例配色 (1 cm/yr 分級; --vlim 80 時每 stop = 5 mm/yr)
// 負側 綠→黃綠→黃→橘→紅→桃紅→粉紅(=-negLim); 正側 綠→青→藍(=+posLim)
const NEG_STOPS=[[128,255,64],[128,255,64],[160,255,64],[192,255,64],[224,240,64],[255,224,64],[255,192,64],[255,160,64],[255,128,64],[255,96,64],[255,96,128],[255,96,192],[255,128,224],[255,160,255],[255,192,255],[255,224,255],[255,240,255]];
const POS_STOPS=[[64,255,64],[64,255,64],[64,255,160],[64,255,255],[64,240,255],[64,224,255],[64,208,255],[64,192,255],[32,160,255],[0,128,255],[0,96,255],[0,64,255],[0,56,255],[0,48,255],[0,40,255],[0,32,255],[0,24,255]];
function lerpStops(stops,s){const x=s*(stops.length-1),
 i=Math.max(0,Math.min(stops.length-2,Math.floor(x))),f=x-i;
 const a=stops[i],b=stops[i+1];return `rgb(${a.map((c,k)=>Math.round(c+(b[k]-c)*f)).join(',')})`;}
function col(v,lims){lims=lims||D.lims.vel;
 return v>=0?lerpStops(POS_STOPS,Math.min(1,v/lims.pos))
            :lerpStops(NEG_STOPS,Math.min(1,-v/lims.neg));}
// meta line + layer/curve option labels: extracted into named functions so a language
// toggle can re-render them without redoing any of the actual grid/contour computation
function renderMeta(){
 const el=document.getElementById('meta');
 let s=' '+t('meta_summary',{n:D.n,nd:D.dates.length,d0:D.dates[0],d1:D.dates[D.dates.length-1]});
 if(D.hasLine)s+=t('meta_corridor',{buf:D.buffer});
 if(D._fname)s+=' · '+D._fname;
 el.textContent=s;
}
// GNSS 校正狀態 badge: D.gnssCorr 不存在(未加 --gnss-correct)或 mode='off' 時不顯示;
// mode='none'=範圍內無可用站(含全被剔除); mode='geom_fail'=LOS 幾何本身取得失敗(與 none 分開,
// 見 S12); mode='const'/'plane'=已校正(rmsWarn=true 時附品質警示); mode='stale'=瀏覽器內
// 載入新 h5 後的提示(新資料未經校正, 見 h5Apply).
function renderGnssCorrBadge(){
 const el=document.getElementById('gnssCorrBadge');
 if(!el)return;
 const gc=D.gnssCorr;
 if(!gc||gc.mode==='off'){el.textContent='';return;}
 if(gc.mode==='none'){el.textContent=' · '+t('gnss_corr_none');return;}
 if(gc.mode==='geom_fail'){el.textContent=' · '+t('gnss_corr_geom_fail');return;}
 if(gc.mode==='stale'){el.textContent=' · '+t('gnss_corr_stale');return;}
 const modelKey=gc.mode==='plane'?'gnss_corr_model_plane':'gnss_corr_model_const';
 let s=' · '+t('gnss_corr_badge',{n:gc.nsta,model:t(modelKey)});
 if(gc.rmsWarn)s+=' '+t('gnss_corr_rms_warn',{rms:gc.rmsAfter});
 el.textContent=s;
}
function renderLayerOptionLabels(){
 document.querySelector('#layerMode option[value=vel]').textContent=t('opt_vel_label',{U:U});
 document.querySelector('#layerMode option[value=cum]').textContent=t('opt_cum_label',{U:U});
 document.querySelector('#curveMode option[value=rate]').textContent=t('curve_opt_rate',{U:U});
 document.querySelector('#curveMode option[value=cum]').textContent=t('opt_cum_label',{U:U});
}
// document.title = user-supplied --title (kept verbatim, untranslated, from #titleTxt) + a
// translated suffix; re-run on toggle so the suffix follows the language, title itself doesn't
function renderTitle(){
 document.title=document.getElementById('titleTxt').textContent+' '+t('ui_title_suffix');
}
window.__i18nHooks.push(renderTitle);
renderTitle();
renderMeta();
renderGnssCorrBadge();
renderLayerOptionLabels();
window.__i18nHooks.push(renderMeta);
window.__i18nHooks.push(renderGnssCorrBadge);
window.__i18nHooks.push(renderLayerOptionLabels);
// 等值線開關 + 間距(mm, 固定; 速度場語意 mm/yr): 需求為預設關閉, 使用者主動開啟才畫也才算.
let contourEnabled=false;
function contourDefaultMM(){                 // 開啟時輸入框預設值: 沿用舊版「10b」慣例換算回 mm
 const b=(D.contours&&D.contours.base)||1;   // (base 0.1 時=1mm, 避免初次開啟就畫出過密的等值線)
 return +(10*b*(U==='cm'?10:1)).toFixed(3);
}
// 具名函式: 載入 HDF5 後 D.contours/base 換新, 直接重呼叫即可重建控制項 (不重複邏輯)
function initContourControls(){
 const cb=document.getElementById('contourEnable'),inp=document.getElementById('contourInt');
 const hasData=!!(D.contours||D.grid);   // 內嵌等值線資料或內嵌格網(可動態算), 任一即可用
 contourEnabled=false;                   // 每次資料源(初載/換 h5)一律先關, 不預先計算
 cb.checked=false;
 cb.disabled=!hasData;
 inp.disabled=true;
 inp.value=contourDefaultMM();
 cb.onchange=()=>{
  contourEnabled=cb.checked;
  inp.disabled=!cb.checked;
  buildContours(document.getElementById('layerMode').value);
 };
 inp.onchange=()=>{if(contourEnabled)buildContours(document.getElementById('layerMode').value);};
 document.getElementById('contourLabels').onchange=
  ()=>buildContours(document.getElementById('layerMode').value);
 // restyle in place on color change (labels follow via the --ctrcol CSS variable)
 document.getElementById('contourColor').oninput=()=>{
  if(!contourLayer)return;
  const cCol=document.getElementById('contourColor').value;
  document.documentElement.style.setProperty('--ctrcol',cCol);
  contourLayer.setStyle(f=>{const lev=f.properties.lev;
   const major=Math.abs(lev/5-Math.round(lev/5))<1e-6;
   return {color:cCol,weight:major?1.6:0.7,opacity:.85,fill:false};});
 };
}
initContourControls();

// 小螢幕/視窗縮小時, 把拖出可視範圍的浮動面板拉回來 (至少保留 minVis px 可見以便再抓回標題列)
const _dragWins=[],_dragFreeEls=[];
function clampToViewport(win){
 const w=win.offsetWidth||300,minVis=40;
 const maxL=window.innerWidth-minVis,minL=-(w-minVis);
 const maxT=window.innerHeight-minVis,minT=0;
 let l=win.offsetLeft,t=win.offsetTop,changed=false;
 if(l>maxL){l=maxL;changed=true;} if(l<minL){l=minL;changed=true;}
 if(t>maxT){t=maxT;changed=true;} if(t<minT){t=minT;changed=true;}
 if(changed){win.style.left=l+'px';win.style.top=t+'px';win.style.right='auto';win.style.bottom='auto';}
}
function clampFreeToViewport(el){
 const r=el.getBoundingClientRect();
 const tx=+(el.dataset.tx||0),ty=+(el.dataset.ty||0);
 let dx=0,dy=0;
 if(r.right>window.innerWidth)dx=window.innerWidth-r.right;
 if(r.left<0)dx=Math.max(dx,-r.left);
 if(r.bottom>window.innerHeight)dy=window.innerHeight-r.bottom;
 if(r.top<0)dy=Math.max(dy,-r.top);
 if(dx||dy){el.dataset.tx=tx+dx;el.dataset.ty=ty+dy;el.style.transform=`translate(${tx+dx}px,${ty+dy}px)`;}
}
window.addEventListener('resize',()=>{_dragWins.forEach(clampToViewport);_dragFreeEls.forEach(clampFreeToViewport);});

// 視窗可拖曳
function drag(win,hdr){let ox,oy,on=false;
 _dragWins.push(win);
 hdr.addEventListener('mousedown',e=>{on=true;ox=e.clientX-win.offsetLeft;oy=e.clientY-win.offsetTop;e.preventDefault();});
 document.addEventListener('mousemove',e=>{if(!on)return;win.style.left=(e.clientX-ox)+'px';win.style.top=(e.clientY-oy)+'px';win.style.right='auto';win.style.bottom='auto';clampToViewport(win);});
 document.addEventListener('mouseup',()=>on=false);}
drag(document.getElementById('tswin'),document.getElementById('tshdr'));
drag(document.getElementById('linewin'),document.getElementById('linehdr'));
drag(document.getElementById('profwin'),document.getElementById('profhdr'));
drag(document.getElementById('cbar'),document.getElementById('cbarhdr'));

// drag-anywhere for panels without a header bar (#ctrl, Leaflet layer control):
// grab any empty spot (interactive children excluded); a real drag suppresses the
// trailing click so layer-control checkboxes don't toggle after a move
function dragFree(el){
 let sx,sy,bx,by,on=false,moved=false;
 _dragFreeEls.push(el);
 el.addEventListener('mousedown',e=>{
  if(e.target.closest('input,select,button,a,textarea'))return;
  on=true;moved=false;sx=e.clientX;sy=e.clientY;
  bx=+(el.dataset.tx||0);by=+(el.dataset.ty||0);
 });
 document.addEventListener('mousemove',e=>{
  if(!on)return;
  const dx=e.clientX-sx,dy=e.clientY-sy;
  if(!moved&&Math.abs(dx)+Math.abs(dy)<=3)return;
  moved=true;
  el.dataset.tx=bx+dx;el.dataset.ty=by+dy;
  el.style.transform=`translate(${bx+dx}px,${by+dy}px)`;
  clampFreeToViewport(el);
  e.preventDefault();
 });
 document.addEventListener('mouseup',()=>{on=false;});
 el.addEventListener('click',e=>{if(moved){e.stopPropagation();e.preventDefault();moved=false;}},true);
}
dragFree(document.getElementById('ctrl'));
// layer control is created later during map init — defer until the script has finished
setTimeout(()=>{if(window._lyrCtrl)dragFree(_lyrCtrl.getContainer());},0);

// ---- 功能 A: 圖層切換 (速度場 / 總累積變形量) ----
function layerVal(p,mode){const last=p.t[p.t.length-1];return mode==='cum'?last:p.v;}
function fmtLim(x){const a=Math.abs(x);return a>=10?x.toFixed(0):a>=1?x.toFixed(1):x.toFixed(2);}
// ---- client 端衍生層: 遮罩後工作格網 / 手動色階範圍 / 重繪 raster / 動態等值線 ----
let manualLims={vel:null,cum:null};          // 手動色階 (per mode), null=自動
let autoLims=null;                           // 遮罩/重算後的自動 lims
let clientF=null;                            // {vel,cum} 由 GRID 重建的顯示場
let clientRasters=null;                      // client 重繪的 raster dataURL
let maskApplied=false,origArr=null;          // 遮罩狀態 + 原始格網副本
// DISPF: 目前地圖 raster 實際上色所用的場 (單一登錄點, 閾值/框線一律讀這裡, 不再各自猜來源).
// 掛在 window 而非 script-scope let: --app 版把本檔 (viewer_js) 包進 startViewer(D) 函式,
// _APP_JS/分享版 bootstrap 是外層獨立 <script>, 唯有 window 屬性能跨這兩種 scope 互通.
window.DISPF=null;                           // {vel,cum,src} src∈'client'|'h5'|'app'|'share', null=尚無登錄→GRID 現場算
const _ctrTplCache=new Map();
function effLims(mode){return manualLims[mode]||(autoLims?autoLims[mode]:D.lims[mode]);}
// 期數時間軸 (自首期起算的十進位年): ISO 日期 (--ts 版) 用 日數/365.25, 與 python
// make_pt_ts_h5.decyears 同式; 純年份字串 (--app 版) 用年差.
// 註: 舊版直接寫 (+D.dates[i]) 取年差, 對 "2025-04-03" 會得到 NaN 而讓 span 恆為 1
// (遮罩後的速度場變成「總位移」而非 mm/yr) — 此函式一併修正該既有問題.
function epochYearsFrom(dates){
 const iso=/^(\d{4})-(\d{2})-(\d{2})$/;
 const day=s=>{const m=iso.exec(String(s));
  return m?Date.UTC(+m[1],+m[2]-1,+m[3])/86400000:null;};
 const d0=day(dates[0]);
 if(d0!==null)return dates.map(s=>{const d=day(s);return d===null?NaN:(d-d0)/365.25;});
 const y0=+dates[0];
 return dates.map(s=>+s-y0);
}
function datesSpanYr(){
 const t=epochYearsFrom(D.dates),s=t[t.length-1];
 return (isFinite(s)&&s>0)?s:1;
}
function lerpRGB(stops,s){const x=s*(stops.length-1),
 i=Math.max(0,Math.min(stops.length-2,Math.floor(x))),f=x-i;
 const a=stops[i],b=stops[i+1];
 return [Math.round(a[0]+(b[0]-a[0])*f),Math.round(a[1]+(b[1]-a[1])*f),Math.round(a[2]+(b[2]-a[2])*f)];}
function asymLimsT(arr){
 const fin=[];
 for(let i=0;i<arr.length;i++)if(isFinite(arr[i]))fin.push(arr[i]);
 if(!fin.length)return {neg:1,pos:1};
 fin.sort((a,b)=>a-b);
 const q=p=>{const t=p*(fin.length-1),i=Math.floor(t);
  return fin[i]+(fin[Math.min(i+1,fin.length-1)]-fin[i])*(t-i);};
 let neg=-q(0.005),pos=q(0.995);
 if(!(neg>0))neg=Math.max(pos*0.05,1e-6);
 if(!(pos>0))pos=Math.max(neg*0.05,1e-6);
 return {neg:+neg.toFixed(3),pos:+pos.toFixed(3)};
}
function rasterPNGT(arr,lims){
 const W=GRID.W,H=GRID.H;
 const cv=document.createElement('canvas');cv.width=W;cv.height=H;
 const ctx=cv.getContext('2d');const img=ctx.createImageData(W,H);const d=img.data;
 for(let i=0;i<W*H;i++){
  const v=arr[i];
  if(!isFinite(v))continue;
  const rgb=v>=0?lerpRGB(POS_STOPS,Math.min(1,v/lims.pos)):lerpRGB(NEG_STOPS,Math.min(1,-v/lims.neg));
  d[i*4]=rgb[0];d[i*4+1]=rgb[1];d[i*4+2]=rgb[2];d[i*4+3]=255;
 }
 ctx.putImageData(img,0,0);
 return cv.toDataURL('image/png');
}
function rebuildClientFields(){              // GRID (工作格網) → 顯示場 (顯示單位)
 const n=GRID.W*GRID.H,off=(GRID.N-1)*n;
 const span=datesSpanYr();
 const vel=new Float32Array(n).fill(NaN),cum=new Float32Array(n).fill(NaN);
 for(let i=0;i<n;i++){
  const q0=GRID.arr[i],q1=GRID.arr[off+i];
  if(q0===GRID.nodata||q1===GRID.nodata)continue;
  vel[i]=(q1-q0)*GRID.scale/span;cum[i]=q1*GRID.scale;
 }
 clientF={vel,cum};
}
function refreshDerived(){                   // 遮罩/手動範圍變更後統一重算重繪
 if(!GRID)return;
 if(maskApplied||manualLims.vel||manualLims.cum){
  rebuildClientFields();
  autoLims={vel:asymLimsT(clientF.vel),cum:asymLimsT(clientF.cum)};
  clientRasters={vel:rasterPNGT(clientF.vel,effLims('vel')),
                 cum:rasterPNGT(clientF.cum,effLims('cum'))};
  // raster 改用 clientF 重繪 → 登錄同一份場, 閾值/等值線與畫面同源
  window.DISPF={vel:clientF.vel,cum:clientF.cum,src:'client'};
 }else{
  clientF=null;autoLims=null;clientRasters=null;
  // 退回未遮罩/未手動調色階狀態: raster 回到 h5Apply 當初設的 D.rasters (h5F 同源),
  // 沒載過 h5 就清空登錄, currentFieldT 改走 GRID 現場算 (依 D.velAlgo 選對算法)
  // 沒載過 h5 時, 退回 --build 模式載入當下算好的原始場 (window.__APPF);
  // 清成 null 會讓 currentFieldT 改走 computeFieldFromGridT 現場算, 與載入時的場不同源
  window.DISPF=h5F?{vel:h5F.vel,cum:h5F.cum,src:'h5'}
   :(window.__APPF?{vel:window.__APPF.vel,cum:window.__APPF.cum,src:'app'}:null);
 }
 _ctrTplCache.clear();
 applyLayer(document.getElementById('layerMode').value);
}
// marching squares (template 版: 遮罩後等值線重建用; 與 app 端同邏輯)
function marchLevelT(f,W,H,lev){
 const segs=[];
 for(let r=0;r<H-1;r++){
  const ro=r*W;
  for(let c=0;c<W-1;c++){
   const v00=f[ro+c],v10=f[ro+c+1],v01=f[ro+W+c],v11=f[ro+W+c+1];
   if(!isFinite(v00)||!isFinite(v10)||!isFinite(v01)||!isFinite(v11))continue;
   let idx=0;
   if(v00>=lev)idx|=1;if(v10>=lev)idx|=2;if(v11>=lev)idx|=4;if(v01>=lev)idx|=8;
   if(idx===0||idx===15)continue;
   const T=[c+(lev-v00)/(v10-v00),r],Rt=[c+1,r+(lev-v10)/(v11-v10)],
         B=[c+(lev-v01)/(v11-v01),r+1],Lf=[c,r+(lev-v00)/(v01-v00)];
   switch(idx){
    case 1:case 14:segs.push([Lf,T]);break;
    case 2:case 13:segs.push([T,Rt]);break;
    case 3:case 12:segs.push([Lf,Rt]);break;
    case 4:case 11:segs.push([Rt,B]);break;
    case 5:segs.push([Lf,T],[Rt,B]);break;
    case 6:case 9:segs.push([T,B]);break;
    case 7:case 8:segs.push([Lf,B]);break;
    case 10:segs.push([T,Rt],[B,Lf]);break;
   }
  }
 }
 return segs;
}
function chainSegmentsT(segs){
 const key=p=>p[0].toFixed(3)+','+p[1].toFixed(3);
 const adj=new Map();
 const add=(k,i)=>{if(!adj.has(k))adj.set(k,[]);adj.get(k).push(i);};
 segs.forEach((s,i)=>{add(key(s[0]),i);add(key(s[1]),i);});
 const used=new Array(segs.length).fill(false);
 const lines=[];
 for(let i=0;i<segs.length;i++){
  if(used[i])continue;
  used[i]=true;
  const line=[segs[i][0],segs[i][1]];
  for(const dir of [1,0]){
   for(;;){
    const end=dir?line[line.length-1]:line[0];
    const k=key(end);
    const cands=(adj.get(k)||[]).filter(si=>!used[si]);
    if(!cands.length)break;
    const si=cands[0];used[si]=true;
    const s=segs[si];
    const nxt=key(s[0])===k?s[1]:s[0];
    if(dir)line.push(nxt);else line.unshift(nxt);
   }
  }
  lines.push(line);
 }
 return lines;
}
function marchingContoursT(f,W,H,gt,interval){
 let vmin=Infinity,vmax=-Infinity;
 for(let i=0;i<W*H;i++){const v=f[i];if(isFinite(v)){if(v<vmin)vmin=v;if(v>vmax)vmax=v;}}
 if(!(vmax>vmin))return [];
 const feats=[];
 for(let lev=Math.ceil(vmin/interval)*interval;lev<=vmax+1e-9;lev+=interval){
  const L2=+lev.toFixed(6);
  for(const ln of chainSegmentsT(marchLevelT(f,W,H,L2))){
   if(ln.length<2)continue;
   feats.push({type:'Feature',properties:{lev:+L2.toFixed(3)},
    geometry:{type:'LineString',coordinates:ln.map(p=>[
     +(gt[0]+(p[0]+0.5)*gt[1]).toFixed(5),+(gt[3]+(p[1]+0.5)*gt[5]).toFixed(5)])}});
  }
 }
 return feats;
}

// 載入 HDF5 後的顯示場 (速度/總累積, 顯示單位): 動態等值線的來源.
// --app 版由 _APP_JS 提供 window.computeContours; --ts 版沒有, 故用本函式頂上.
let h5F=null;
function h5Contours(mode,interval){
 if(!h5F||!GRID)return [];
 const key='h5_'+mode+'_'+interval;
 if(!_ctrTplCache.has(key))
  _ctrTplCache.set(key,marchingContoursT(mode==='cum'?h5F.cum:h5F.vel,
   GRID.W,GRID.H,GRID.gt,interval));
 return _ctrTplCache.get(key);
}

let lastLayerMode='vel';   // for the language-toggle hook to re-render cbar/threshold-unit text
// cbarTitle/cbarDesc/thUnit text for a given mode: pulled out of applyLayer so a language
// toggle can redraw just the text without recomputing raster/contours/markers
function renderCbarText(mode){
 document.getElementById('cbarTitle').textContent=mode==='cum'?t('cbar_title_cum',{U:U}):t('cbar_title_vel',{U:U});
 document.getElementById('cbarDesc').textContent=mode==='cum'?t('cbar_desc_cum'):t('cbar_desc_vel');
 document.getElementById('thUnit').textContent=mode==='cum'?U:`${U}/yr`;
 // 等值線間距輸入框單位: 固定 mm, 速度場圖層語意為 mm/yr (與 U/顯示單位無關, 見需求)
 document.getElementById('contourIntUnit').textContent=mode==='cum'?'mm':'mm/yr';
}
window.__i18nHooks.push(()=>renderCbarText(lastLayerMode));
function applyLayer(mode){
 lastLayerMode=mode;
 const lims=effLims(mode);
 const R=clientRasters||D.rasters;
 if(rasterOv&&R)rasterOv.setUrl(mode==='cum'?R.cum:R.vel);
 markers.forEach((mk,i)=>{const v=layerVal(D.points[i],mode);
  mk.setStyle({fillColor:(v==null||!isFinite(v))?'#888':col(v,lims)});});
 renderCbarText(mode);
 // 漸層 stop 直接取 NEG_STOPS/POS_STOPS (與 col() 同一來源), 依實際
 // negLim/posLim 比例擺放; 零點自左=neg/(neg+pos)
 const z=lims.neg/(lims.neg+lims.pos)*100;
 const stops=[];
 for(let j=NEG_STOPS.length-1;j>=0;j--)
  stops.push(`rgb(${NEG_STOPS[j].join(',')}) ${(z*(1-j/(NEG_STOPS.length-1))).toFixed(1)}%`);
 for(let j=0;j<POS_STOPS.length;j++)
  stops.push(`rgb(${POS_STOPS[j].join(',')}) ${(z+(100-z)*j/(POS_STOPS.length-1)).toFixed(1)}%`);
 document.getElementById('cbarGrad').style.background=`linear-gradient(90deg,${stops.join(',')})`;
 renderCbarTicks(lims);
 buildContours(mode);
 refreshThresholdIfActive();
}
document.getElementById('layerMode').onchange=e=>{
 // 雙向連動: colorbar 視窗的緊湊切換鈕跟著同步 (單一 applyLayer 呼叫鏈, 不重複邏輯)
 const cbarSel=document.getElementById('cbarLayerSel');
 if(cbarSel&&cbarSel.value!==e.target.value)cbarSel.value=e.target.value;
 applyLayer(e.target.value);
};
document.getElementById('cbarLayerSel').onchange=e=>{
 // colorbar 切換鈕本身不呼叫 applyLayer: 改設 #layerMode 的值再觸發同一個 change handler
 const lm=document.getElementById('layerMode');
 lm.value=e.target.value;
 lm.dispatchEvent(new Event('change'));
};

// ---- cbar 刻度: nice-step 均勻刻度 + 碰撞略過 (端點與 0 必標, 重疊時 0 降到第二列) ----
let curCbarLims=null;
function renderCbarTicks(lims){
 curCbarLims=lims||curCbarLims;
 if(!curCbarLims)return;
 const neg=curCbarLims.neg,pos=curCbarLims.pos,span=neg+pos;
 const el=document.getElementById('cbarTicks');
 const wpx=el.clientWidth||200;
 const pct=v=>(v+neg)/span*100;
 const lw=v=>fmtLim(v).length*6.5+4;        // rough label width (px)
 // nice step: 整數/半整數等間隔, 全帶 ≤6 格
 const steps=[0.1,0.2,0.25,0.5,1,2,2.5,5,10,20,25,50,100,200,250,500];
 const step=steps.find(s=>span/s<=6)||Math.pow(10,Math.ceil(Math.log10(span/6)));
 const placed=[];                            // {v,x,w,row}
 function collide(x,w,row){return placed.some(p=>p.row===row&&Math.abs(p.x-x)<(p.w+w)/2);}
 // 端點與 0 必標; 0 與端點重疊時放到第二列
 for(const v of [-neg,pos,0]){
  const x=pct(v)/100*wpx,w=lw(v);
  placed.push({v,x,w,row:collide(x,w,0)?1:0});
 }
 // 內插 nice 值: 放得下才加 (自動略過)
 for(let v=Math.ceil(-neg/step)*step;v<pos-step*0.25;v+=step){
  const vv=+v.toFixed(6);
  if(Math.abs(vv)<1e-9)continue;
  const x=pct(vv)/100*wpx,w=lw(vv);
  if(!collide(x,w,0))placed.push({v:vv,x,w,row:0});
 }
 const rows=placed.some(p=>p.row===1)?2:1;
 el.style.height=(rows*13)+'px';
 el.innerHTML=placed.map(p=>{
  const pc100=pct(p.v);
  const tx=pc100<4?'0':pc100>96?'-100%':'-50%';
  return `<span style="position:absolute;left:${pc100.toFixed(2)}%;top:${p.row*13}px;`+
   `transform:translateX(${tx})">${fmtLim(p.v)}</span>`;
 }).join('');
}
// 視窗縮放時依新寬度重排刻度
new ResizeObserver(()=>renderCbarTicks()).observe(document.getElementById('cbarTicks'));

// ---- 等值線 overlay: 隨圖層切換換組, 依下拉間距過濾, 5 的倍數畫粗, hover 顯示數值 ----
let contourLayer=null;
window.__i18nRegisterLayer(()=>contourLayer,'lyr_contour');
// level 是否為選定間距的整數倍 (浮點容差 1e-6)
function isLevelMultiple(lev,interval){
 const r=lev/interval;
 return Math.abs(r-Math.round(r))<1e-6;
}
function filterContourFeatures(features,interval){
 return features.filter(f=>isLevelMultiple(f.properties.lev,interval));
}
function buildContours(mode){
 const wasOn=contourLayer?map.hasLayer(contourLayer):true;   // 預設開
 if(contourLayer){map.removeLayer(contourLayer);_lyrCtrl.removeLayer(contourLayer);contourLayer=null;}
 const st=document.getElementById('ctrStatus');
 st.textContent='';
 if(!contourEnabled)return;                                  // 開關關閉(預設): 不畫也不算
 const mmVal=parseFloat(document.getElementById('contourInt').value);
 if(!(mmVal>0)){st.textContent=t('ctr_invalid_interval');return;}
 const sel=mmVal/(U==='cm'?10:1);        // 輸入框固定 mm; 顯示單位=cm 時內部換算成顯示單位
 if(maskApplied&&clientF){
  // 遮罩後: 一律用 client 場即時重算 (静態 D.contours 已不代表遮罩後資料)
  st.textContent=t('ctr_computing');
  setTimeout(()=>{
   try{
    const key=mode+'_'+sel;
    if(!_ctrTplCache.has(key))
     _ctrTplCache.set(key,marchingContoursT(mode==='cum'?clientF.cum:clientF.vel,
      GRID.W,GRID.H,GRID.gt,sel));
    st.textContent='';
    const feats=_ctrTplCache.get(key);
    if(feats.length)renderContours(feats,mode,wasOn);
   }catch(err){st.textContent=t('ctr_failed');console.error(err);}
  },30);
  return;
 }
 if(D.contours&&D.contours.dynamic){
  // 動態模式 (--app / 載入 HDF5 後): JS marching squares 即時計算 (含快取), 計算中顯示提示
  st.textContent=t('ctr_computing');
  setTimeout(()=>{
   try{
    const feats=window.computeContours?window.computeContours(mode,sel):h5Contours(mode,sel);
    st.textContent='';
    if(feats&&feats.length)renderContours(feats,mode,wasOn);
   }catch(err){st.textContent=t('ctr_failed');console.error(err);}
  },30);
  return;
 }
 // 快路徑: 純靜態頁面內嵌等值線, 間距為底稿(base)整數倍時直接濾用既有幾何, 免即時運算
 if(D.contours){
  const src=D.contours[mode],base=D.contours.base;
  if(src&&src.features.length&&base&&isLevelMultiple(sel,base)){
   const feats=filterContourFeatures(src.features,sel);
   if(feats.length){renderContours(feats,mode,wasOn);return;}
  }
 }
 // 間距非底稿整數倍(或本檔未內嵌等值線): 唯有內嵌 GRID 才能對任意間距動態算 marching squares;
 // 無 GRID (--no-grid 產出) 則只能提示 (currentFieldT 與閾值區域同源, 保證與畫面同一份場)
 if(!GRID){st.textContent=t('ctr_no_grid_interval');return;}
 st.textContent=t('ctr_computing');
 setTimeout(()=>{
  try{
   const key='g_'+mode+'_'+sel;
   if(!_ctrTplCache.has(key)){
    const f=currentFieldT(mode);
    if(!f){st.textContent=t('ctr_no_grid_interval');return;}
    _ctrTplCache.set(key,marchingContoursT(f,GRID.W,GRID.H,GRID.gt,sel));
   }
   st.textContent='';
   const feats=_ctrTplCache.get(key);
   if(feats&&feats.length)renderContours(feats,mode,wasOn);
  }catch(err){st.textContent=t('ctr_failed');console.error(err);}
 },30);
}
function renderContours(feats,mode,wasOn){
 const gj={type:'FeatureCollection',features:feats};
 const unitStr=mode==='cum'?U:`${U}/yr`;
 const cCol=document.getElementById('contourColor').value||'#222';
 document.documentElement.style.setProperty('--ctrcol',cCol);
 const gjLayer=L.geoJSON(gj,{
  style:f=>{const lev=f.properties.lev;
   const major=Math.abs(lev/5-Math.round(lev/5))<1e-6;       // 值為 5 的倍數畫粗
   return {color:cCol,weight:major?1.6:0.7,opacity:.85,fill:false};},
  onEachFeature:(f,l)=>l.bindTooltip(`${f.properties.lev} ${unitStr}`,{sticky:true})});
 // value labels at each line's midpoint vertex; skip tiny fragments to limit clutter
 const labGrp=L.layerGroup();
 if(document.getElementById('contourLabels').checked){
  for(const f of feats){
   const cs=f.geometry.coordinates;
   if(!cs||cs.length<8)continue;
   const mid=cs[Math.floor(cs.length/2)];
   L.marker([mid[1],mid[0]],{interactive:false,icon:L.divIcon({className:'ctr-lab',
    html:`<span>${f.properties.lev}</span>`,iconSize:null})}).addTo(labGrp);
  }
 }
 contourLayer=L.featureGroup([gjLayer,labGrp]);   // featureGroup so setStyle propagates to lines
 if(wasOn)contourLayer.addTo(map);
 _lyrCtrl.addOverlay(contourLayer,t('lyr_contour'));
}

// ---- 閾值區域: 目前圖層數值場中 "值<閾值" 的面積統計 + 邊界框線 ----
// 數值場來源 (與畫面 raster 上色同源, 修復同源缺陷): 一律先讀 window.DISPF —— 這是每個
// 會重繪 raster 的地方 (refreshDerived 的 clientF 分支/h5F 分支、--app 的 velC、分享版
// 的 _shF) 各自登錄「這次實際拿去上色的那份場」的單一登錄點, 因此不論頁面目前處在哪個
// 狀態 (遮罩/手動色階/載入 h5/--app/分享版), 只要 DISPF 有值就保證與畫面同源.
//  1) window.DISPF 有值: 直接用它 (與目前 raster 同一份場, 見上)
//  2) DISPF 為 null (初始 python 直出頁面, 未載過 h5、未遮罩、未手動調色階): 用內嵌
//     GRID 現場重算 vel/cum —— vel 依 D.velAlgo (python main() 依實際速度來源設定:
//     "ols"=逐格最小二乘回歸與 --vel 外部 GeoTIFF 同式; "endpoint"=首尾期簡單差分,
//     即 _vel_from_ts() 分支) 選對應算法, 與該次頁面實際顯示的 vel raster 一致.
//     D.velAlgo 缺席 (--app 版 D 由瀏覽器端組出, 不經 python) 時預設 "endpoint",
//     與 --app 自己算 velC 的公式 (a1-a0)/spanYr 相同.
//  3) 無內嵌 GRID (--no-grid 版): 回傳 null, 功能停用
const _gridFieldCache=new Map();             // key=mode(+算法), GRID 換掉 (h5Apply) 時清空
function computeFieldFromGridT(mode){
 const n=GRID.W*GRID.H;
 const algo=mode==='vel'?(D.velAlgo||'endpoint'):'cum';
 const key=mode+'_'+algo;
 if(_gridFieldCache.has(key))return _gridFieldCache.get(key);
 let out;
 if(mode==='cum'){
  const off=(GRID.N-1)*n;
  out=new Float32Array(n).fill(NaN);
  for(let i=0;i<n;i++){
   const q=GRID.arr[off+i];
   out[i]=q===GRID.nodata?NaN:q*GRID.scale;
  }
 }else if(algo==='endpoint'){
  const off=(GRID.N-1)*n,span=datesSpanYr();
  out=new Float32Array(n).fill(NaN);
  for(let i=0;i<n;i++){
   const q0=GRID.arr[i],q1=GRID.arr[off+i];
   if(q0===GRID.nodata||q1===GRID.nodata)continue;
   out[i]=(q1-q0)*GRID.scale/span;
  }
 }else{
  const vals=new Float32Array(GRID.N*n);
  for(let i=0;i<vals.length;i++){
   const q=GRID.arr[i];
   vals[i]=q===GRID.nodata?NaN:q*GRID.scale;
  }
  out=h5OlsVel(vals,GRID.N,n,epochYearsFrom(D.dates));
 }
 _gridFieldCache.set(key,out);
 return out;
}
function currentFieldT(mode){
 if(window.DISPF)return window.DISPF[mode];
 if(GRID)return computeFieldFromGridT(mode);
 return null;
}
// 逐列以列緯度算像元面積 (km²): (|gt[1]|*111.320*cos(lat)) * (|gt[5]|*110.574)
function computeThresholdAreaT(f,W,H,gt,thresh){
 let area=0,count=0;
 const dLon=Math.abs(gt[1]),dLat=Math.abs(gt[5]);
 for(let r=0;r<H;r++){
  const lat=gt[3]+(r+0.5)*gt[5];
  const pxArea=dLon*111.320*Math.cos(lat*Math.PI/180)*dLat*110.574;
  const ro=r*W;
  for(let c=0;c<W;c++){
   const v=f[ro+c];
   if(!isFinite(v))continue;
   if(v<thresh){count++;area+=pxArea;}
  }
 }
 return {area,count};
}
// 閾值邊界: 對同一數值場以 level=thresh 跑單層 marching squares (與等值線同一引擎,
// NaN 四角自動跳過故不會穿越無資料區), 回傳 [[[lon,lat],...], ...] 折線陣列
function thresholdBoundaryT(f,W,H,gt,thresh){
 const lines=chainSegmentsT(marchLevelT(f,W,H,thresh));
 return lines.filter(ln=>ln.length>=2).map(ln=>ln.map(p=>
  [gt[0]+(p[0]+0.5)*gt[1],gt[3]+(p[1]+0.5)*gt[5]]));
}
let thLayer=null,thEnabled=false;
window.__i18nRegisterLayer(()=>thLayer,'lyr_threshold');
let lastThresholdResult=null;   // {thresh,unit,area,count}, re-formatted on language toggle
function renderThresholdText(){
 if(!lastThresholdResult)return;
 document.getElementById('thArea').textContent=t('thr_result',lastThresholdResult);
}
window.__i18nHooks.push(renderThresholdText);
function applyThreshold(){
 const areaEl=document.getElementById('thArea');
 const thresh=parseFloat(document.getElementById('thVal').value);
 if(!isFinite(thresh)){areaEl.textContent=t('thr_invalid');return;}
 const mode=document.getElementById('layerMode').value;
 // GRID 情境 4 (無遮罩/無手動範圍/無 h5) 的 vel 場需全期最小二乘回歸, 585x714x40 這類
 // 規模約需 0.1-0.3 秒: 先顯示提示再 setTimeout 讓畫面先重繪, 與 buildContours 同慣例
 areaEl.textContent=t('msg_computing');
 // 沿用 buildContours 的開關延續慣例: 重算時保留使用者原本對「閾值區域邊界」層的開/關狀態
 const wasOn=thLayer?map.hasLayer(thLayer):true;
 setTimeout(()=>{
  try{
   const f=currentFieldT(mode);
   if(!f){areaEl.textContent=t('thr_no_grid');return;}
   const {W,H,gt}=GRID;
   const {area,count}=computeThresholdAreaT(f,W,H,gt,thresh);
   const unitStr=mode==='cum'?U:`${U}/yr`;
   lastThresholdResult={thresh:thresh,unit:unitStr,area:area.toFixed(2),count:count};
   areaEl.textContent=t('thr_result',lastThresholdResult);
   if(thLayer){map.removeLayer(thLayer);_lyrCtrl.removeLayer(thLayer);thLayer=null;}
   const col=document.getElementById('thColor').value||'#ff00ff';
   const lines=thresholdBoundaryT(f,W,H,gt,thresh);
   thLayer=L.layerGroup(lines.map(coords=>L.polyline(
    coords.map(([lon,lat])=>[lat,lon]),{color:col,weight:2.5,opacity:.9,fill:false})));
   if(wasOn)thLayer.addTo(map);
   _lyrCtrl.addOverlay(thLayer,t('lyr_threshold'));
   thEnabled=true;
  }catch(err){areaEl.textContent=t('thr_failed')+err;console.error(err);}
 },30);
}
function clearThreshold(){
 if(thLayer){map.removeLayer(thLayer);_lyrCtrl.removeLayer(thLayer);thLayer=null;}
 document.getElementById('thArea').textContent='';
 lastThresholdResult=null;
 thEnabled=false;
}
// 圖層切換/遮罩套用清除/h5 載入後由 applyLayer 尾端呼叫: 已套用過才重算, 否則不動作
function refreshThresholdIfActive(){
 if(!thEnabled||document.getElementById('thVal').value==='')return;
 applyThreshold();
}
document.getElementById('thApply').onclick=applyThreshold;
document.getElementById('thClear').onclick=clearThreshold;
document.getElementById('thVal').addEventListener('keydown',e=>{if(e.key==='Enter')applyThreshold();});
document.getElementById('thColor').oninput=()=>{
 if(!thLayer)return;
 const col=document.getElementById('thColor').value;
 thLayer.eachLayer(l=>l.setStyle({color:col}));
};

// 地圖 + 衛星底圖 (可旋轉)
// 讓廊道起點在左、終點在右(與剖面x軸0→40km同向): bearing+90 (leaflet-rotate正負向實測校準)
// 廊道版: bearing+90 讓起點在左; 無廊道線時預設正北 (bearing 0)
const initB=D.hasLine?Math.round((((D.bearing||0)+90)%360+360)%360):0;
// 注意: 不可在建構時傳 bearing(leaflet-rotate 會在設視圖前套旋轉→"Set map center and zoom first"錯,
// 導致 tiles/markers 全不出來=空白地圖). 必須 fitBounds 設好視圖後才 setBearing.
const map=L.map('map',{rotate:true,touchRotate:true,shiftKeyRotate:true,
 rotateControl:{closeOnZeroBearing:false}});
/*__BASEMAP__*/
// 確保有 layer control 可掛 overlay (離線版沒有底圖切換 control 時自建一個)
if(!window._lyrCtrl)window._lyrCtrl=L.control.layers(null,null,{position:'topleft',collapsed:false}).addTo(map);
const lts=D.points.map(p=>p.lat),lns=D.points.map(p=>p.lon);
map.fitBounds([[Math.min(...lts),Math.min(...lns)],[Math.max(...lts),Math.max(...lns)]]);
if(window.__miniAttach)__miniAttach(map);          // 總覽小地圖: 紅框同步 + 點擊跳轉
// 變形場 raster 半透明疊加 (速度場/總累積由「圖層」選單切換), 預設開啟
let rasterOv=null;
window.__i18nRegisterLayer(()=>rasterOv,'lyr_raster');
if(D.rasters){
 rasterOv=L.imageOverlay(D.rasters.vel,D.rasters.bounds,{opacity:0.6}).addTo(map);
 _lyrCtrl.addOverlay(rasterOv,t('lyr_raster'));
}
// PS-point overlay removed (leveling grids carry no PS input); keep the empty
// array so recolor/highlight loops stay no-ops and D.points still feeds CLI profiles
const markers=[];
applyLayer('vel');
// 透明度滑桿: 即時調整 raster 疊加透明度
const rop=document.getElementById('rasterOpacity');
rop.oninput=e=>{const v=+e.target.value;
 document.getElementById('rasterOpacityVal').textContent=v+'%';
 if(rasterOv)rasterOv.setOpacity(v/100);};
// 剖面線顏色: 單一來源, 供 CLI 中線 / 互動剖面線 / 畫線中暫時線共用 (選色器見 #profLineColor)
let profColor=document.getElementById('profLineColor').value||'#ffcc00';
let cliLine=null;                                  // CLI --line 中線 handle, 供選色器就地換色
// 中線 + 起訖點里程
if(D.line&&D.line.length){const ll=D.line.map(c=>[c[1],c[0]]);
 cliLine=L.polyline(ll,{color:profColor,weight:2,opacity:.9,dashArray:'5 4'}).addTo(map);
 const endkm=(D.points.reduce((m,p)=>p.s>m?p.s:m,0)/1000).toFixed(1);
 L.circleMarker(ll[0],{radius:6,color:'#fff',weight:1.5,fillColor:'#22c55e',fillOpacity:1})
  .addTo(map).bindTooltip(t('prof_start'),{permanent:true,direction:'right',className:'muted'});
 L.circleMarker(ll[ll.length-1],{radius:6,color:'#fff',weight:1.5,fillColor:'#ef4444',fillOpacity:1})
  .addTo(map).bindTooltip(t('prof_end',{km:endkm}),{permanent:true,direction:'left'});}
// GNSS 站位疊加: 三角形 divIcon + 站名 tooltip, 掛 layer control 當 overlay(預設開啟)
// 包成可重呼叫的函式: 瀏覽器內載入新 GNSS CSV 後可重建整層 (見 gnssCsvFile 載入邏輯)
let gnssLayer=null;
window.__i18nRegisterLayer(()=>gnssLayer,'lyr_gnss');
// 未過濾的完整站清單: D.gnss 會依目前資料範圍裁掉範圍外的站 (載入 HDF5 時),
// 保留原始清單才不會在切換區域時把使用者載入的測站永久刪掉
let gnssAll=(D.gnss||[]).slice();
function buildGnssLayer(){
 const wasOn=gnssLayer?map.hasLayer(gnssLayer):true;         // 預設開; 重建時延續原開關狀態
 if(gnssLayer){map.removeLayer(gnssLayer);_lyrCtrl.removeLayer(gnssLayer);gnssLayer=null;}
 if(!D.gnss||!D.gnss.length)return;
 gnssLayer=L.layerGroup();
 const triIcon=L.divIcon({className:'',iconSize:[14,12],iconAnchor:[7,6],
  html:'<div class="gnss-tri"></div>'});
 D.gnss.forEach(g=>{
  const mk=L.marker([g.lat,g.lon],{icon:triIcon})
   .bindTooltip(g.n,{permanent:true,direction:'top',offset:[0,-7],className:'gnss-name'})
   .bindPopup(`<b>${t('gnss_popup_title',{n:g.n})}</b><br>${t('gnss_popup_lon',{v:g.lon})}<br>${t('gnss_popup_lat',{v:g.lat})}`)
   .addTo(gnssLayer);
  // 載入過 GNSS 時序資料夾後, 點測站直接在時序視窗畫 E/N/U 三分量
  mk.on('click',()=>{if(window.GNSS_TS&&window.GNSS_TS[g.n])showGnssTS(g.n);});});
 if(wasOn)gnssLayer.addTo(map);
 _lyrCtrl.addOverlay(gnssLayer,t('lyr_gnss'));
}
buildGnssLayer();

// ==================== GNSS 時序資料夾 (瀏覽器端讀 .xlsx / .csv / .txt) ====================
// 欄位順序由 UI 的「欄位」欄指定 (date/year/doy/n/e/u/ignore), 座標系與單位亦由 UI 決定;
// NAS 預設格式 <站碼>_f_all.xlsx 表頭為 Date/DOY/Year/N/E/h → "date,doy,year,n,e,u"。
window.GNSS_TS=window.GNSS_TS||{};
const GNSS_FILE_RE=/^([A-Za-z0-9]{3,6})(?:_f_all)?\.(xlsx|xls|csv|txt)$/i;
function gnssTsStatus(m){document.getElementById('gnssTsStatus').textContent=m||'';}
// Excel 序號 / Date 物件 / YYYYMMDD / YYYY-MM-DD → 'YYYY-MM-DD'
function gnssParseDate(v){
 if(v==null||v==='')return null;
 // Date 物件用「本地時間分量」取日期: SheetJS 產生的 Date 帶時區偏移,
 // 用 toISOString() (UTC) 會讓整條時序的日期軸整體偏一天
 if(v instanceof Date&&isFinite(v))
  return `${v.getFullYear()}-${String(v.getMonth()+1).padStart(2,'0')}-${String(v.getDate()).padStart(2,'0')}`;
 if(typeof v==='number'&&isFinite(v)){
  if(v>10000&&v<80000){                        // Excel 1900 日期序號
   const ms=Date.UTC(1899,11,30)+v*86400000;
   return new Date(ms).toISOString().slice(0,10);
  }
  v=String(Math.round(v));
 }
 const s=String(v).trim();
 let m=/^(\d{4})-(\d{2})-(\d{2})/.exec(s);
 if(m)return `${m[1]}-${m[2]}-${m[3]}`;
 m=/^(\d{4})(\d{2})(\d{2})$/.exec(s);
 if(m)return `${m[1]}-${m[2]}-${m[3]}`;
 m=/^(\d{4})\/(\d{1,2})\/(\d{1,2})/.exec(s);
 if(m)return `${m[1]}-${String(+m[2]).padStart(2,'0')}-${String(+m[3]).padStart(2,'0')}`;
 return null;
}
function gnssYearDoy(y,doy){
 if(!isFinite(y)||!isFinite(doy))return null;
 return new Date(Date.UTC(y,0,1)+(doy-1)*86400000).toISOString().slice(0,10);
}
// rows (二維陣列, 含表頭) + 欄位定義 → {dates, e, n, u} (皆 mm, 相對首筆)
function gnssParseRows(rows,fields,unitK,crs){
 const ix={};
 fields.forEach((f,i)=>{if(f!=='ignore')ix[f]=i;});
 const out={dates:[],e:[],n:[],u:[],lon:null,lat:null};
 const rawN=[],rawE=[],rawU=[];
 for(let r=1;r<rows.length;r++){          // 第 1 列固定視為表頭
  const row=rows[r];
  if(!row||!row.length)continue;
  let d=null;
  if(ix.date!=null)d=gnssParseDate(row[ix.date]);
  if(!d&&ix.year!=null&&ix.doy!=null)d=gnssYearDoy(+row[ix.year],+row[ix.doy]);
  if(!d)continue;
  const vn=+row[ix.n],ve=+row[ix.e],vu=(ix.u!=null)?+row[ix.u]:NaN;
  if(!isFinite(vn)||!isFinite(ve))continue;
  out.dates.push(d);rawN.push(vn);rawE.push(ve);rawU.push(vu);
 }
 if(!out.dates.length)return null;
 // 站位: 取首筆座標
 if(crs==='4326'){out.lon=rawE[0];out.lat=rawN[0];}
 else{const ll=twd97ToWgs84(rawE[0],rawN[0]);out.lon=ll[0];out.lat=ll[1];}
 // 位移: 相對首筆, 轉 mm。4326 時水平為度 → 以當地換算成公尺再轉 mm
 const degN=110574000,degE=111320000*Math.cos((out.lat||0)*Math.PI/180);  // mm/度
 for(let i=0;i<out.dates.length;i++){
  if(crs==='4326'){
   out.n.push((rawN[i]-rawN[0])*degN);
   out.e.push((rawE[i]-rawE[0])*degE);
  }else{
   out.n.push((rawN[i]-rawN[0])*unitK);
   out.e.push((rawE[i]-rawE[0])*unitK);
  }
  out.u.push(isFinite(rawU[i])?(rawU[i]-rawU[0])*unitK:null);
 }
 return out;
}
async function gnssReadFile(file,fields,unitK,crs){
 const ext=(file.name.split('.').pop()||'').toLowerCase();
 let rows;
 if(ext==='xlsx'||ext==='xls'){
  if(typeof XLSX==='undefined')throw t('gnss_sheetjs_missing');
  // 不用 cellDates: 讓日期維持 Excel 序號, 由 gnssParseDate 以 UTC 換算 (不受瀏覽器時區影響)
  const wb=XLSX.read(new Uint8Array(await file.arrayBuffer()),{type:'array'});
  rows=XLSX.utils.sheet_to_json(wb.Sheets[wb.SheetNames[0]],{header:1,raw:true,defval:null});
 }else{
  const txt=await file.text();
  rows=txt.split(/\r?\n/).filter(s=>s.trim()).map(s=>
   (s.indexOf(',')>=0?s.split(','):s.trim().split(/\s+/)).map(v=>{
    const x=v.trim();const num=+x;
    return (x!==''&&isFinite(num))?num:x;}));
 }
 return gnssParseRows(rows,fields,unitK,crs);
}
document.getElementById('gnssDirLoad').onclick=async()=>{
 const btn=document.getElementById('gnssDirLoad');
 const files=[...document.getElementById('gnssTsDir').files];
 const fields=document.getElementById('gnssFmt').value.split(',').map(s=>s.trim().toLowerCase());
 const unitK=document.getElementById('gnssUnit').value==='mm'?1:1000;
 const crs=document.getElementById('gnssCrs').value;
 // 圖上已有站位時只讀那些站; 沒有站位資料時全部讀, 讀完再依圖幅範圍篩選
 const want=new Set((D.gnss||[]).map(g=>g.n.toUpperCase()));
 const list=files.filter(f=>{
  const m=GNSS_FILE_RE.exec(f.name);
  return m&&(!want.size||want.has(m[1].toUpperCase()));});
 if(!list.length){gnssTsStatus(t('gnss_dir_none'));return;}
 btn.disabled=true;
 let ok=0;
 const g=GRID?GRID.gt:null,gW=GRID?GRID.W:0,gH=GRID?GRID.H:0;
 const inView=(lon,lat)=>!g||(lon>=g[0]&&lon<=g[0]+g[1]*gW&&lat<=g[3]&&lat>=g[3]+g[5]*gH);
 try{
  for(let i=0;i<list.length;i++){
   const sta=GNSS_FILE_RE.exec(list[i].name)[1].toUpperCase();
   gnssTsStatus(t('gnss_dir_reading',{i:i+1,n:list.length,name:list[i].name}));
   await new Promise(r=>setTimeout(r,0));
   try{
    const s=await gnssReadFile(list[i],fields,unitK,crs);
    if(!s||!inView(s.lon,s.lat))continue;
    window.GNSS_TS[sta]=s;ok++;
    // 站位未內嵌時 (預設情況), 由時序檔自己的座標補上圖層
    if(!(D.gnss||[]).some(x=>x.n.toUpperCase()===sta)){
     D.gnss=(D.gnss||[]).concat([{n:sta,lon:+s.lon.toFixed(6),lat:+s.lat.toFixed(6)}]);
    }
   }catch(err){console.error(list[i].name,err);}
  }
  buildGnssLayer();
  gnssTsStatus(t('gnss_dir_loaded',{n:ok}));
 }finally{btn.disabled=false;}
};
// GNSS 站時序 → tswin (E/N/U 三分量, mm)
function showGnssTS(sta){
 const s=window.GNSS_TS[sta];
 if(!s)return;
 clearTsPtMarker();
 document.getElementById('tstitle').textContent=t('gnss_ts_title',{sta:sta,comp:'E/N/U'});
 const fit=v=>{
  const yr=s.dates.map(d=>(Date.parse(d)-Date.parse(s.dates[0]))/(365.25*86400000));
  const g=[];for(let i=0;i<v.length;i++)if(v[i]!=null&&isFinite(v[i]))g.push(i);
  if(g.length<2)return null;
  const mt=g.reduce((a,i)=>a+yr[i],0)/g.length,my=g.reduce((a,i)=>a+v[i],0)/g.length;
  let num=0,den=0;
  for(const i of g){num+=(yr[i]-mt)*(v[i]-my);den+=(yr[i]-mt)**2;}
  return den>0?num/den:null;};
 const ve=fit(s.e),vn=fit(s.n),vu=fit(s.u);
 document.getElementById('tsinfo').textContent=
  `Ve=${ve==null?'—':ve.toFixed(1)}  Vn=${vn==null?'—':vn.toFixed(1)}  `+
  `Vu=${vu==null?'—':vu.toFixed(1)} mm/yr  (n=${s.dates.length})`;
 if(tsc)tsc.destroy();
 tsc=new Chart(document.getElementById('tschart'),{type:'line',
  data:{labels:s.dates,datasets:[
   {label:'E',data:s.e,borderColor:'#dc2626',pointRadius:0,borderWidth:1.1,spanGaps:true,pointStyle:'line'},
   {label:'N',data:s.n,borderColor:'#16a34a',pointRadius:0,borderWidth:1.1,spanGaps:true,pointStyle:'line'},
   {label:'U',data:s.u,borderColor:'#2563eb',pointRadius:0,borderWidth:1.1,spanGaps:true,pointStyle:'line'}]},
  options:gnssChartOptions()});
}
// GNSS 一律以 mm 呈現 (與圖面顯示單位 U 無關) → y 軸標題不能沿用 tsChartOptions 的 U
function gnssChartOptions(){
 const o=tsChartOptions();
 o.scales.y.title.text=t('axis_displacement',{U:'mm'});
 return o;
}
// 旋轉控制
const brg=document.getElementById('brg');
function setB(v){brg.value=((v%360)+360)%360;if(map.setBearing)map.setBearing(+brg.value);}
brg.oninput=e=>{if(map.setBearing)map.setBearing(+e.target.value);};
setB(initB);
document.getElementById('align').onclick=()=>setB(initB);
document.getElementById('north').onclick=()=>setB(0);

// 每期相對首期的年數 (畫速度趨勢線用)
const d0=new Date(D.dates[0]);
const yrs=D.dates.map(d=>(new Date(d)-d0)/(365.25*864e5));
// 時間序列圖 (點 PS): 雜訊時序(黃) + 速度趨勢線(青虛線, =顏色所反映的整段速率)
let tsc=null;
// tswin 共用圖表選項: 白底深色文字, 圖例用線樣式 (虛線 dataset 圖例同步顯示虛線)
function tsChartOptions(){
 return {responsive:true,maintainAspectRatio:false,
  plugins:{legend:{display:true,labels:{usePointStyle:true,color:'#000',boxWidth:30,font:{size:10}}}},
  scales:{x:{ticks:{color:'#000',maxTicksLimit:6}},
   y:{title:{display:true,text:t('axis_displacement',{U:U}),color:'#000'},ticks:{color:'#000'}}}};
}
let tsPtMarker=null;   // map marker highlighting the coordinate whose timeseries is shown
function clearTsPtMarker(){if(tsPtMarker){map.removeLayer(tsPtMarker);tsPtMarker=null;}}
// NOTE: charts/popups intentionally are not retranslated live on a language toggle (only
// re-render on their next trigger, e.g. clicking a point again) — see the i18n report.
function showTS(i){const p=D.points[i];
 clearTsPtMarker();
 document.getElementById('tstitle').textContent=
  t('ts_ps_title',{i:i})+(p.c!=null?t('ts_coh',{c:p.c}):'')+(p.s!=null?t('ts_along_line',{km:km(p.s).toFixed(1)}):'');
 const last=p.t[p.t.length-1];
 document.getElementById('tsinfo').textContent=
  t('ts_velocity',{sign:p.v>0?'+':'',v:p.v,U:U,dir:p.v<0?t('common_subsidence'):t('common_uplift')})+' ・ '+t('ts_cum_prefix')+
  (last!=null?`${last>0?'+':''}${last.toFixed(2)} ${U}`:'—');
 const trend=yrs.map(y=>+(p.v*y).toFixed(1));   // 首期歸零→趨勢線 = v*年數
 if(tsc)tsc.destroy();
 tsc=new Chart(document.getElementById('tschart'),{type:'line',
  data:{labels:D.dates,datasets:[
    {label:t('chart_ts_noise'),data:p.t,borderColor:'#e09112',pointRadius:2,borderWidth:1.3,tension:.15,spanGaps:true,pointStyle:'line'},
    {label:t('chart_vel_trend'),data:trend,borderColor:'#0891b2',borderDash:[6,4],borderWidth:2,pointRadius:0,pointStyle:'line'}]},
  options:tsChartOptions()});
 markers.forEach((m,k)=>m.setStyle(k===i?{stroke:true,color:'#fff',weight:2,radius:6}:{stroke:false,radius:4}));
}

// 格網/座標點時序: 資料線 + 最小二乘回歸趨勢虛線; 標題含經緯度/回歸速度/總累積
// (param is named `t` here, shadowing the i18n t() helper -- use window.t explicitly)
function showGridTS(lon,lat,t){
 clearTsPtMarker();
 tsPtMarker=L.circleMarker([lat,lon],{radius:7,color:'#000',weight:2,
  fillColor:'#ffee00',fillOpacity:.95,interactive:false}).addTo(map);
 const fit=lsqSlope(yrs,t);                 // least-squares fit over valid epochs
 let fv=null,lv=null;                        // first/last valid values → total cumulative
 for(const v of t){if(v!=null&&isFinite(v)){if(fv==null)fv=v;lv=v;}}
 const cum=(fv!=null&&lv!=null)?lv-fv:null;
 document.getElementById('tstitle').textContent=
  window.t('ts_coord',{lon:lon.toFixed(4),lat:lat.toFixed(4)});
 document.getElementById('tsinfo').textContent=
  (fit?window.t('ts_reg_vel',{sign:fit.slope>0?'+':'',v:fit.slope.toFixed(2),U:U}):window.t('ts_reg_vel_na'))+' ・ '+
  (cum!=null?window.t('ts_cum_val',{sign:cum>0?'+':'',v:cum.toFixed(2),U:U,dir:cum<0?window.t('common_subsidence'):window.t('common_uplift')}):window.t('ts_cum_na'));
 const ds=[{label:window.t('chart_ts_grid'),data:t,borderColor:'#e09112',pointRadius:2,borderWidth:1.3,tension:.15,spanGaps:true,pointStyle:'line'}];
 if(fit)ds.push({label:window.t('chart_reg_trend'),data:yrs.map(y=>+(fit.intercept+fit.slope*y).toFixed(2)),
  borderColor:'#0891b2',borderDash:[6,4],borderWidth:2,pointRadius:0,pointStyle:'line'});
 if(tsc)tsc.destroy();
 tsc=new Chart(document.getElementById('tschart'),{type:'line',
  data:{labels:D.dates,datasets:ds},
  options:tsChartOptions()});
 markers.forEach(m=>m.setStyle({stroke:false,radius:4}));
}

// 沿線剖面: 全期灰線 + 當期紅線 + y2軸頂端上色曲線, 滑桿選期
// (CLI --line 初始剖面 與 互動剖面 皆呼叫此同一函式呈現)
let pc=null,di=0,timer=null;
const slider=document.getElementById('slider');
function stopTimer(){if(timer){clearInterval(timer);timer=null;document.getElementById('play').textContent=t('prof_play');}}
function renderProfilePane(pts,getTop,limTop){
 stopTimer();
 if(pc){pc.destroy();pc=null;}
 document.getElementById('profwin').style.display='';
 slider.max=D.dates.length-1;
 const GN=Math.min(pts.length,600),gs=Math.max(1,Math.floor(pts.length/GN));
 const gidx=[];for(let i=0;i<pts.length;i+=gs)gidx.push(i);
 const grayDs=D.dates.map((d,e)=>({data:gidx.map(i=>({x:km(pts[i].s),y:pts[i].t[e]})),
   borderColor:'rgba(90,95,105,0.15)',borderWidth:0.6,pointRadius:0,tension:0,spanGaps:true,order:3}));
 const redDs={label:t('chart_current_epoch'),data:pts.map(p=>({x:km(p.s),y:p.t[0]})),
   borderColor:'#e11',borderWidth:1.8,pointRadius:0,tension:0,spanGaps:true,order:1};
 // 頂端上色曲線: CLI 剖面=沿線 PS 速度; 互動剖面=依「曲線選擇」為速率或總累積; 掛隱藏 y2 軸固定在頂端
 const topVal=pts.map(p=>getTop(p));
 const topDs={type:'scatter',yAxisID:'y2',order:0,
   data:pts.map((p,i)=>topVal[i]==null?null:{x:km(p.s),y:0.95}),
   pointBackgroundColor:topVal.map(v=>v==null?'rgba(0,0,0,0)':col(v,limTop)),pointBorderWidth:0,pointRadius:3.2};
 pc=new Chart(document.getElementById('profchart'),{type:'line',
  plugins:[cursorPlugin],                    // 剖面 hover 垂直虛線游標
  data:{datasets:[...grayDs,redDs,topDs]},
  options:{responsive:true,maintainAspectRatio:false,animation:false,normalized:true,
   parsing:true,plugins:{legend:{display:false}},elements:{point:{radius:0}},
   scales:{x:{type:'linear',title:{display:true,text:t('axis_along_dist'),color:'#333'},
     grid:{color:'#ececec'},ticks:{color:'#333'}},
    y:{title:{display:true,text:t('axis_displacement',{U:U}),color:'#333'},grid:{color:'#eee'},ticks:{color:'#333'}},
    y2:{position:'right',min:0,max:1,display:false}}}});
 const RED=grayDs.length;   // datasets 順序: [...gray, red, top] → red 在 index=grayDs.length
 function setEpoch(k){di=k;slider.value=k;
  document.getElementById('dlabel').textContent=D.dates[k];
  document.getElementById('dval').textContent=D.dates[k];
  pc.data.datasets[RED].data=pts.map(p=>({x:km(p.s),y:p.t[k]}));
  pc.update('none');}
 slider.oninput=e=>setEpoch(+e.target.value);
 document.getElementById('play').onclick=function(){
  if(timer){clearInterval(timer);timer=null;this.textContent=t('prof_play');return;}
  this.textContent=t('prof_pause');
  timer=setInterval(()=>setEpoch((di+1)%D.dates.length),600);};
 setEpoch(0);
}
function cliProfilePts(){return D.points.filter(p=>p.s!=null).slice().sort((a,b)=>a.s-b.s);}

// ---- 剖面線 hover/click 同步 ----
// 局部等距近似常數 (供 nearestStation 與 功能C polylineStations 共用)
const M_PER_DEG_LAT=111320;
const mPerDegLon=lat=>111320*Math.cos(lat*Math.PI/180);
// least-squares slope ignoring null epochs; xs in years, ys displacement (display units)
function lsqSlope(xs,ys){
 let n=0,sx=0,sy=0;
 for(let i=0;i<xs.length;i++){if(ys[i]==null||!isFinite(ys[i]))continue;n++;sx+=xs[i];sy+=ys[i];}
 if(n<2)return null;
 const mx=sx/n,my=sy/n;let num=0,den=0;
 for(let i=0;i<xs.length;i++){if(ys[i]==null||!isFinite(ys[i]))continue;
  num+=(xs[i]-mx)*(ys[i]-my);den+=(xs[i]-mx)*(xs[i]-mx);}
 return den>0?{slope:num/den,intercept:my-num/den*mx,n:n}:null;
}
// nearest station index by local-equidistant metric (m); -1 when empty
function nearestStation(stations,lat,lon){
 let best=-1,bd=Infinity;
 const ml=mPerDegLon(lat);
 for(let i=0;i<stations.length;i++){
  const dx=(stations[i].lon-lon)*ml,dy=(stations[i].lat-lat)*M_PER_DEG_LAT;
  const d=dx*dx+dy*dy;
  if(d<bd){bd=d;best=i;}}
 return best;
}
// 剖面圖游標垂直虛線 (inline Chart.js plugin, 不引外部套件)
let profCursorS=null;                       // km; null=隱藏
const cursorPlugin={id:'profCursor',afterDraw(chart){
 if(profCursorS==null)return;
 const x=chart.scales.x.getPixelForValue(profCursorS);
 if(!isFinite(x)||x<chart.chartArea.left||x>chart.chartArea.right)return;
 const c=chart.ctx;c.save();c.strokeStyle='#e11';c.setLineDash([5,4]);c.lineWidth=1;
 c.beginPath();c.moveTo(x,chart.chartArea.top);c.lineTo(x,chart.chartArea.bottom);c.stroke();c.restore();}};
let profStations=null,hitLine=null,hoverDot=null,hoverClearTimer=null;
function clearProfileHover(){
 profStations=null;profCursorS=null;
 if(hoverClearTimer){clearTimeout(hoverClearTimer);hoverClearTimer=null;}
 if(hitLine){map.removeLayer(hitLine);hitLine=null;}
 if(hoverDot){map.removeLayer(hoverDot);hoverDot=null;}
 document.getElementById('profcur').textContent='';
}
// CLI --line 剖面: 以 D.line 依取樣間距重建 station 陣列 (同 polylineStations 邏輯)
function cliStations(){
 if(!D.line||D.line.length<2)return null;
 const step=Math.max(1,+document.getElementById('sampleStep').value||50);
 const {stations,atS}=polylineStations(D.line,step);
 return stations.map(s=>{const p=atS(s);return {s:s,lon:p[0],lat:p[1]};});
}
function setupProfileHover(stations){
 clearProfileHover();
 if(!stations||stations.length<2)return;
 profStations=stations;
 // 寬觸控線: 透明 weight 20, interactive; 為 leaflet-interactive → 畫線模式時被 .drawing CSS 穿透
 hitLine=L.polyline(stations.map(p=>[p.lat,p.lon]),
  {color:'#000',opacity:0,weight:20,interactive:true,bubblingMouseEvents:false}).addTo(map);
 hitLine.on('mousemove',e=>{
  if(hoverClearTimer){clearTimeout(hoverClearTimer);hoverClearTimer=null;}
  const i=nearestStation(profStations,e.latlng.lat,e.latlng.lng);
  if(i<0)return;
  const st=profStations[i];
  if(!hoverDot)hoverDot=L.circleMarker([st.lat,st.lon],{radius:5,color:'#fff',weight:1.5,
   fillColor:'#e11',fillOpacity:1,interactive:false}).addTo(map);
  else hoverDot.setLatLng([st.lat,st.lon]);
  profCursorS=km(st.s);
  document.getElementById('profcur').textContent=t('prof_mileage',{km:km(st.s).toFixed(2)});
  if(pc)pc.update('none');
 });
 // debounce 清除: 線上方壓著 GNSS 三角形/站名 tooltip (marker/tooltip pane 在 overlay pane 之上)
 // 會對觸控線觸發短暫 mouseout, 立即清除會讓圖表游標在滑過時不斷消失 → 延遲 150ms, 回到線上就取消
 hitLine.on('mouseout',()=>{
  if(hoverClearTimer)clearTimeout(hoverClearTimer);
  hoverClearTimer=setTimeout(()=>{
   hoverClearTimer=null;
   if(hoverDot){map.removeLayer(hoverDot);hoverDot=null;}
   profCursorS=null;
   document.getElementById('profcur').textContent='';
   if(pc)pc.update('none');
  },150);
 });
 hitLine.on('click',e=>{
  if(drawing||gifRectMode||!GRID)return;
  if(Date.now()<gifSuppressUntil)return;           // 框選拖曳剛好停在觸控線上時別彈時序視窗
  const i=nearestStation(profStations,e.latlng.lat,e.latlng.lng);
  if(i<0)return;
  const st=profStations[i];
  const t=[];let any=false;
  for(let k=0;k<D.dates.length;k++){
   const v=gridSample(st.lon,st.lat,k);
   t.push(v==null?null:+v.toFixed(2));
   if(v!=null)any=true;}
  if(any)showGridTS(st.lon,st.lat,t);
 });
}

if(D.hasLine){renderProfilePane(cliProfilePts(),p=>p.v,effLims('vel'));setupProfileHover(cliStations());}
else{document.getElementById('profwin').style.display='none';}

// ---- 功能 B: 全解析度時序格網解壓 (供互動剖面雙線性取樣) ----
let GRID=null;
function setInteractiveEnabled(ok,msg){
 document.getElementById('drawBtn').disabled=!ok;
 document.getElementById('lineFile').disabled=!ok;
 document.getElementById('lineStatus').textContent=msg||'';
 const bt=document.getElementById('exportTif');
 bt.disabled=!ok;
 bt.title=ok?t('title_export_tif_ok')
            :t('title_export_tif_bad')+(msg||'');
 // 遮罩/手動色階範圍/閾值區域同樣依賴內嵌格網
 for(const id of ['maskFile','maskClear','drawMaskBtn','cbMin','cbMax','cbApply','cbAuto',
                   'thVal','thColor','thApply','thClear']){
  const el=document.getElementById(id);
  el.disabled=!ok;
  el.title=ok?'':t('title_needs_grid');
 }
}
async function initGrid(){
 if(!D.grid){setInteractiveEnabled(false,t('msg_no_grid_disabled'));return;}
 if(D.grid.arr){                                // --app 直讀模式: 已是記憶體中的 typed array
  const [N,H,W]=D.grid.shape;
  GRID={arr:D.grid.arr,N,H,W,gt:D.grid.gt,scale:D.grid.scale,nodata:D.grid.nodata};
  setInteractiveEnabled(true,'');return;}
 if(typeof DecompressionStream==='undefined'){
  setInteractiveEnabled(false,t('msg_no_decompression'));
  alert(t('alert_no_decompression'));
  return;}
 try{
  setInteractiveEnabled(false,t('msg_decompressing'));
  const bin=atob(D.grid.b64);
  const bytes=new Uint8Array(bin.length);
  for(let i=0;i<bin.length;i++)bytes[i]=bin.charCodeAt(i);
  const stream=new Blob([bytes]).stream().pipeThrough(new DecompressionStream('gzip'));
  const buf=await new Response(stream).arrayBuffer();
  const [N,H,W]=D.grid.shape;
  GRID={arr:new Int16Array(buf),N,H,W,gt:D.grid.gt,scale:D.grid.scale,nodata:D.grid.nodata};
  setInteractiveEnabled(true,'');
 }catch(err){console.error(err);setInteractiveEnabled(false,t('msg_decompress_failed')+err);}
}
setInteractiveEnabled(false,D.grid?t('msg_decompressing'):t('msg_no_grid_short'));
initGrid();

// 雙線性內插: 任一角 nodata 改用最近有效鄰點(限縮半徑內搜尋); 搜尋後四角仍全無效 → null
function gridPixel(r,c,e){
 if(r<0||r>=GRID.H||c<0||c>=GRID.W)return null;
 const v=GRID.arr[e*GRID.H*GRID.W+r*GRID.W+c];
 return v===GRID.nodata?null:v*GRID.scale;
}
function gridPixelNearest(r,c,e){
 let v=gridPixel(r,c,e);
 if(v!==null)return v;
 for(let rad=1;rad<=5;rad++)
  for(let dr=-rad;dr<=rad;dr++)
   for(let dc=-rad;dc<=rad;dc++){
    if(Math.max(Math.abs(dr),Math.abs(dc))!==rad)continue;
    v=gridPixel(r+dr,c+dc,e);
    if(v!==null)return v;}
 return null;
}
function gridSample(lon,lat,e){
 const gt=GRID.gt;
 const fc=(lon-gt[0])/gt[1]-0.5, fr=(lat-gt[3])/gt[5]-0.5;
 const c0=Math.floor(fc), r0=Math.floor(fr), fx=fc-c0, fy=fr-r0;
 const v00=gridPixelNearest(r0,c0,e), v10=gridPixelNearest(r0,c0+1,e),
       v01=gridPixelNearest(r0+1,c0,e), v11=gridPixelNearest(r0+1,c0+1,e);
 // 仍為 null 的角剔除, 有效角雙線性權重重新歸一化 (避免 nodata 角以 0 參與加權拉低取樣值)
 const vals=[v00,v10,v01,v11];
 const wts=[(1-fx)*(1-fy),fx*(1-fy),(1-fx)*fy,fx*fy];
 let sw=0,sv=0;
 for(let i=0;i<4;i++){if(vals[i]!==null){sw+=wts[i];sv+=wts[i]*vals[i];}}
 return sw>0?sv/sw:null;
}

// ---- 匯出 GeoTIFF: 手寫最小 writer (單 band float32, little-endian, 單一 strip) ----
// geo tags: ModelPixelScale(33550) / ModelTiepoint(33922, 外緣角點) /
// GeoKeyDirectory(34735: geographic + PixelIsArea + EPSG:4326) / GDAL_NODATA(42113)="nan"
function writeGeoTiff(arr,W,H,gt){
 const nTags=14;
 const ifdOff=8,ifdSize=2+nTags*12+4;
 let off=Math.ceil((ifdOff+ifdSize)/8)*8;   // pad: doubles 8-aligned, pixel data 4-aligned
 const scaleOff=off;off+=24;                // 3 doubles
 const tieOff=off;off+=48;                  // 6 doubles
 const geoOff=off;off+=32;                  // 16 shorts
 const dataOff=off;
 const buf=new ArrayBuffer(dataOff+W*H*4);
 const dv=new DataView(buf);
 dv.setUint8(0,0x49);dv.setUint8(1,0x49);   // "II" little-endian
 dv.setUint16(2,42,true);dv.setUint32(4,ifdOff,true);
 let p=ifdOff;
 dv.setUint16(p,nTags,true);p+=2;
 function tag(id,type,count,value){
  dv.setUint16(p,id,true);dv.setUint16(p+2,type,true);
  dv.setUint32(p+4,count,true);dv.setUint32(p+8,value,true);p+=12;
 }
 tag(256,3,1,W);                            // ImageWidth
 tag(257,3,1,H);                            // ImageLength
 tag(258,3,1,32);                           // BitsPerSample
 tag(259,3,1,1);                            // Compression=none
 tag(262,3,1,1);                            // Photometric=BlackIsZero
 tag(273,4,1,dataOff);                      // StripOffsets
 tag(277,3,1,1);                            // SamplesPerPixel
 tag(278,4,1,H);                            // RowsPerStrip (single strip)
 tag(279,4,1,W*H*4);                        // StripByteCounts
 tag(339,3,1,3);                            // SampleFormat=IEEE float
 tag(33550,12,3,scaleOff);                  // ModelPixelScale
 tag(33922,12,6,tieOff);                    // ModelTiepoint
 tag(34735,3,16,geoOff);                    // GeoKeyDirectory
 dv.setUint16(p,42113,true);dv.setUint16(p+2,2,true);dv.setUint32(p+4,4,true);
 dv.setUint8(p+8,110);dv.setUint8(p+9,97);dv.setUint8(p+10,110);dv.setUint8(p+11,0);p+=12; // "nan\0"
 dv.setUint32(p,0,true);                    // next IFD = none
 const wd=(o,v)=>dv.setFloat64(o,v,true);
 wd(scaleOff,gt[1]);wd(scaleOff+8,Math.abs(gt[5]));wd(scaleOff+16,0);
 [0,0,0,gt[0],gt[3],0].forEach((v,i)=>wd(tieOff+i*8,v));   // raster(0,0)=格網外緣角點
 [1,1,0,3, 1024,0,1,2, 1025,0,1,1, 2048,0,1,4326]
  .forEach((v,i)=>dv.setUint16(geoOff+i*2,v,true));
 new Float32Array(buf,dataOff,W*H).set(arr);
 return buf;
}
// 純函式: 目前圖層 → GeoTIFF ArrayBuffer (顯示單位; vel=顯示單位/yr)
function exportLayerTiff(mode){
 if(!GRID)throw t('err_grid_not_ready');
 const N=GRID.N,H=GRID.H,W=GRID.W;
 const span=datesSpanYr();
 const out=new Float32Array(W*H).fill(NaN);
 const off=(N-1)*H*W;
 for(let i=0;i<W*H;i++){
  const q0=GRID.arr[i],q1=GRID.arr[off+i];
  if(q0===GRID.nodata||q1===GRID.nodata)continue;
  out[i]=mode==='cum'?q1*GRID.scale:(q1-q0)*GRID.scale/span;
 }
 return {buf:writeGeoTiff(out,W,H,GRID.gt),
  name:`ks_${mode}_${D.dates[0]}_${D.dates[D.dates.length-1]}.tif`};
}
document.getElementById('exportTif').onclick=()=>{
 try{
  const r=exportLayerTiff(document.getElementById('layerMode').value);
  const a=document.createElement('a');
  a.href=URL.createObjectURL(new Blob([r.buf],{type:'image/tiff'}));
  a.download=r.name;a.click();
  setTimeout(()=>URL.revokeObjectURL(a.href),5000);
 }catch(err){alert(t('alert_export_tif_failed')+err);}
};

// ---- 匯出畫面 GeoTIFF: 目前視野合成 RGB 三波段 (EPSG:3857, 圖磚原生投影) ----
function writeRGBGeoTiff(rgb,W,H,gt,epsg){
 const nTags=13;
 const ifdOff=8,ifdSize=2+nTags*12+4;
 let off=Math.ceil((ifdOff+ifdSize)/8)*8;
 const bitsOff=off;off+=8;                  // 3 shorts (padded)
 const scaleOff=off;off+=24;
 const tieOff=off;off+=48;
 const geoOff=off;off+=32;
 const dataOff=off;
 const buf=new ArrayBuffer(dataOff+W*H*3);
 const dv=new DataView(buf);
 dv.setUint8(0,0x49);dv.setUint8(1,0x49);
 dv.setUint16(2,42,true);dv.setUint32(4,ifdOff,true);
 let p=ifdOff;
 dv.setUint16(p,nTags,true);p+=2;
 const tag=(id,type,count,value)=>{
  dv.setUint16(p,id,true);dv.setUint16(p+2,type,true);
  dv.setUint32(p+4,count,true);dv.setUint32(p+8,value,true);p+=12;};
 tag(256,3,1,W);
 tag(257,3,1,H);
 tag(258,3,3,bitsOff);                      // BitsPerSample [8,8,8]
 tag(259,3,1,1);
 tag(262,3,1,2);                            // Photometric=RGB
 tag(273,4,1,dataOff);
 tag(277,3,1,3);                            // SamplesPerPixel=3
 tag(278,4,1,H);
 tag(279,4,1,W*H*3);
 tag(284,3,1,1);                            // PlanarConfig=chunky
 tag(33550,12,3,scaleOff);
 tag(33922,12,6,tieOff);
 tag(34735,3,16,geoOff);
 dv.setUint32(p,0,true);
 for(let i=0;i<3;i++)dv.setUint16(bitsOff+i*2,8,true);
 const wd=(o,v)=>dv.setFloat64(o,v,true);
 wd(scaleOff,gt[1]);wd(scaleOff+8,Math.abs(gt[5]));wd(scaleOff+16,0);
 [0,0,0,gt[0],gt[3],0].forEach((v,i)=>wd(tieOff+i*8,v));
 // GTModelType=1 projected, PixelIsArea, ProjectedCSType=epsg
 [1,1,0,3, 1024,0,1,1, 1025,0,1,1, 3072,0,1,epsg]
  .forEach((v,i)=>dv.setUint16(geoOff+i*2,v,true));
 new Uint8Array(buf,dataOff,W*H*3).set(rgb);
 return buf;
}
// 純函式: 目前視野 → {buf,name,baseOk} (async: SVG 圖層需解碼)
async function exportViewTiff(){
 const brg=map.getBearing?((map.getBearing()%360)+360)%360:0;
 if(Math.abs(brg)>0.01&&Math.abs(brg-360)>0.01)throw t('err_rotate_north_first');
 const cont=map.getContainer(),rect=cont.getBoundingClientRect();
 const W=Math.round(rect.width),H=Math.round(rect.height);
 const cv=document.createElement('canvas');cv.width=W;cv.height=H;
 const ctx=cv.getContext('2d');
 ctx.fillStyle='#fff';ctx.fillRect(0,0,W,H);
 // 1. 底圖圖磚 → 先畫進獨立 canvas 驗證未被 taint (CORS 失敗時放棄底圖, 不輸出全黑)
 let baseOk=true;
 const tcv=document.createElement('canvas');tcv.width=W;tcv.height=H;
 const tctx=tcv.getContext('2d');
 tctx.fillStyle='#fff';tctx.fillRect(0,0,W,H);
 for(const img of cont.querySelectorAll('.leaflet-tile-pane img')){
  const r=img.getBoundingClientRect();
  try{tctx.drawImage(img,r.left-rect.left,r.top-rect.top,r.width,r.height);}catch(e){}
 }
 try{tctx.getImageData(0,0,1,1);ctx.drawImage(tcv,0,0);}
 catch(e){baseOk=false;}
 // 2. 變形場 raster overlay (dataURL, 依目前透明度)
 if(rasterOv&&rasterOv.getElement()){
  const el=rasterOv.getElement(),r=el.getBoundingClientRect();
  ctx.save();ctx.globalAlpha=parseFloat(el.style.opacity||'1');
  ctx.drawImage(el,r.left-rect.left,r.top-rect.top,r.width,r.height);
  ctx.restore();
 }
 // 3. 向量層 SVG (等值線/剖面線/遮罩邊界; data URL 不 taint)
 for(const svg of cont.querySelectorAll('.leaflet-overlay-pane svg')){
  const r=svg.getBoundingClientRect();
  if(!r.width||!r.height)continue;
  const clone=svg.cloneNode(true);
  clone.setAttribute('width',r.width);clone.setAttribute('height',r.height);
  const im=new Image();
  await new Promise(res=>{im.onload=res;im.onerror=res;
   im.src='data:image/svg+xml;charset=utf-8,'
    +encodeURIComponent(new XMLSerializer().serializeToString(clone));});
  try{ctx.drawImage(im,r.left-rect.left,r.top-rect.top,r.width,r.height);}catch(e){}
 }
 // 4. GNSS 三角形+站名 (CSS divIcon 無法序列化 → 依站位手動重繪)
 if(D.gnss&&D.gnss.length){
  const b=map.getBounds();
  for(const g of D.gnss){
   if(!b.contains([g.lat,g.lon]))continue;
   const q=map.latLngToContainerPoint([g.lat,g.lon]);
   ctx.beginPath();ctx.moveTo(q.x,q.y-6);ctx.lineTo(q.x-7,q.y+6);ctx.lineTo(q.x+7,q.y+6);
   ctx.closePath();ctx.fillStyle='#ffd21e';ctx.fill();
   ctx.lineWidth=1;ctx.strokeStyle='#000';ctx.stroke();
   ctx.font='10px sans-serif';ctx.textAlign='center';
   ctx.lineWidth=3;ctx.strokeStyle='#fff';ctx.strokeText(g.n,q.x,q.y-9);
   ctx.fillStyle='#111';ctx.fillText(g.n,q.x,q.y-9);
  }
 }
 // 5. RGB + EPSG:3857 geotransform (map bounds 的 mercator 座標)
 const px=ctx.getImageData(0,0,W,H).data;
 const rgb=new Uint8Array(W*H*3);
 for(let i=0;i<W*H;i++){rgb[i*3]=px[i*4];rgb[i*3+1]=px[i*4+1];rgb[i*3+2]=px[i*4+2];}
 const bb=map.getBounds();
 const nw=L.CRS.EPSG3857.project(bb.getNorthWest()),se=L.CRS.EPSG3857.project(bb.getSouthEast());
 const gt=[nw.x,(se.x-nw.x)/W,0,nw.y,0,(se.y-nw.y)/H];
 const mode=document.getElementById('layerMode').value;
 return {buf:writeRGBGeoTiff(rgb,W,H,gt,3857),
  name:`ks_view_${mode}_${map.getZoom()}.tif`,baseOk};
}
document.getElementById('exportViewTif').onclick=async()=>{
 try{
  const r=await exportViewTiff();
  if(!r.baseOk)alert(t('alert_basemap_capture_failed'));
  const a=document.createElement('a');
  a.href=URL.createObjectURL(new Blob([r.buf],{type:'image/tiff'}));
  a.download=r.name;a.click();
  setTimeout(()=>URL.revokeObjectURL(a.href),5000);
 }catch(err){alert(t('alert_export_view_failed')+err);}
};

// ---- GeoJSON 遮罩 (client 端): scanline rasterize → 工作格網 (GRID.arr) 遮罩外設 nodata ----
function geojsonPolygonRings(gj){
 const feats=gj.type==='FeatureCollection'?gj.features:(gj.type==='Feature'?[gj]:[{geometry:gj}]);
 const rings=[];
 for(const f of feats){
  const g=f.geometry||f;if(!g)continue;
  if(g.type==='Polygon')for(const r of g.coordinates)rings.push(r.map(c=>[c[0],c[1]]));
  else if(g.type==='MultiPolygon')for(const p of g.coordinates)for(const r of p)rings.push(r.map(c=>[c[0],c[1]]));
 }
 if(!rings.length)throw t('err_no_polygon');
 return rings.map(r=>{
  let rr=normalizeCoords(r);                 // |x|>360 → TWD97 轉 WGS84 (沿用既有函式)
  if(rr.length&&(rr[0][0]!==rr[rr.length-1][0]||rr[0][1]!==rr[rr.length-1][1]))rr=rr.concat([rr[0]]);
  return rr;});
}
function rasterizeMaskT(rings){              // 逐列 even-odd 交點掃描
 const W=GRID.W,H=GRID.H,gt=GRID.gt;
 const mask=new Uint8Array(W*H);
 for(let r=0;r<H;r++){
  const lat=gt[3]+(r+0.5)*gt[5];
  const xs=[];
  for(const ring of rings){
   for(let i=0;i<ring.length-1;i++){
    const y1=ring[i][1],y2=ring[i+1][1];
    if((y1<=lat&&y2>lat)||(y2<=lat&&y1>lat)){
     const t=(lat-y1)/(y2-y1);
     xs.push(ring[i][0]+t*(ring[i+1][0]-ring[i][0]));
    }
   }
  }
  xs.sort((a,b)=>a-b);
  for(let k=0;k+1<xs.length;k+=2){
   const c0=Math.max(0,Math.ceil((xs[k]-gt[0])/gt[1]-0.5));
   const c1=Math.min(W-1,Math.floor((xs[k+1]-gt[0])/gt[1]-0.5));
   for(let c=c0;c<=c1;c++)mask[r*W+c]=1;
  }
 }
 return mask;
}
let maskBoundary=null;                       // 手繪遮罩邊界 (虛線多邊形)
function clearMaskBoundary(){
 if(maskBoundary){map.removeLayer(maskBoundary);maskBoundary=null;}
}
function applyMask(rings){
 if(!GRID)throw t('err_grid_not_ready');
 clearMaskBoundary();                        // 換新遮罩時移除舊邊界 (手繪路徑事後再畫新的)
 const mask=rasterizeMaskT(rings);
 if(!origArr)origArr=GRID.arr.slice();       // 保留原始副本供清除還原
 else GRID.arr.set(origArr);                 // 重套時從原始資料開始
 const n=GRID.W*GRID.H;
 let kept=0;
 for(let i=0;i<n;i++){
  if(mask[i]){if(GRID.arr[(GRID.N-1)*n+i]!==GRID.nodata)kept++;continue;}
  for(let e=0;e<GRID.N;e++)GRID.arr[e*n+i]=GRID.nodata;
 }
 maskApplied=true;
 refreshDerived();
 document.getElementById('lineStatus').textContent=t('msg_mask_applied',{n:kept});
}
function clearMask(){
 if(origArr)GRID.arr.set(origArr);
 clearMaskBoundary();
 maskApplied=false;
 refreshDerived();
 document.getElementById('lineStatus').textContent=t('msg_mask_cleared');
}
document.getElementById('maskFile').onchange=async e=>{
 const file=e.target.files[0];
 if(!file)return;
 try{applyMask(geojsonPolygonRings(JSON.parse(await file.text())));}
 catch(err){console.error(err);document.getElementById('lineStatus').textContent=t('msg_mask_failed')+err;}
 e.target.value='';
};
document.getElementById('maskClear').onclick=clearMask;

// ---- 手動畫遮罩多邊形: 單擊加頂點(即時預覽), 雙擊閉合→applyMask, ESC/再按取消 ----
let maskDrawing=false,maskVerts=[],maskTempLine=null;
function cancelMaskDraw(){
 maskDrawing=false;maskVerts=[];
 if(maskTempLine){map.removeLayer(maskTempLine);maskTempLine=null;}
 map.getContainer().classList.remove('drawing');
 map.getContainer().style.cursor='';
 map.doubleClickZoom.enable();
 document.getElementById('drawMaskBtn').textContent=t('mask_draw_btn');
}
document.getElementById('drawMaskBtn').onclick=()=>{
 if(maskDrawing){cancelMaskDraw();return;}
 if(drawing)cancelDraw();                    // 互斥: 退出畫線
 if(gifRectMode)endGifRectMode();            // 互斥: 退出 GIF 框選
 maskDrawing=true;maskVerts=[];
 map.doubleClickZoom.disable();
 map.getContainer().classList.add('drawing');   // 沿用穿透 CSS: marker 不擋點擊
 map.getContainer().style.cursor='crosshair';
 document.getElementById('drawMaskBtn').textContent=t('mask_drawing_btn');
 document.getElementById('lineStatus').textContent=t('msg_click_add_vertex');
};
map.on('click',e=>{
 if(!maskDrawing)return;
 maskVerts.push([e.latlng.lng,e.latlng.lat]);
 if(maskTempLine)map.removeLayer(maskTempLine);
 maskTempLine=L.polyline(maskVerts.map(v=>[v[1],v[0]]),
  {color:'#e11',weight:2,dashArray:'6 4'}).addTo(map);
});
map.on('dblclick',()=>{
 if(!maskDrawing)return;
 const verts=maskVerts.slice();
 cancelMaskDraw();
 if(verts.length<3){
  document.getElementById('lineStatus').textContent=t('msg_mask_not_enough_vertex');
  return;
 }
 try{
  applyMask([verts.concat([verts[0]])]);     // 閉合 ring (WGS84)
  maskBoundary=L.polygon(verts.map(v=>[v[1],v[0]]),
   {color:'#e11',weight:2,dashArray:'6 4',fill:false,interactive:false}).addTo(map);
 }catch(err){document.getElementById('lineStatus').textContent=t('msg_mask_failed')+err;}
});
document.addEventListener('keydown',e=>{if(e.key==='Escape'&&maskDrawing)cancelMaskDraw();});

// ---- colorbar 手動範圍 ----
document.getElementById('cbApply').onclick=()=>{
 if(!GRID)return;
 const mn=parseFloat(document.getElementById('cbMin').value);
 const mx=parseFloat(document.getElementById('cbMax').value);
 if(!isFinite(mn)||!isFinite(mx)||mn>=mx)return;   // 驗證: 非數字忽略, 需 最小<最大
 const mode=document.getElementById('layerMode').value;
 manualLims[mode]={neg:mn<0?-mn:1e-6,pos:mx>0?mx:1e-6};
 refreshDerived();
};
document.getElementById('cbAuto').onclick=()=>{
 if(!GRID)return;
 manualLims[document.getElementById('layerMode').value]=null;
 refreshDerived();
};

// ---- 功能 C: 互動剖面 (手動畫線 / GeoJSON / SHP 匯入 → 沿線取樣) ----
// (M_PER_DEG_LAT / mPerDegLon 已於剖面 hover 區塊宣告)

// TWD97(EPSG:3826) → WGS84 逆解橫麥卡托 (GRS80, 中央經線121E, k0=0.9999, 假東距250000, 假北距0)
function twd97ToWgs84(x,y){
 const a=6378137.0, f=1/298.257222101, e2=f*(2-f), k0=0.9999, fx=250000, fy=0;
 const lon0=121*Math.PI/180;
 const M=(y-fy)/k0;
 const e1=(1-Math.sqrt(1-e2))/(1+Math.sqrt(1-e2));
 const mu=M/(a*(1-e2/4-3*e2*e2/64-5*e2*e2*e2/256));
 const phi1=mu+(3*e1/2-27*Math.pow(e1,3)/32)*Math.sin(2*mu)
   +(21*e1*e1/16-55*Math.pow(e1,4)/32)*Math.sin(4*mu)
   +(151*Math.pow(e1,3)/96)*Math.sin(6*mu)
   +(1097*Math.pow(e1,4)/512)*Math.sin(8*mu);
 const e2p=e2/(1-e2), C1=e2p*Math.cos(phi1)**2, T1=Math.tan(phi1)**2;
 const N1=a/Math.sqrt(1-e2*Math.sin(phi1)**2), R1=a*(1-e2)/Math.pow(1-e2*Math.sin(phi1)**2,1.5);
 const Dd=(x-fx)/(N1*k0);
 const lat=phi1-(N1*Math.tan(phi1)/R1)*(Dd*Dd/2-(5+3*T1+10*C1-4*C1*C1-9*e2p)*Math.pow(Dd,4)/24
   +(61+90*T1+298*C1+45*T1*T1-252*e2p-3*C1*C1)*Math.pow(Dd,6)/720);
 const lon=lon0+(Dd-(1+2*T1+C1)*Math.pow(Dd,3)/6
   +(5-2*C1+28*T1-3*C1*C1+8*e2p+24*T1*T1)*Math.pow(Dd,5)/120)/Math.cos(phi1);
 return [lon*180/Math.PI, lat*180/Math.PI];
}
function normalizeCoords(verts){
 const needTM=verts.some(v=>Math.abs(v[0])>360||Math.abs(v[1])>360);
 return needTM?verts.map(v=>twd97ToWgs84(v[0],v[1])):verts;
}

// 折線里程(m)工具: 局部等距近似 (m/deg_lat=111320, m/deg_lon=111320*cos(lat))
function polylineStations(verts,stepM){
 const segs=[];let total=0;
 for(let i=0;i<verts.length-1;i++){
  const [lo1,la1]=verts[i],[lo2,la2]=verts[i+1];
  const latm=(la1+la2)/2;
  const dx=(lo2-lo1)*mPerDegLon(latm), dy=(la2-la1)*M_PER_DEG_LAT;
  const len=Math.hypot(dx,dy);
  segs.push({lo1,la1,lo2,la2,len,cum0:total});
  total+=len;
 }
 const stations=[];
 for(let s=0;s<total;s+=stepM)stations.push(s);
 if(stations.length===0||stations[stations.length-1]<total-1e-6)stations.push(total);
 function atS(s){
  let seg=segs[segs.length-1];
  for(const sg of segs){if(s<=sg.cum0+sg.len+1e-9){seg=sg;break;}}
  const t=seg.len>0?(s-seg.cum0)/seg.len:0;
  return [seg.lo1+(seg.lo2-seg.lo1)*t, seg.la1+(seg.la2-seg.la1)*t];
 }
 return {total,stations,atS};
}
function buildProfileFromLine(verts,stepM){
 const {stations,atS}=polylineStations(verts,stepM);
 const N=D.dates.length;
 return stations.map(s=>{
  const [lon,lat]=atS(s);
  const t=[];for(let e=0;e<N;e++)t.push(gridSample(lon,lat,e));
  return {s,lon,lat,t};
 });
}

// GeoJSON: 支援 LineString / MultiLineString (各段依序接續, 里程連續累計)
function geojsonToVerts(gj){
 const feats=gj.type==='FeatureCollection'?gj.features:(gj.type==='Feature'?[gj]:[{geometry:gj}]);
 const verts=[];
 for(const f of feats){
  const g=f.geometry||f; if(!g)continue;
  if(g.type==='LineString'){for(const c of g.coordinates)verts.push([c[0],c[1]]);}
  else if(g.type==='MultiLineString'){for(const line of g.coordinates)for(const c of line)verts.push([c[0],c[1]]);}
 }
 if(verts.length<2)throw t('err_no_linestring');
 return verts;
}
// SHP 手寫解析器: 僅支援 shape type 3/13/23 (PolyLine/Z/M); 多 part 依序接續同 MultiLineString 處理
function parseShpToVerts(buf){
 const dv=new DataView(buf);
 const shapeType=dv.getInt32(32,true);
 if(![3,13,23].includes(shapeType))throw t('err_unsupported_shp',{t:shapeType});
 const len=buf.byteLength;
 const verts=[];
 let off=100;
 while(off+8<=len){
  const contentWords=dv.getInt32(off+4,false);        // record header: content length(BE, 16-bit words)
  const contentLen=contentWords*2;
  let p=off+8;
  const recEnd=p+contentLen;
  const rType=dv.getInt32(p,true); p+=4;
  if(rType!==0){
   p+=32;                                              // bbox: 4 doubles
   const numParts=dv.getInt32(p,true); p+=4;
   const numPoints=dv.getInt32(p,true); p+=4;
   p+=numParts*4;                                       // parts index(不需拆段, 依序接續即可)
   for(let i=0;i<numPoints;i++){
    const x=dv.getFloat64(p,true); p+=8;
    const y=dv.getFloat64(p,true); p+=8;
    verts.push([x,y]);
   }
  }
  off=recEnd;
 }
 if(verts.length<2)throw t('err_shp_no_coords');
 return verts;
}

const interactiveLineLayer=L.layerGroup().addTo(map);
let activeInteractivePts=null;
let interactiveLine=null;                          // 互動剖面線 handle (不含起訖點圓標記, 換色時只動線本身)
function drawInteractiveLine(verts){
 interactiveLineLayer.clearLayers();
 const ll=verts.map(v=>[v[1],v[0]]);
 interactiveLine=L.polyline(ll,{color:profColor,weight:2,opacity:.9,dashArray:'5 4'}).addTo(interactiveLineLayer);
 const totalKm=(polylineStations(verts,1e9).total/1000).toFixed(1);
 L.circleMarker(ll[0],{radius:6,color:'#fff',weight:1.5,fillColor:'#22c55e',fillOpacity:1})
  .addTo(interactiveLineLayer).bindTooltip(t('prof_start'),{permanent:true,direction:'right',className:'muted'});
 L.circleMarker(ll[ll.length-1],{radius:6,color:'#fff',weight:1.5,fillColor:'#ef4444',fillOpacity:1})
  .addTo(interactiveLineLayer).bindTooltip(t('prof_end',{km:totalKm}),{permanent:true,direction:'left'});
}
function applyInteractiveCurve(){
 if(!activeInteractivePts)return;
 const mode=document.getElementById('curveMode').value;
 const N=D.dates.length;
 // 頂端色帶必須與地圖圖層同源, 否則同一個位置在圖上是綠色、色帶上卻是橘色:
 //  - 速率: 用全期最小二乘回歸 (與速度場圖層同式), 不是首尾期差分
 //    (首期常是全零參考期, 差分對含雜訊的時序會高估數倍)
 //  - 總累積: 末期 − 首期 (與 cum 圖層同式), 不是末期絕對值
 function getTop(p){
  if(mode==='rate'){const f=lsqSlope(yrs,p.t);return f?f.slope:null;}
  const first=p.t[0],last=p.t[N-1];
  return(first==null||last==null)?null:last-first;
 }
 // 色階同樣取生效中的 (手動色階/遮罩後自動色階), 不是頁面產生時的 D.lims
 renderProfilePane(activeInteractivePts,getTop,mode==='rate'?effLims('vel'):effLims('cum'));
}
document.getElementById('curveMode').onchange=applyInteractiveCurve;
function finishLine(verts){
 if(verts.length<2){document.getElementById('lineStatus').textContent=t('msg_not_enough_line_pts');return;}
 clearGifRange();                                  // 換線了, 舊的 GIF 框選里程區間不再適用
 drawInteractiveLine(verts);
 const step=Math.max(1,+document.getElementById('sampleStep').value||50);
 document.getElementById('lineStatus').textContent=t('msg_sampling');
 setTimeout(()=>{
  activeInteractivePts=buildProfileFromLine(verts,step);
  applyInteractiveCurve();
  setupProfileHover(activeInteractivePts);   // 互動剖面: station 直接用取樣結果
  document.getElementById('lineStatus').textContent=t('msg_sampled_n',{n:activeInteractivePts.length});
 },10);
}

// 手動畫線: 進入模式後地圖單擊加點(即時畫暫時折線), 雙擊結束
let drawing=false, drawVerts=[], tempLine=null;
const drawBtn=document.getElementById('drawBtn');
function cancelDraw(){
 drawing=false; drawVerts=[];
 if(tempLine){map.removeLayer(tempLine); tempLine=null;}
 map.getContainer().classList.remove('drawing');   // 唯一移除點: 所有退出路徑都經過 cancelDraw
 map.getContainer().style.cursor='';
 map.doubleClickZoom.enable();
 drawBtn.textContent=t('line_draw_btn');
}
drawBtn.onclick=()=>{
 if(drawing){cancelDraw();return;}
 if(typeof maskDrawing!=='undefined'&&maskDrawing)cancelMaskDraw();   // 互斥: 退出畫遮罩
 if(gifRectMode)endGifRectMode();                  // 互斥: 退出 GIF 框選
 drawing=true; drawVerts=[];
 map.doubleClickZoom.disable();
 map.getContainer().classList.add('drawing');      // 讓 marker/tooltip/PS 點穿透, 見 CSS .drawing 規則
 map.getContainer().style.cursor='crosshair';
 drawBtn.textContent=t('line_drawing_btn');
 document.getElementById('lineStatus').textContent=t('msg_click_add_point');
};
map.on('click',e=>{
 if(!drawing)return;
 drawVerts.push([e.latlng.lng,e.latlng.lat]);
 if(tempLine)map.removeLayer(tempLine);
 tempLine=L.polyline(drawVerts.map(v=>[v[1],v[0]]),{color:profColor,weight:2,dashArray:'5 4'}).addTo(map);
});
// 剖面線選色器: 更新單一來源 profColor (供之後新畫的線用) + 就地重繪目前地圖上的三種線
document.getElementById('profLineColor').oninput=e=>{
 profColor=e.target.value;
 if(cliLine)cliLine.setStyle({color:profColor});
 if(interactiveLine)interactiveLine.setStyle({color:profColor});
 if(tempLine)tempLine.setStyle({color:profColor});
};
map.on('dblclick',()=>{
 if(!drawing)return;
 const verts=drawVerts.slice();
 cancelDraw();
 if(verts.length>=2)finishLine(verts);
});

// ============ 剖面播放輸出 GIF ============
// 地圖上拖曳矩形 → 取剖面線落在框內的里程區間 → 把該區間的播放(第一期→最後一期)
// 逐幀擷取剖面圖 canvas, 用內建 GIF89a 編碼器打包下載 (不依賴任何外部函式庫)

// ---- 依賴無關的 GIF89a 編碼器 (全域調色盤 median-cut + LZW) ----
function _gifWriter(){
 let buf=new Uint8Array(1<<16),len=0;
 const need=n=>{if(len+n<=buf.length)return;let cap=buf.length;
  while(cap<len+n)cap*=2;const nb=new Uint8Array(cap);nb.set(buf.subarray(0,len));buf=nb;};
 return {byte(b){need(1);buf[len++]=b&255;},
  bytes(a){need(a.length);buf.set(a,len);len+=a.length;},
  u16(v){need(2);buf[len++]=v&255;buf[len++]=(v>>8)&255;},
  str(s){need(s.length);for(let i=0;i<s.length;i++)buf[len++]=s.charCodeAt(i);},
  done(){return buf.slice(0,len);}};
}
// 5-5-5 直方圖 → median-cut 調色盤 + 32768 格最近色查表
function gifBuildPalette(frames,maxColors){
 const cnt=new Uint32Array(32768);
 const sr=new Float64Array(32768),sg=new Float64Array(32768),sb=new Float64Array(32768);
 for(const f of frames){
  for(let i=0;i<f.length;i+=4){
   const r=f[i],g=f[i+1],b=f[i+2];
   const k=((r>>3)<<10)|((g>>3)<<5)|(b>>3);
   cnt[k]++;sr[k]+=r;sg[k]+=g;sb[k]+=b;}}
 const bins=[];
 for(let k=0;k<32768;k++){if(!cnt[k])continue;
  bins.push({r:(k>>10)&31,g:(k>>5)&31,b:k&31,n:cnt[k],
   ar:sr[k]/cnt[k],ag:sg[k]/cnt[k],ab:sb[k]/cnt[k]});}
 const mkBox=list=>{
  let n=0,r0=99,r1=-1,g0=99,g1=-1,b0=99,b1=-1;
  for(const p of list){n+=p.n;
   if(p.r<r0)r0=p.r;if(p.r>r1)r1=p.r;
   if(p.g<g0)g0=p.g;if(p.g>g1)g1=p.g;
   if(p.b<b0)b0=p.b;if(p.b>b1)b1=p.b;}
  return {list:list,n:n,ext:Math.max(r1-r0,g1-g0,b1-b0),
   axis:(r1-r0)>=(g1-g0)&&(r1-r0)>=(b1-b0)?'r':(g1-g0)>=(b1-b0)?'g':'b'};};
 let boxes=[mkBox(bins)];
 while(boxes.length<maxColors){
  let bi=-1,best=0;
  for(let i=0;i<boxes.length;i++){const bx=boxes[i];
   if(bx.list.length<2||bx.ext<1)continue;
   const score=bx.n*(bx.ext+1);
   if(score>best){best=score;bi=i;}}
  if(bi<0)break;
  const bx=boxes[bi],ax=bx.axis;
  const sorted=bx.list.slice().sort((p,q)=>p[ax]-q[ax]);
  let half=bx.n/2,acc=0,cut=0;
  for(;cut<sorted.length-1;cut++){acc+=sorted[cut].n;if(acc>=half)break;}
  // 單一 bin 佔壓倒性多數時(例如剖面圖的大片白底), 中位數會落在最後一個元素,
  // 右半變空 → 整個切割停擺、調色盤退化成 1 色, 故把切點夾到倒數第二個
  if(cut>sorted.length-2)cut=sorted.length-2;
  const left=sorted.slice(0,cut+1),right=sorted.slice(cut+1);
  if(!left.length||!right.length){bx.ext=0;continue;}
  boxes.splice(bi,1,mkBox(left),mkBox(right));}
 const pal=boxes.map(bx=>{let n=0,r=0,g=0,b=0;
  for(const p of bx.list){n+=p.n;r+=p.ar*p.n;g+=p.ag*p.n;b+=p.ab*p.n;}
  return [Math.round(r/n),Math.round(g/n),Math.round(b/n)];});
 const lut=new Uint8Array(32768);
 for(let k=0;k<32768;k++){
  const r=(((k>>10)&31)<<3)|4,g=(((k>>5)&31)<<3)|4,b=((k&31)<<3)|4;
  let bestI=0,bestD=Infinity;
  for(let i=0;i<pal.length;i++){
   const dr=r-pal[i][0],dg=g-pal[i][1],db=b-pal[i][2];
   const d=dr*dr+dg*dg+db*db;
   if(d<bestD){bestD=d;bestI=i;}}
  lut[k]=bestI;}
 return {pal:pal,lut:lut};
}
function gifLZW(px,minCodeSize){
 const clear=1<<minCodeSize,eoi=clear+1;
 let codeSize=minCodeSize+1,next=clear+2,dict=new Map();
 const out=[];let cur=0,nbits=0;
 const emit=c=>{cur|=c<<nbits;nbits+=codeSize;
  while(nbits>=8){out.push(cur&255);cur>>>=8;nbits-=8;}};
 emit(clear);
 let prefix=px[0];
 for(let i=1;i<px.length;i++){
  const k=px[i],key=(prefix<<8)|k;
  const hit=dict.get(key);
  if(hit!==undefined){prefix=hit;continue;}
  emit(prefix);
  if(next<4096){dict.set(key,next++);
   if(next>(1<<codeSize)&&codeSize<12)codeSize++;}
  else{emit(clear);dict=new Map();codeSize=minCodeSize+1;next=clear+2;}
  prefix=k;}
 emit(prefix);emit(eoi);
 if(nbits>0)out.push(cur&255);
 return out;
}
// frames: RGBA 位元組陣列; delayCs: 每幀延遲 (1/100 秒)
function gifEncode(frames,w,h,delayCs,maxColors){
 maxColors=Math.max(4,Math.min(256,maxColors||256));
 const pl=gifBuildPalette(frames,maxColors),pal=pl.pal,lut=pl.lut;
 let bits=1;while((1<<bits)<pal.length)bits++;
 if(bits<2)bits=2;                                 // GIF 最小 code size 為 2
 const tableSize=1<<bits,W=_gifWriter();
 W.str('GIF89a');W.u16(w);W.u16(h);
 W.byte(0xF0|(bits-1));W.byte(0);W.byte(0);
 for(let i=0;i<tableSize;i++){const c=pal[i]||[0,0,0];W.byte(c[0]);W.byte(c[1]);W.byte(c[2]);}
 W.byte(0x21);W.byte(0xFF);W.byte(0x0B);W.str('NETSCAPE2.0');   // 無限循環
 W.byte(0x03);W.byte(0x01);W.u16(0);W.byte(0x00);
 const idx=new Uint8Array(w*h);
 for(const f of frames){
  for(let p=0,q=0;p<idx.length;p++,q+=4)
   idx[p]=lut[((f[q]>>3)<<10)|((f[q+1]>>3)<<5)|(f[q+2]>>3)];
  W.byte(0x21);W.byte(0xF9);W.byte(0x04);W.byte(1<<2);
  W.u16(Math.max(1,delayCs|0));W.byte(0);W.byte(0);
  W.byte(0x2C);W.u16(0);W.u16(0);W.u16(w);W.u16(h);W.byte(0);
  W.byte(bits);
  const data=gifLZW(idx,bits);
  for(let o=0;o<data.length;o+=255){
   const n=Math.min(255,data.length-o);
   W.byte(n);W.bytes(data.slice(o,o+n));}
  W.byte(0);}
 W.byte(0x3B);
 return W.done();
}

// ==================== AOI 框選 → PNG (含框外 colorbar) / 數值 GeoTIFF ====================
// AOI 以地圖拖曳矩形決定, 對齊到格網像元邊界; 數值一律取自 currentFieldT (與畫面同源)。
let aoiMode=false,aoiStart=null,aoiRect=null,aoiBounds=null;
const aoiDrawBtn=document.getElementById('aoiDrawBtn');
const aoiStatusEl=document.getElementById('aoiStatus');
function aoiStatus(m){aoiStatusEl.textContent=m||'';}
function endAoiMode(){
 aoiMode=false;aoiStart=null;
 map.dragging.enable();
 map.getContainer().style.cursor='';
 aoiDrawBtn.textContent=t('aoi_draw_btn');
}
function clearAoi(){
 if(aoiRect){map.removeLayer(aoiRect);aoiRect=null;}
 aoiBounds=null;aoiStatus(t('aoi_none'));
}
aoiDrawBtn.onclick=()=>{
 if(aoiMode){endAoiMode();aoiStatus('');return;}
 if(drawing)cancelDraw();
 if(maskDrawing)cancelMaskDraw();
 aoiMode=true;aoiStart=null;
 map.dragging.disable();
 map.getContainer().style.cursor='crosshair';
 aoiDrawBtn.textContent=t('aoi_clear_btn');
 aoiStatus(t('aoi_hint'));
};
document.getElementById('aoiClearBtn').onclick=()=>{if(aoiMode)endAoiMode();clearAoi();};
function aoiDocUp(){setTimeout(()=>{if(aoiMode){endAoiMode();aoiStatus('');}},0);}
map.on('mousedown',e=>{
 if(!aoiMode)return;
 aoiStart=e.latlng;clearAoi();aoiMode=true;      // clearAoi 會清狀態字, 這裡保持模式
 document.addEventListener('mouseup',aoiDocUp,{once:true});
});
map.on('mousemove',e=>{
 if(!aoiMode||!aoiStart)return;
 const b=L.latLngBounds(aoiStart,e.latlng);
 if(aoiRect)aoiRect.setBounds(b);
 else aoiRect=L.rectangle(b,{color:'#111',weight:1.5,dashArray:'5 3',
  fillColor:'#111',fillOpacity:.04,interactive:false}).addTo(map);
});
map.on('mouseup',e=>{
 if(!aoiMode||!aoiStart)return;
 const b=L.latLngBounds(aoiStart,e.latlng);
 endAoiMode();
 gifSuppressUntil=Date.now()+400;                // 別讓框選那一下順便開時序視窗
 aoiBounds=b;
 const s=aoiSubset(document.getElementById('layerMode').value);
 aoiStatus(s?t('aoi_set',{w:s.W,h:s.H}):t('aoi_empty'));
});
// AOI → 子集場 {arr,W,H,gt}; 未框選時回傳整張圖
function aoiSubset(mode){
 if(!GRID)return null;
 const f=currentFieldT(mode);
 if(!f)return null;
 const gt=GRID.gt,W=GRID.W,H=GRID.H;
 let c0=0,c1=W,r0=0,r1=H;
 if(aoiBounds){
  const lo0=aoiBounds.getWest(),lo1=aoiBounds.getEast(),
        la0=aoiBounds.getSouth(),la1=aoiBounds.getNorth();
  c0=Math.max(0,Math.floor((lo0-gt[0])/gt[1]));
  c1=Math.min(W,Math.ceil((lo1-gt[0])/gt[1]));
  r0=Math.max(0,Math.floor((la1-gt[3])/gt[5]));   // gt[5]<0: 北緯 → 小 row
  r1=Math.min(H,Math.ceil((la0-gt[3])/gt[5]));
 }
 const Wa=c1-c0,Ha=r1-r0;
 if(!(Wa>0&&Ha>0))return null;
 const arr=new Float32Array(Wa*Ha).fill(NaN);
 let nval=0;
 for(let r=0;r<Ha;r++)for(let c=0;c<Wa;c++){
  const v=f[(r0+r)*W+(c0+c)];
  arr[r*Wa+c]=v;
  if(isFinite(v))nval++;
 }
 if(!nval)return null;
 return {arr,W:Wa,H:Ha,gt:[gt[0]+c0*gt[1],gt[1],0,gt[3]+r0*gt[5],0,gt[5]],n:nval};
}
function aoiFileTag(mode){
 const d0=D.dates[0],d1=D.dates[D.dates.length-1];
 return `${mode}_${d0}_${d1}${aoiBounds?'_aoi':''}`;
}
document.getElementById('aoiTifBtn').onclick=()=>{
 try{
  const mode=document.getElementById('layerMode').value;
  const s=aoiSubset(mode);
  if(!s){aoiStatus(t('aoi_empty'));return;}
  const buf=writeGeoTiff(s.arr,s.W,s.H,s.gt);
  const a=document.createElement('a');
  a.href=URL.createObjectURL(new Blob([buf],{type:'image/tiff'}));
  a.download=`insar_${aoiFileTag(mode)}.tif`;a.click();
  setTimeout(()=>URL.revokeObjectURL(a.href),5000);
  aoiStatus(t('aoi_set',{w:s.W,h:s.H}));
 }catch(err){console.error(err);alert(t('alert_export_tif_failed')+err);}
};
// AOI PNG: 左為資料影像, 右側 (影像框外) 為 colorbar 與刻度, 上方標題, 下方座標範圍
function drawAoiPNG(mode){
 const s=aoiSubset(mode);
 if(!s)return null;
 const lims=effLims(mode);
 const MAXW=1100,MAXH=900;
 const z=Math.max(1,Math.min(Math.floor(MAXW/s.W),Math.floor(MAXH/s.H)))||1;
 const iw=s.W*z,ih=s.H*z;
 const padL=14,padT=48,padB=40,cbW=104;
 const cv=document.createElement('canvas');
 cv.width=padL+iw+cbW+16;cv.height=padT+ih+padB;
 const ctx=cv.getContext('2d');
 ctx.fillStyle='#ffffff';ctx.fillRect(0,0,cv.width,cv.height);
 // 影像
 const off=document.createElement('canvas');off.width=s.W;off.height=s.H;
 const octx=off.getContext('2d'),img=octx.createImageData(s.W,s.H),dd=img.data;
 for(let i=0;i<s.W*s.H;i++){
  const v=s.arr[i];
  if(!isFinite(v))continue;
  // col() 回傳 CSS 字串, ImageData 需要數值 → 用 lerpRGB (同一組 stops)
  const rgb=v>=0?lerpRGB(POS_STOPS,Math.min(1,v/lims.pos)):lerpRGB(NEG_STOPS,Math.min(1,-v/lims.neg));
  dd[i*4]=rgb[0];dd[i*4+1]=rgb[1];dd[i*4+2]=rgb[2];dd[i*4+3]=255;
 }
 octx.putImageData(img,0,0);
 ctx.imageSmoothingEnabled=false;
 ctx.drawImage(off,0,0,s.W,s.H,padL,padT,iw,ih);
 ctx.strokeStyle='#888';ctx.lineWidth=1;
 ctx.strokeRect(padL+0.5,padT+0.5,iw,ih);
 // 標題
 const unit=mode==='vel'?`${U}/yr`:U;
 ctx.fillStyle='#222';ctx.font='bold 17px sans-serif';ctx.textAlign='left';
 const title=document.getElementById('titleTxt').textContent+' — '+
  (mode==='vel'?t('cbar_opt_vel'):t('cbar_opt_cum'))+` (${unit})`;
 ctx.fillText(title,padL,26);
 ctx.font='12px sans-serif';ctx.fillStyle='#555';
 ctx.fillText(`${D.dates[0]} ~ ${D.dates[D.dates.length-1]}`,padL,42);
 // colorbar (影像框外, 右側; 高度上限 460px 並垂直置中, 免得細長到看不出分級)
 const cbBarW=20,cbH=Math.min(ih,460),cbX=padL+iw+18,cbY=padT+Math.round((ih-cbH)/2);
 const grad=ctx.createLinearGradient(0,cbY,0,cbY+cbH);
 for(let k=0;k<=40;k++){
  const frac=k/40;                                  // 0=頂端(正/抬升) → 1=底(負/下沉)
  const v=lims.pos-(lims.pos+lims.neg)*frac;
  grad.addColorStop(frac,col(v,lims));              // col() 已是 CSS 'rgb(...)' 字串
 }
 ctx.fillStyle=grad;ctx.fillRect(cbX,cbY,cbBarW,cbH);
 ctx.strokeStyle='#888';ctx.strokeRect(cbX+0.5,cbY+0.5,cbBarW,cbH);
 ctx.fillStyle='#222';ctx.font='11px sans-serif';ctx.textAlign='left';
 const nTick=6;
 for(let k=0;k<=nTick;k++){
  const frac=k/nTick,y=cbY+frac*cbH;
  const v=lims.pos-(lims.pos+lims.neg)*frac;
  ctx.beginPath();ctx.moveTo(cbX+cbBarW,y);ctx.lineTo(cbX+cbBarW+4,y);ctx.stroke();
  ctx.fillText(v.toFixed(U==='cm'?1:0),cbX+cbBarW+7,y+4);
 }
 ctx.fillText(unit,cbX,cbY-8);
 // 座標範圍
 const g=s.gt,lo0=g[0],lo1=g[0]+g[1]*s.W,la1=g[3],la0=g[3]+g[5]*s.H;
 ctx.fillStyle='#555';ctx.font='11px sans-serif';
 ctx.fillText(`${lo0.toFixed(4)}E ~ ${lo1.toFixed(4)}E, ${la0.toFixed(4)}N ~ ${la1.toFixed(4)}N`
  +`  |  ${s.W}×${s.H} px, ${Math.abs(g[1]).toFixed(5)}°`,padL,padT+ih+22);
 return {cv,s};
}
document.getElementById('aoiPngBtn').onclick=()=>{
 try{
  const mode=document.getElementById('layerMode').value;
  const r=drawAoiPNG(mode);
  if(!r){aoiStatus(t('aoi_empty'));return;}
  r.cv.toBlob(b=>{
   const a=document.createElement('a');
   a.href=URL.createObjectURL(b);
   a.download=`insar_${aoiFileTag(mode)}.png`;a.click();
   setTimeout(()=>URL.revokeObjectURL(a.href),5000);
  },'image/png');
  aoiStatus(t('aoi_set',{w:r.s.W,h:r.s.H}));
 }catch(err){console.error(err);alert(t('alert_export_failed')+err);}
};

// ==================== 點時間序列 → PNG ====================
// 直接沿用目前 tswin 圖表的資料, 以離屏 Chart 重繪成高解析靜態圖 (含標題與速度摘要)
function drawTsPNG(){
 if(!tsc)return null;
 const Wpx=Math.max(400,Math.min(4000,+document.getElementById('tsPngW').value||1200));
 const Hpx=Math.round(Wpx*0.5);
 const head=52,foot=34;
 const cv=document.createElement('canvas');
 cv.width=Wpx;cv.height=Hpx+head+foot;
 const ctx=cv.getContext('2d');
 ctx.fillStyle='#ffffff';ctx.fillRect(0,0,cv.width,cv.height);
 const c2=document.createElement('canvas');c2.width=Wpx-24;c2.height=Hpx;
 const data=JSON.parse(JSON.stringify(tsc.data));
 const ch=new Chart(c2,{type:'line',data:data,options:{
  responsive:false,animation:false,devicePixelRatio:1,
  plugins:{legend:{display:true,labels:{usePointStyle:true,color:'#000',boxWidth:30,
   font:{size:Math.round(Wpx/100)}}}},
  scales:{x:{ticks:{color:'#000',maxTicksLimit:8,font:{size:Math.round(Wpx/110)}}},
   y:{title:{display:true,text:t('axis_displacement',{U:U}),color:'#000',
    font:{size:Math.round(Wpx/100)}},ticks:{color:'#000',font:{size:Math.round(Wpx/110)}}}}}});
 ctx.drawImage(c2,12,head);
 ch.destroy();
 ctx.fillStyle='#222';ctx.font=`bold ${Math.round(Wpx/70)}px sans-serif`;ctx.textAlign='left';
 ctx.fillText(document.getElementById('tstitle').textContent,14,Math.round(Wpx/45));
 ctx.fillStyle='#444';ctx.font=`${Math.round(Wpx/95)}px sans-serif`;
 ctx.fillText(document.getElementById('tsinfo').textContent,14,head-8);
 ctx.fillStyle='#777';ctx.font=`${Math.round(Wpx/120)}px sans-serif`;
 ctx.fillText(document.getElementById('titleTxt').textContent+'  |  '+
  D.dates[0]+' ~ '+D.dates[D.dates.length-1],14,cv.height-12);
 return cv;
}
document.getElementById('tsPngExport').onclick=()=>{
 const cv=drawTsPNG();
 if(!cv){document.getElementById('tsinfo').textContent=t('ts_png_none');return;}
 cv.toBlob(b=>{
  const a=document.createElement('a');
  a.href=URL.createObjectURL(b);
  a.download='insar_timeseries.png';a.click();
  setTimeout(()=>URL.revokeObjectURL(a.href),5000);
 },'image/png');
};

// ---- 地圖拖曳框選: 決定 GIF 只輸出哪一段里程 ----
let gifRectMode=false,gifRectStart=null,gifRect=null,gifRangeS=null,gifSuppressUntil=0;
const gifRangeBtn=document.getElementById('gifRangeBtn');
const gifStatusEl=document.getElementById('gifStatus');
function endGifRectMode(){
 gifRectMode=false;gifRectStart=null;
 map.dragging.enable();
 map.getContainer().style.cursor='';
 gifRangeBtn.textContent=t('gif_range_btn');
}
function clearGifRange(){
 if(gifRect){map.removeLayer(gifRect);gifRect=null;}
 gifRangeS=null;
}
gifRangeBtn.onclick=()=>{
 if(gifRectMode){endGifRectMode();gifStatusEl.textContent=t('gif_msg_cancelled');return;}
 // 地圖旋轉時螢幕上的拖曳軌跡與地理正交矩形對不上 → 比照匯出畫面 GeoTIFF 的作法先擋掉
 const bg=map.getBearing?((map.getBearing()%360)+360)%360:0;
 if(Math.abs(bg)>0.01&&Math.abs(bg-360)>0.01){gifStatusEl.textContent=t('gif_msg_rotate');return;}
 if(drawing)cancelDraw();                          // 與畫線/畫遮罩互斥
 if(maskDrawing)cancelMaskDraw();
 gifRectMode=true;gifRectStart=null;
 map.dragging.disable();
 map.getContainer().style.cursor='crosshair';
 gifRangeBtn.textContent=t('gif_range_btn_active');
 gifStatusEl.textContent=t('gif_msg_drag');
};
document.getElementById('gifFullBtn').onclick=()=>{
 if(gifRectMode)endGifRectMode();
 clearGifRange();
 gifStatusEl.textContent=t('gif_msg_full');
};
// 放開在地圖容器外時 map 的 mouseup 不會觸發 → 用 document 層級收尾, 避免卡在框選模式
function gifDocUp(){setTimeout(()=>{
 if(gifRectMode){endGifRectMode();gifStatusEl.textContent=t('gif_msg_cancelled');}},0);}
map.on('mousedown',e=>{
 if(!gifRectMode)return;
 gifRectStart=e.latlng;
 clearGifRange();
 document.addEventListener('mouseup',gifDocUp,{once:true});
});
map.on('mousemove',e=>{
 if(!gifRectMode||!gifRectStart)return;
 const b=L.latLngBounds(gifRectStart,e.latlng);
 if(gifRect)gifRect.setBounds(b);
 else gifRect=L.rectangle(b,{color:'#2563eb',weight:1.5,dashArray:'4 3',
  fillColor:'#2563eb',fillOpacity:.06,interactive:false}).addTo(map);
});
map.on('mouseup',e=>{
 if(!gifRectMode||!gifRectStart)return;
 const b=L.latLngBounds(gifRectStart,e.latlng);
 endGifRectMode();
 gifSuppressUntil=Date.now()+400;                  // 別讓框選結束的那一下順便開時序視窗
 applyGifRange(b);
});
// 框內剖面站點的里程 min/max 即為 GIF 的 x 軸範圍
function applyGifRange(b){
 const st=profStations;
 if(!st||st.length<2){clearGifRange();gifStatusEl.textContent=t('gif_msg_noline');return;}
 let lo=Infinity,hi=-Infinity,n=0;
 for(const p of st){
  if(!b.contains(L.latLng(p.lat,p.lon)))continue;
  n++;if(p.s<lo)lo=p.s;if(p.s>hi)hi=p.s;}
 if(n<2||!(hi>lo)){clearGifRange();gifStatusEl.textContent=t('gif_msg_toofew');return;}
 gifRangeS=[lo,hi];
 gifStatusEl.textContent=t('gif_msg_range',
  {a:km(lo).toFixed(2),b:km(hi).toFixed(2),n:n});
}

// ---- 逐幀擷取剖面圖 → GIF ----
const GIF_MAX_W=900;                               // 幀寬上限, 控制記憶體與檔案大小
const GIF_FRAME_BUDGET=360*1024*1024;              // 全部幀緩衝的總量上限 (RGBA bytes)
async function exportProfileGif(){
 if(!pc){gifStatusEl.textContent=t('gif_msg_nochart');return;}
 const btn=document.getElementById('gifExport'),playBtn=document.getElementById('play');
 btn.disabled=true;playBtn.disabled=true;stopTimer();   // 匯出期間鎖住播放, 免得 timer 與擷取迴圈互搶期數
 const chart0=pc;                                  // 匯出途中圖表若被清除/重建就中止, 不去動新的實例
 const N=D.dates.length,oldDi=di;
 // Chart.js v4 會在 update 過程重建 options 物件, 快取的參考會失效 → 每次重新取得
 const setX=(mn,mx)=>{const s=pc.options.scales.x;
  if(mn===undefined)delete s.min;else s.min=mn;
  if(mx===undefined)delete s.max;else s.max=mx;};
 const oldMin=pc.options.scales.x.min,oldMax=pc.options.scales.x.max;
 const src=pc.canvas;
 // 幀寬取「上限」與「記憶體預算容許值」較小者: bytes = N*w*(w*ratio)*4
 const ratio=src.height/src.width;
 const wBudget=Math.floor(Math.sqrt(GIF_FRAME_BUDGET/(N*4*ratio)));
 const capW=Math.max(240,Math.min(GIF_MAX_W,wBudget));
 const sc=Math.min(1,capW/src.width);
 const w=Math.max(2,Math.round(src.width*sc)),h=Math.max(2,Math.round(src.height*sc));
 if(capW<GIF_MAX_W&&sc<1)gifStatusEl.textContent=t('gif_msg_shrunk',{w:w});
 const off=document.createElement('canvas');off.width=w;off.height=h;
 const octx=off.getContext('2d');
 const frames=[];
 try{
  for(let i=0;i<N;i++){
   if(pc!==chart0)throw t('gif_msg_aborted');                  // 圖表被換掉(清除線/載入新 h5)
   slider.value=i;slider.dispatchEvent(new Event('input'));    // 走既有 setEpoch, 與手動播放同一路徑
   if(gifRangeS)setX(km(gifRangeS[0]),km(gifRangeS[1]));       // 每幀重申, 避免 options 被重建後失效
   pc.update('none');
   await new Promise(r=>requestAnimationFrame(()=>requestAnimationFrame(r)));
   octx.fillStyle='#fff';octx.fillRect(0,0,w,h);               // chart canvas 本身透明, 先鋪白底
   octx.drawImage(src,0,0,w,h);
   const fs=Math.max(11,Math.round(h*0.055)),lab=`${D.dates[i]}  (${i+1}/${N})`;
   octx.textBaseline='top';octx.font='bold '+fs+'px sans-serif';
   octx.fillStyle='rgba(255,255,255,0.86)';                    // 底框: 避免壓在頂端色階條上看不清
   octx.fillRect(4,3,octx.measureText(lab).width+9,fs+6);
   octx.fillStyle='#111';octx.fillText(lab,8,6);
   frames.push(octx.getImageData(0,0,w,h).data);
   gifStatusEl.textContent=t('gif_msg_capture',{i:i+1,n:N});
  }
  gifStatusEl.textContent=t('gif_msg_encoding');
  await new Promise(r=>setTimeout(r,20));
  const fps=Math.max(1,Math.min(20,+document.getElementById('gifFps').value||4));
  const gif=gifEncode(frames,w,h,Math.max(2,Math.round(100/fps)),256);
  const tag=(D.dates[0]+'_'+D.dates[N-1]).replace(/-/g,'');
  const a=document.createElement('a');
  a.href=URL.createObjectURL(new Blob([gif],{type:'image/gif'}));
  a.download='profile_'+tag+'.gif';
  document.body.appendChild(a);a.click();a.remove();
  setTimeout(()=>URL.revokeObjectURL(a.href),8000);
  gifStatusEl.textContent=t('gif_msg_done',
   {n:N,w:w,h:h,mb:(gif.length/1048576).toFixed(2)});
 }catch(err){
  gifStatusEl.textContent=t('gif_msg_fail',{err:err});
 }finally{
  if(pc&&pc===chart0){                                         // 圖表換過就別動它, 新的實例有自己的狀態
   setX(oldMin,oldMax);                                        // 還原 x 軸與原本停留的期數
   slider.value=oldDi;slider.dispatchEvent(new Event('input'));
   setX(oldMin,oldMax);                                        // setEpoch 的 update 可能再度重建 options
   pc.update('none');
  }
  btn.disabled=false;playBtn.disabled=false;
 }
}
document.getElementById('gifExport').onclick=exportProfileGif;

// ============ 速率剖面圖 PNG (QGIS 風格) ============
// 沿線每個取樣點的變化速率畫成單條綠線+方形標記, 只有水平格線, 無圖例/無色階條。
// 配色與版面比照 dolphin202504-202607-profile.png (2026-08-06 由該圖取樣)。
const RP_LINE='#518137',RP_FILL='#6ea84a',RP_FRAME='#c8c8c8',RP_GRID='#e9e9e9',RP_TEXT='#5f5f5f';
// 沿線速率 [{s(公尺), v(顯示單位/yr 或 null)}]: 互動剖面用首末期差, CLI 剖面直接用點速度
function rateProfilePts(){
 const N=D.dates.length;
 const yearSpan=(new Date(D.dates[N-1])-new Date(D.dates[0]))/(365.25*864e5);
 if(activeInteractivePts&&activeInteractivePts.length>1)
  return activeInteractivePts.map(p=>{
   const a=p.t[0],b=p.t[N-1];
   return {s:p.s,v:(a==null||b==null||yearSpan<=0)?null:(b-a)/yearSpan};});
 if(D.hasLine)return cliProfilePts().map(p=>({s:p.s,v:(p.v==null||!isFinite(p.v))?null:p.v}));
 return null;
}
// 1/2/5 級距的漂亮刻度; 回傳 {ticks,min,max}
function rpNiceTicks(lo,hi,target){
 if(!(hi>lo)){hi=lo+1;}
 const raw=(hi-lo)/Math.max(2,target);
 const mag=Math.pow(10,Math.floor(Math.log10(raw)));
 const nm=raw/mag,step=(nm<=1?1:nm<=2?2:nm<=5?5:10)*mag;
 const t0=Math.floor(lo/step)*step,t1=Math.ceil(hi/step)*step;
 const ticks=[];
 for(let i=0;i<=Math.round((t1-t0)/step);i++)ticks.push(+(t0+i*step).toPrecision(12));
 return {ticks:ticks,min:t0,max:t1};
}
function rpFmt(v){return Math.abs(v)>=1000?String(Math.round(v)):String(+v.toFixed(2));}
function drawRateProfile(pts,W,H,dpr){
 const cv=document.createElement('canvas');
 cv.width=Math.round(W*dpr);cv.height=Math.round(H*dpr);
 const c=cv.getContext('2d');c.scale(dpr,dpr);
 c.fillStyle='#fff';c.fillRect(0,0,W,H);
 const fsTick=Math.max(10,Math.round(H*0.031)),fsAxis=Math.max(11,Math.round(H*0.037));
 const FONT=`"Microsoft JhengHei","PingFang TC","Noto Sans TC","Heiti TC",sans-serif`;
 const vals=pts.filter(p=>p.v!=null).map(p=>p.v);
 const sMax=pts.reduce((m,p)=>p.s>m?p.s:m,0);
 const vmin=Math.min.apply(null,vals),vmax=Math.max.apply(null,vals);
 const padV=(vmax-vmin)*0.08||1;
 const Y=rpNiceTicks(vmin-padV,vmax+padV,8),X=rpNiceTicks(0,sMax,9);
 // 右邊距至少要放得下最後一個 x 刻度標籤的一半, 否則像 "25000" 會被畫布切掉
 c.font=fsTick+'px '+FONT;
 const halfLast=c.measureText(rpFmt(X.ticks[X.ticks.length-1])).width/2;
 const pad={l:Math.round(W*0.068),r:Math.round(Math.max(W*0.014,halfLast+4)),
            t:Math.round(H*0.035),b:Math.round(H*0.175)};
 const pw=W-pad.l-pad.r,ph=H-pad.t-pad.b;
 const px=s=>pad.l+(s-X.min)/(X.max-X.min)*pw;
 const py=v=>pad.t+(Y.max-v)/(Y.max-Y.min)*ph;
 // 水平格線 (無垂直格線, 照範例圖)
 c.strokeStyle=RP_GRID;c.lineWidth=1;
 Y.ticks.forEach(v=>{const y=Math.round(py(v))+0.5;
  c.beginPath();c.moveTo(pad.l,y);c.lineTo(pad.l+pw,y);c.stroke();});
 // 外框
 c.strokeStyle=RP_FRAME;c.lineWidth=1;
 c.strokeRect(pad.l+0.5,pad.t+0.5,pw,ph);
 // 刻度文字
 c.fillStyle=RP_TEXT;c.font=fsTick+'px '+FONT;
 c.textAlign='right';c.textBaseline='middle';
 Y.ticks.forEach(v=>c.fillText(rpFmt(v),pad.l-8,py(v)));
 c.textAlign='center';c.textBaseline='top';
 X.ticks.forEach(v=>c.fillText(rpFmt(v),px(v),pad.t+ph+7));
 // 軸標題
 c.font=fsAxis+'px '+FONT;
 c.fillText(t('rate_axis_x'),pad.l+pw/2,pad.t+ph+7+fsTick+Math.round(H*0.045));
 c.save();c.translate(Math.round(W*0.016),pad.t+ph/2);c.rotate(-Math.PI/2);
 c.textBaseline='middle';c.fillText(t('rate_axis_y',{U:U}),0,0);c.restore();
 // 折線 (null 處斷開)
 c.strokeStyle=RP_LINE;c.lineWidth=2;c.lineJoin='round';
 let open=false;c.beginPath();
 pts.forEach(p=>{
  if(p.v==null){open=false;return;}
  if(open)c.lineTo(px(p.s),py(p.v));else{c.moveTo(px(p.s),py(p.v));open=true;}});
 c.stroke();
 // 方形標記: 取樣點很密時縮小到不超過相鄰點間距, 免得糊成一整片
 const gap=pw/Math.max(1,pts.length-1);
 const ms=Math.max(3,Math.min(Math.round(H*0.016),Math.round(gap*0.85)));
 c.lineWidth=1;
 pts.forEach(p=>{if(p.v==null)return;
  const x=Math.round(px(p.s))-ms/2,y=Math.round(py(p.v))-ms/2;
  c.fillStyle=RP_FILL;c.fillRect(x,y,ms,ms);
  // 標記太小時描邊會把填色整個蓋掉, 只有夠大才畫深綠外框
  if(ms>=5){c.strokeStyle=RP_LINE;c.strokeRect(x+0.5,y+0.5,ms-1,ms-1);}});
 return cv;
}
function exportRateProfilePNG(){
 const pts=rateProfilePts();
 if(!pts||pts.length<2){gifStatusEl.textContent=t('rate_msg_noline');return;}
 const good=pts.filter(p=>p.v!=null);
 if(good.length<2){gifStatusEl.textContent=t('rate_msg_nodata');return;}
 const W=Math.max(400,Math.min(4000,+document.getElementById('ratePngW').value||1400));
 const H=Math.round(W*0.355);                      // 比照範例圖 1394×495 的長寬比
 const cv=drawRateProfile(pts,W,H,2);              // 2x 解析度輸出, 報告裡放大不糊
 cv.toBlob(b=>{
  const a=document.createElement('a');
  a.href=URL.createObjectURL(b);
  a.download='rate_profile_'+(D.dates[0]+'_'+D.dates[D.dates.length-1]).replace(/-/g,'')+'.png';
  document.body.appendChild(a);a.click();a.remove();
  setTimeout(()=>URL.revokeObjectURL(a.href),8000);
  gifStatusEl.textContent=t('rate_msg_done',{w:cv.width,h:cv.height,n:good.length});
 },'image/png');
}
document.getElementById('ratePngExport').onclick=exportRateProfilePNG;

// 地圖點擊(非畫線模式且格網已載入)→ 以 gridSample 取該處各期值畫時序
map.on('click',e=>{
 if(Date.now()<gifSuppressUntil)return;            // 剛結束框選拖曳, 這一下不算點選
 if(drawing||maskDrawing||gifRectMode||!GRID)return;
 const lon=e.latlng.lng,lat=e.latlng.lat;
 const t=[];let any=false;
 for(let k=0;k<D.dates.length;k++){
  const v=gridSample(lon,lat,k);
  t.push(v==null?null:+v.toFixed(1));
  if(v!=null)any=true;}
 if(any)showGridTS(lon,lat,t);
});

// 檔案匯入 (GeoJSON/.json 或 .shp)
document.getElementById('lineFile').onchange=async(e)=>{
 const file=e.target.files[0];
 if(!file)return;
 document.getElementById('lineStatus').textContent=t('msg_reading_file');
 try{
  let verts;
  if(/\.shp$/i.test(file.name)){
   verts=parseShpToVerts(await file.arrayBuffer());
  }else{
   verts=geojsonToVerts(JSON.parse(await file.text()));
  }
  verts=normalizeCoords(verts);
  finishLine(verts);
 }catch(err){
  console.error(err);
  document.getElementById('lineStatus').textContent=t('msg_import_failed')+err;
  alert(t('msg_line_import_failed')+err);
 }
 e.target.value='';
};

// 清除互動剖面: 移除地圖線 + 還原 CLI 剖面(若有)或隱藏剖面窗
// (具名函式: 載入 HDF5 換資料時直接重用同一段清理邏輯)
function resetProfilePane(){
 activeInteractivePts=null;
 interactiveLineLayer.clearLayers();
 interactiveLine=null;                             // 圖層已清空, handle 一併歸零避免選色器操作到已移除的線
 clearGifRange();                                  // 線沒了, 舊的 GIF 框選里程區間一併作廢
 if(drawing)cancelDraw();
 if(gifRectMode)endGifRectMode();
 if(D.hasLine){renderProfilePane(cliProfilePts(),p=>p.v,effLims('vel'));setupProfileHover(cliStations());}
 else{stopTimer();if(pc){pc.destroy();pc=null;}document.getElementById('profwin').style.display='none';
  clearProfileHover();}
 document.getElementById('lineStatus').textContent='';
}
document.getElementById('clearLineBtn').onclick=resetProfilePane;

// ---- GNSS CSV 瀏覽器內載入 + 匯出更新後的自包含 HTML ----
// 座標合理性判斷: |a|或|b|>1000 視為 TWD97 TM2(EPSG:3826) 平面座標, 沿用既有 twd97ToWgs84
// (量級 x約5e4~5e5, y約2.3e6~2.9e6, 不會與年份/速度等小數值欄混淆);
// 否則需落在 WGS84 台灣範圍(lon 118~124, lat 21~27)才視為合法座標, 並自動判斷欄位順序,
// 兩者皆不符 → 回傳 null(不合理, 呼叫端須略過, 不得靜默當座標用)
function gnssCoordPair(a,b){
 if(Math.abs(a)>1000||Math.abs(b)>1000)return twd97ToWgs84(a,b);
 const aLon=a>=118&&a<=124,aLat=a>=21&&a<=27,bLon=b>=118&&b<=124,bLat=b>=21&&b<=27;
 if(aLon&&bLat)return [a,b];                  // 欄序 經度,緯度
 if(aLat&&bLon)return [b,a];                  // 欄序偵測為 緯度,經度 → 交換回 [lon,lat]
 return null;                                  // 兩欄皆非合理座標範圍 → 不接受
}
const GNSS_NUM_RE=/^-?\d+(\.\d+)?$/;
function gnssSplitLine(line){
 if(line.indexOf(',')>=0)return line.split(',').map(s=>s.trim());
 if(line.indexOf('\t')>=0)return line.split('\t').map(s=>s.trim());
 return line.trim().split(/\s+/);
}
// header 欄名比對(不分大小寫, 精確相等避免 "Ve"/"start" 等誤配): 站名/經度/緯度三欄都對得到才採用
const GNSS_HDR_NAME=new Set(['station','site','name','stationname','sitename','站','站名','測站','测站']);
const GNSS_HDR_LON=new Set(['lon','long','longitude','x','e','經度','经度']);
const GNSS_HDR_LAT=new Set(['lat','latitude','y','n','緯度','纬度']);
function gnssMatchHeaderCols(headerRow){
 let nameIdx=-1,lonIdx=-1,latIdx=-1;
 headerRow.forEach((h,i)=>{
  const v=h.trim().toLowerCase();
  if(nameIdx<0&&GNSS_HDR_NAME.has(v))nameIdx=i;
  else if(lonIdx<0&&GNSS_HDR_LON.has(v))lonIdx=i;
  else if(latIdx<0&&GNSS_HDR_LAT.has(v))latIdx=i;
 });
 return (nameIdx>=0&&lonIdx>=0&&latIdx>=0)?{nameIdx,lonIdx,latIdx}:null;
}
// 純函式: CSV/TXT 文字 → {stations:[{n,lon,lat}], skipped:略過列數}
function parseGnssCsv(text){
 const lines=text.split(/\r\n|\r|\n/).map(l=>l.trim()).filter(l=>l.length);
 if(!lines.length)return {stations:[],skipped:0};
 let rows=lines.map(gnssSplitLine);
 const numCount=r=>r.filter(f=>GNSS_NUM_RE.test(f)).length;
 let cols=null;
 if(numCount(rows[0])<2){                     // 首列非數值欄 ≥2 → 視為 header
  cols=gnssMatchHeaderCols(rows[0]);           // 優先依欄名精確對欄(站名/lon/lat)
  rows=rows.slice(1);
 }
 const stations=[];let skipped=0;
 for(const r of rows){
  let name,lon,lat;
  if(cols){                                    // header 對欄成功: 直接取指名欄位, 不再自動偵測順序
   const a=parseFloat(r[cols.lonIdx]),b=parseFloat(r[cols.latIdx]);
   const pair=(isFinite(a)&&isFinite(b))
    ?((Math.abs(a)>1000||Math.abs(b)>1000)?twd97ToWgs84(a,b):[a,b])
    :null;
   if(!pair){skipped++;continue;}
   [lon,lat]=pair;
   name=(r[cols.nameIdx]||'').trim();
  }else{                                       // 無 header 或對欄失敗: 逐列自行判斷
   const numIdx=[],numVal=[];
   for(let i=0;i<r.length;i++)if(GNSS_NUM_RE.test(r[i])){numIdx.push(i);numVal.push(parseFloat(r[i]));}
   let pair=null;
   for(let i=0;i+1<numVal.length&&!pair;i++)pair=gnssCoordPair(numVal[i],numVal[i+1]);
   if(!pair){skipped++;continue;}              // 整列找不到合法座標對 → 略過
   [lon,lat]=pair;
   const nameIdx=r.findIndex(f=>!GNSS_NUM_RE.test(f));   // 站名 = 第一個非數值欄
   name=nameIdx>=0?r[nameIdx].trim():'';
  }
  if(!isFinite(lon)||!isFinite(lat)){skipped++;continue;}
  stations.push({n:name||`P${stations.length+1}`,lon:+lon.toFixed(6),lat:+lat.toFixed(6)});
 }
 return {stations,skipped};
}
document.getElementById('gnssCsvFile').addEventListener('change',e=>{
 const file=e.target.files[0];
 if(!file)return;
 const st=document.getElementById('gnssLoadStatus');
 const reader=new FileReader();
 reader.onload=()=>{
  try{
   const {stations,skipped}=parseGnssCsv(reader.result);
   if(!stations.length){st.textContent=t('gnss_no_stations_skipped',{n:skipped});return;}
   D.gnss=stations;gnssAll=stations.slice();
   buildGnssLayer();
   st.textContent=t('gnss_loaded_n',{n:stations.length})+(skipped?t('gnss_skipped_suffix',{n:skipped}):'');
   document.getElementById('exportGnssHtml').disabled=false;
  }catch(err){console.error(err);st.textContent=t('gnss_csv_parse_failed')+err;}
 };
 reader.onerror=()=>{st.textContent=t('gnss_read_failed');};
 reader.readAsText(file,'utf-8');
 e.target.value='';
});
// 匯出「帶新 GNSS 測站」的自包含 HTML: 在原始頁面原始碼(__SRC_HTML)中把 GNSS 起訖標記
// (GNSS_S/GNSS_E, 與 Python _embed_gnss_json / buildShareHTML 同一約定)間的片段換成目前
// D.gnss, 其餘位元組不動 (含 CDN/pack/offline 版函式庫、既有 DATA), 使匯出檔可在無 Python
// 環境的電腦開啟, 也可再次載入 CSV / 再匯出(標記本身會被保留).
// 標記常數刻意用字串相加組成: 若原始碼本身直接寫出完整連續標記文字, 會被下面這個掃描
// 整份原始頁面原始碼的正規表示式誤配對到 (而非只配對到真正內嵌的資料片段).
const GNSS_S='/*__GNSS'+'_S__*/', GNSS_E='/*__GNSS'+'_E__*/';
const GNSS_MARK_RE=/\/\*__GNSS_S__\*\/[\s\S]*?\/\*__GNSS_E__\*\//;
function buildGnssUpdatedHTML(){
 if(!GNSS_MARK_RE.test(__SRC_HTML))throw t('err_gnss_marker_missing');
 const seg=GNSS_S+JSON.stringify(D.gnss||[])+GNSS_E;
 return __SRC_HTML.replace(GNSS_MARK_RE,seg);
}
document.getElementById('exportGnssHtml').addEventListener('click',()=>{
 const st=document.getElementById('gnssLoadStatus');
 try{
  const html=buildGnssUpdatedHTML();
  const blob=new Blob([html],{type:'text/html'});
  const a=document.createElement('a');
  a.href=URL.createObjectURL(blob);
  const base=(decodeURIComponent((location.pathname.split('/').pop()||'gmtsar_timeseries_html_viewer.html')))
   .replace(/\.html?$/i,'')||'gmtsar_timeseries_html_viewer';
  a.download=base+t('gnss_export_suffix_filename');
  a.click();
  setTimeout(()=>URL.revokeObjectURL(a.href),5000);
 }catch(err){console.error(err);st.textContent=t('alert_export_failed')+err;}
});

// ==================== 瀏覽器端直讀 HDF5 (h5wasm) ====================
// 選一個 make_pt_ts_h5.py 格式的 .h5 (+ 新標題) → 就地換掉整個檢視器狀態, 不需重跑 Python.
// 全部重用既有機制: 量化格網 GRID → refreshDerived/applyLayer/rasterPNGT/asymLimsT/
// buildContours(dynamic)/buildGnssLayer/resetProfilePane, 不另建平行資料流.
const H5_QSTEP=0.1;                          // 量化步長 (mm), 與 python 端 GRID_SCALE 同值
const H5_NODATA=-32768;
const H5_MAX_POINTS=20000;                   // 對應 python --max-points 預設值
// 產生此頁時的色階錨點: --vlim 會讓速度場 neg==pos (對稱), 換資料時保持同一上限,
// 讓既有分級色不因新資料的百分位而改變
const LIMS_ANCHOR={vel:Object.assign({},D.lims.vel),cum:Object.assign({},D.lims.cum)};
const LIMS_SYMMETRIC=Math.abs(D.lims.vel.neg-D.lims.vel.pos)<1e-9;
function h5Status(msg){document.getElementById('h5Status').textContent=msg||'';}
const h5Tick=()=>new Promise(r=>setTimeout(r,0));

// 逐像元 OLS 速度 (單位/yr), 與 python make_pt_ts_h5.ols_slope 同式:
//   tm = t - mean(t);  slope = Σ(tm*(y-mean(y))) / Σ(tm²)
// 這裡展開成一次逐期掃描累積的等價式 (平均以該像元「有值期數」為基準;
// 全期有值時與 python 逐項相同). 逐期外圈 + 連續索引內圈, 避免跨步取值與物件包裝.
function h5OlsVel(vals,N,n,t){
 const cnt=new Float64Array(n),sT=new Float64Array(n),sY=new Float64Array(n),
       sTY=new Float64Array(n),sTT=new Float64Array(n);
 for(let e=0;e<N;e++){
  const off=e*n,te=t[e];
  for(let i=0;i<n;i++){
   const y=vals[off+i];
   if(!isFinite(y))continue;
   cnt[i]++;sT[i]+=te;sY[i]+=y;sTY[i]+=te*y;sTT[i]+=te*te;
  }
 }
 const vel=new Float32Array(n).fill(NaN);
 for(let i=0;i<n;i++){
  if(cnt[i]<2)continue;
  const den=sTT[i]-sT[i]*sT[i]/cnt[i];       // = Σ(t-t̄)²
  if(!(den>1e-12))continue;
  vel[i]=(sTY[i]-sT[i]*sY[i]/cnt[i])/den;    // = Σ(t-t̄)(y-ȳ)/Σ(t-t̄)²
 }
 return vel;
}
// 代表點: 有效速度像元 → 等間隔抽樣至上限 (對應 python main() 的「速度有限」遮罩 +
// --max-points; python 超量時隨機抽樣, 此處改等間隔以求可重現, 空間涵蓋等價).
// 本模板不畫 PS 點圖層 (markers=[]), 這組點供 fitBounds/剖面/匯出使用.
function h5BuildPoints(velF,vals,N,n,W,gt,uscale){
 const idx=[];
 for(let i=0;i<n;i++)if(isFinite(velF[i]))idx.push(i);
 if(!idx.length)return [];
 const stride=Math.max(1,Math.ceil(idx.length/H5_MAX_POINTS));
 const nd=(U==='cm')?2:1;
 const pts=[];
 for(let k=0;k<idx.length;k+=stride){
  const i=idx[k],r=Math.floor(i/W),c=i%W;
  const rec={lon:+(gt[0]+(c+0.5)*gt[1]).toFixed(6),lat:+(gt[3]+(r+0.5)*gt[5]).toFixed(6),
   v:+velF[i].toFixed(nd),t:[]};
  for(let e=0;e<N;e++){const v=vals[e*n+i];rec.t.push(isFinite(v)?+(v*uscale).toFixed(nd):null);}
  pts.push(rec);
 }
 return pts;
}
// 解析 + 驗證 + 計算; 任一驗證失敗 → throw (呼叫端顯示錯誤, 此前不動任何既有狀態)
// NOTE: this function declares `const t=epochYearsFrom(dates)` partway through, which
// (due to JS's temporal-dead-zone scoping) shadows the i18n t() helper for the ENTIRE
// function body, not just after that line -- so every translated string below uses
// window.t explicitly instead of the bare t() used elsewhere in this template.
async function h5Parse(file){
 if(typeof h5wasm==='undefined')
  throw window.t('h5_not_loaded');
 h5Status(window.t('h5_initializing'));
 const {FS}=await h5wasm.ready;
 h5Status(window.t('msg_reading_file'));await h5Tick();
 const tmp='/h5load_'+Date.now()+'.h5';
 FS.writeFile(tmp,new Uint8Array(await file.arrayBuffer()));
 let f=null;
 try{
  // h5wasm 對非 HDF5 檔會丟出內部訊息 (name not defined 之類), 改包成看得懂的說明
  try{f=new h5wasm.File(tmp,'r');}
  catch(e){throw window.t('h5_open_failed',{name:file.name});}
  const keys=f.keys();
  if(keys.indexOf('date')<0||keys.indexOf('timeseries')<0)
   throw window.t('h5_missing_datasets');
  const at=f.attrs;
  if(!at||!at.GEOTRANSFORM)throw window.t('h5_missing_geotransform');
  const gt=Array.prototype.slice.call(at.GEOTRANSFORM.value).map(Number);
  if(gt.length!==6||!gt.every(isFinite)||!(Math.abs(gt[1])>0)||!(Math.abs(gt[5])>0))
   throw window.t('h5_bad_geotransform');
  const dsT=f.get('timeseries'),dsD=f.get('date');
  const shp=dsT.shape;
  if(!shp||shp.length!==3)throw window.t('h5_bad_shape',{shp:shp});
  const N=+shp[0],H=+shp[1],W=+shp[2],n=H*W;
  if(N<2)throw window.t('h5_too_few_epochs',{n:N});
  const dates=Array.prototype.slice.call(dsD.value).map(v=>{
   const s=String(v);
   return s.length===8?s.slice(0,4)+'-'+s.slice(4,6)+'-'+s.slice(6,8):s;});
  if(dates.length!==N)throw window.t('h5_date_mismatch',{a:dates.length,b:N});
  const t=epochYearsFrom(dates);
  if(!t.every(isFinite)||!(t[N-1]>0))throw window.t('h5_bad_dates');
  const srcUnit=at.UNIT?String(at.UNIT.value):'mm';
  if(srcUnit!=='mm'&&srcUnit!=='cm')throw window.t('h5_bad_unit',{u:srcUnit});
  h5Status(window.t('h5_parsing',{n:N,h:H,w:W}));await h5Tick();
  const vals=dsT.value;                      // 攤平的 Float32Array (N*H*W)
  if(!vals||vals.length!==N*n)
   throw window.t('h5_size_mismatch',{a:(vals?vals.length:0),b:`${N}x${H}x${W}`});
  if(srcUnit==='cm')for(let i=0;i<vals.length;i++)vals[i]*=10;   // 內部一律 mm
  h5Status(window.t('h5_computing_vel'));await h5Tick();
  const velMM=h5OlsVel(vals,N,n,t);
  // 顯示單位沿用本頁產生時的設定 (UI 文字已烘進 U), h5 的 DISPLAY_UNIT 僅供參考
  const uscale=(U==='cm')?0.1:1;
  const velF=new Float32Array(n),cumF=new Float32Array(n);
  const lastOff=(N-1)*n;
  for(let i=0;i<n;i++){
   velF[i]=velMM[i]*uscale;
   const c=vals[lastOff+i];
   cumF[i]=isFinite(c)?c*uscale:NaN;
  }
  const points=h5BuildPoints(velF,vals,N,n,W,gt,uscale);
  if(!points.length)throw window.t('h5_no_valid_px');
  h5Status(window.t('h5_quantizing'));await h5Tick();
  const q=new Int16Array(N*n);
  for(let i=0;i<vals.length;i++){
   const v=vals[i];
   q[i]=isFinite(v)?Math.max(-32767,Math.min(32767,Math.round(v/H5_QSTEP))):H5_NODATA;
  }
  return {dates,N,H,W,gt,q,velF,cumF,points,uscale};
 }finally{
  if(f)try{f.close();}catch(e){}
  try{FS.unlink(tmp);}catch(e){}
 }
}
// 就地套用: 換掉 GRID / D 的資料欄位 → 沿用既有 refresh 鏈重畫
function h5Apply(P,newTitle,fname){
 const {dates,N,H,W,gt,q,velF,cumF,points,uscale}=P;
 GRID={arr:q,N,H,W,gt,scale:H5_QSTEP*uscale,nodata:H5_NODATA};
 _gridFieldCache.clear();                    // GRID 換掉 (新 h5), 舊快取的 GRID 後援場已失效
 h5F={vel:velF,cum:cumF};
 D.dates=dates;D.points=points;D.n=points.length;
 D.hasLine=false;D.line=[];D.bearing=0;
 D.grid={shape:[N,H,W],gt,scale:GRID.scale,nodata:H5_NODATA};
 D.lims={vel:LIMS_SYMMETRIC?Object.assign({},LIMS_ANCHOR.vel):asymLimsT(velF),
         cum:asymLimsT(cumF)};
 D.contours={dynamic:true,base:(D.contours&&D.contours.base)||0.1};
 // imageOverlay 角點 = 格網外緣 (gt 原點是外緣角, 不是像元中心)
 const rb=[[Math.min(gt[3],gt[3]+gt[5]*H),Math.min(gt[0],gt[0]+gt[1]*W)],
           [Math.max(gt[3],gt[3]+gt[5]*H),Math.max(gt[0],gt[0]+gt[1]*W)]];
 D.rasters={vel:rasterPNGT(velF,D.lims.vel),cum:rasterPNGT(cumF,D.lims.cum),bounds:rb};
 // 由完整清單重新過濾 (不是從已過濾的 D.gnss 再篩), 換回原區域時測站才會回來
 D.gnss=gnssAll.filter(g=>g.lon>=rb[0][1]&&g.lon<=rb[1][1]&&
                          g.lat>=rb[0][0]&&g.lat<=rb[1][0]);
 // 舊的遮罩/手動色階/剖面/時序圖狀態一律清掉 (它們都綁在舊資料上)
 maskApplied=false;origArr=null;clientF=null;clientRasters=null;autoLims=null;
 manualLims={vel:null,cum:null};
 clearMaskBoundary();
 resetProfilePane();
 if(tsc){tsc.destroy();tsc=null;}
 document.getElementById('tstitle').textContent=t('ts_default_title');
 document.getElementById('tsinfo').textContent='';
 document.getElementById('cbMin').value='';document.getElementById('cbMax').value='';
 if(rasterOv)rasterOv.setBounds(rb);
 else{
  rasterOv=L.imageOverlay(D.rasters.vel,rb,
   {opacity:(+document.getElementById('rasterOpacity').value)/100}).addTo(map);
  _lyrCtrl.addOverlay(rasterOv,t('lyr_raster'));
 }
 buildGnssLayer();
 initContourControls();
 setInteractiveEnabled(true,'');
 map.fitBounds(rb);                          // 總覽小地圖固定全台灣, 紅框會自行跟著 moveend 更新
 // 檔名一併顯示: 沒填新標題時, 標題仍是原區域名稱, 靠檔名才分得出目前看的是哪份資料
 D._fname=fname;
 // 新資料未經 Python 端 GNSS 校正 (校正是產出 HTML 當下針對原始資料算的), 提示未校正;
 // 但只在原頁面本來就有 --gnss-correct 結果(const/plane/none)時才提示 -- 原本沒加
 // --gnss-correct 的頁面(D.gnssCorr 不存在或 mode='off')不該無中生有冒出「未經校正」字樣(S13)
 if(D.gnssCorr&&(D.gnssCorr.mode==='const'||D.gnssCorr.mode==='plane'||D.gnssCorr.mode==='none')){
  D.gnssCorr={mode:'stale'};
 }
 renderMeta();
 renderGnssCorrBadge();
 if(newTitle){
  document.getElementById('titleTxt').textContent=newTitle;
 }
 renderTitle();
 // __SRC_HTML 是頁面原始碼快照, 換資料後已不代表目前內容 → 停用 GNSS 匯出, 避免匯出到舊區域
 const eg=document.getElementById('exportGnssHtml');
 eg.disabled=true;
 eg.removeAttribute('data-i18n-title');   // no longer follows the generic title_export_gnss_html key
 eg.title=t('title_gnss_export_stale');
 window.__i18nHooks.push(()=>{if(eg.disabled&&!eg.hasAttribute('data-i18n-title'))eg.title=t('title_gnss_export_stale');});
 refreshDerived();                           // → applyLayer → buildContours (動態)
 h5Status(t('h5_loaded_summary',{fname:fname,n:N,h:H,w:W,pts:D.n})+
  t('h5_loaded_scale',{neg:D.lims.vel.neg,pos:D.lims.vel.pos,U:U}));
}
async function loadH5File(file,newTitle){
 const P=await h5Parse(file);
 h5Status(window.t('h5_redrawing'));await h5Tick();
 h5Apply(P,newTitle,file.name);
 return P;
}
document.getElementById('h5Load').onclick=async()=>{
 const btn=document.getElementById('h5Load');
 const file=document.getElementById('h5File').files[0];
 if(!file){h5Status(t('h5_select_first'));return;}
 btn.disabled=true;
 try{
  await loadH5File(file,document.getElementById('h5Title').value.trim());
 }catch(err){
  console.error(err);
  h5Status(t('h5_load_failed')+err);
  alert(t('alert_h5_load_failed')+err);
 }finally{btn.disabled=false;}
};

// E2E 測試掛鉤 (供 headless 驗證; 不影響功能)
window.__test={get map(){return map},get GRID(){return GRID},get contourLayer(){return contourLayer},
 get pc(){return pc},get tsc(){return tsc},showGridTS,buildContours,applyLayer,exportLayerTiff,
 gridSample,applyMask,clearMask,geojsonPolygonRings,refreshDerived,exportViewTiff,
 get drawing(){return drawing},get maskDrawing(){return maskDrawing},
 get maskBoundary(){return maskBoundary},
 get rasterUrl(){return rasterOv?rasterOv._url:null},
 get clientRasters(){return clientRasters},get clientF(){return clientF},
 get effLimsSnapshot(){return {vel:effLims('vel'),cum:effLims('cum')}},
 parseGnssCsv,buildGnssLayer,buildGnssUpdatedHTML,get gnssLayer(){return gnssLayer},
 loadH5File,h5Parse,h5Apply,h5OlsVel,h5BuildPoints,epochYearsFrom,datesSpanYr,
 get h5F(){return h5F},D,
 aoiSubset,drawAoiPNG,drawTsPNG,showGnssTS,gnssParseRows,gnssParseDate,
 lsqSlope,yrs,applyInteractiveCurve,get profStations(){return profStations},
 get activeInteractivePts(){return activeInteractivePts},renderProfilePane,
 get aoiBounds(){return aoiBounds},set aoiBounds(v){aoiBounds=v},
 currentFieldT,computeThresholdAreaT,thresholdBoundaryT,applyThreshold,clearThreshold,
 get thLayer(){return thLayer},get thEnabled(){return thEnabled},
 get DISPF(){return window.DISPF}};
</script></body></html>"""


_APP_PANEL = r"""
<div id="apppanel" class="panel" style="top:50%;left:50%;transform:translate(-50%,-50%);
 padding:16px 20px;max-width:min(600px,94vw);max-height:92vh;overflow:auto;z-index:2000">
 <div class="big" data-i18n="app_title">InSAR Viewer — 時序檢視器</div>
 <div style="margin:8px 0 10px;line-height:1.5" data-i18n-html="app_desc">選擇一種資料來源。
  讀檔、內插、速度回歸與繪圖<b>全部在瀏覽器內完成</b>, 資料不會上傳。</div>

 <div style="border:1px solid #ccc;border-radius:4px;padding:8px 10px;margin-bottom:8px">
  <div style="font-weight:600;margin-bottom:4px" data-i18n="src_h5_title">① 時序 HDF5 (MintPy / dolphin / gmtsar2h5)</div>
  <div class="ctrls"><input type="file" id="srcH5" accept=".h5,.hdf5" style="flex:1;min-width:0"></div>
  <div class="ctrls" style="margin-top:4px">
   <label class="muted" data-i18n="src_coh_label">同調性遮罩檔 (選填, 兩種來源皆適用)</label>
   <input type="file" id="srcCoh" accept=".h5,.hdf5" style="flex:1;min-width:0">
   <label class="muted" data-i18n="src_coh_min">門檻</label>
   <input type="number" id="srcCohMin" value="0.4" min="0" max="1" step="0.05" style="width:56px;flex:none">
  </div>
  <div class="muted" style="margin-top:3px;font-size:var(--fs-sm)" data-i18n="src_h5_hint">
   支援 MintPy 屬性 (X_FIRST/Y_FIRST/X_STEP/Y_STEP, 單位 m) 與 gmtsar2h5 的 GEOTRANSFORM (單位 mm/cm)。</div>
 </div>

 <div style="border:1px solid #ccc;border-radius:4px;padding:8px 10px;margin-bottom:8px">
  <div style="font-weight:600;margin-bottom:4px" data-i18n="src_gmt_title">② GMTSAR 時序資料夾</div>
  <div class="ctrls"><input type="file" id="srcGmtDir" webkitdirectory multiple style="flex:1;min-width:0"></div>
  <div class="ctrls" style="margin-top:4px">
   <label class="muted" data-i18n="src_gmt_step">格網間距 (m)</label>
   <input type="number" id="srcGmtStep" value="100" min="10" max="2000" step="10" style="width:64px;flex:none">
   <label class="muted" data-i18n="src_gmt_buffer">資料緩衝 (m)</label>
   <input type="number" id="srcGmtBuf" value="500" min="0" max="5000" step="50" style="width:64px;flex:none"
    data-i18n-title="src_gmt_buffer_title">
  </div>
  <div class="muted" style="margin-top:3px;font-size:var(--fs-sm)" data-i18n="src_gmt_hint">
   需含 disp_NNN_ll.xy (lon lat 位移mm) 與 data_date.txt。散點以格網平均聚合 (非克利金), 屬快視結果; 有效範圍由 PS 點分布＋資料緩衝決定, 不需外部遮罩。</div>
 </div>

 <div style="border:1px solid #ccc;border-radius:4px;padding:8px 10px;margin-bottom:8px">
  <div style="font-weight:600;margin-bottom:4px" data-i18n="src_opt_title">共同選項</div>
  <div class="ctrls">
   <label class="muted" data-i18n="src_vel_label" data-i18n-title="src_vel_title">速度場校正檔 (選填)</label>
   <input type="file" id="srcVel" accept=".geojson,.json,.h5,.hdf5" style="flex:1;min-width:0">
   <label class="muted" data-i18n="src_vel_field">屬性欄</label>
   <input type="text" id="srcVelField" value="field_3" style="width:62px;flex:none">
   <select id="srcVelUnit" style="flex:none"><option value="mm">mm/yr</option><option value="m">m/yr</option></select>
  </div>
  <div class="ctrls">
   <label class="muted" data-i18n="src_vel_deramp" data-i18n-title="src_vel_deramp_title">坡面扣除 C0,C1,C2,LON0,LAT0</label>
   <input type="text" id="srcVelDeramp" placeholder="選填" style="flex:1;min-width:0">
  </div>
  <div class="ctrls">
   <label class="muted"><input type="checkbox" id="optDeramp"> <span data-i18n="src_opt_deramp">去除線性坡面 (deramp)</span></label>
   <label class="muted" data-i18n="src_opt_unit">顯示單位</label>
   <select id="optUnit" style="flex:none"><option value="mm">mm</option><option value="cm">cm</option></select>
   <label class="muted" data-i18n="src_opt_vlim" data-i18n-title="src_opt_vlim_title">色階上限 (0=自動)</label>
   <input type="number" id="optVlim" value="0" min="0" step="1" style="width:60px;flex:none"
    data-i18n-title="src_opt_vlim_title">
  </div>
 </div>

 <div class="ctrls">
  <button id="runLoad" style="font-weight:600" data-i18n="src_run">載入並繪圖</button>
  <span class="muted" data-i18n="src_gnss_note">GNSS 時序資料夾在載入後於左上面板選取</span>
 </div>
 <div id="appstatus" class="muted" style="margin-top:10px;min-height:34px;white-space:pre-line"></div>
</div>
<!-- top:46px 而非 10px: 語言切換鈕與分享按鈕都在 top:10 那一排, 訊息條常駐後會蓋住它們 -->
<div id="appwarn" class="panel" style="display:none;top:46px;left:50%;transform:translateX(-50%);
 padding:6px 12px;z-index:1600;background:#fff7e0;border:1px solid #d97706;max-width:56%">
 <span id="appwarnText"></span>
 <button class="sb" style="margin-left:8px" onclick="this.parentElement.style.display='none'">✕</button>
</div>
"""

_APP_JS = r"""
// ==================== 純瀏覽器 InSAR 時序直讀計算模組 ====================
// 兩種來源 → 共同的中間格式 → buildD() → startViewer(D):
//   ① 時序 HDF5 (MintPy/dolphin: X_FIRST.. + m; gmtsar2h5: GEOTRANSFORM + mm/cm)
//   ② GMTSAR 散點資料夾 (disp_NNN_ll.xy + data_date.txt) → 等經緯格網平均聚合
// 中間格式 S = {dates:[YYYY-MM-DD], N, H, W, gt:[6], mm:[Float32Array(H*W) × N]}
function appStatus(msg){document.getElementById('appstatus').textContent=msg;}
const appTick=()=>new Promise(r=>setTimeout(r,0));

// ---- 共用: h5wasm 開檔 (回傳 {f, cleanup}) ----
async function openH5(file){
 if(typeof h5wasm==='undefined')throw t('h5_not_loaded');
 const {FS}=await h5wasm.ready;
 const tmp='/insar_'+(window.__h5seq=(window.__h5seq||0)+1)+'.h5';
 FS.writeFile(tmp,new Uint8Array(await file.arrayBuffer()));
 let f;
 try{f=new h5wasm.File(tmp,'r');}
 catch(e){try{FS.unlink(tmp);}catch(e2){}throw t('h5_open_failed',{name:file.name});}
 return {f,cleanup:()=>{try{f.close();}catch(e){}try{FS.unlink(tmp);}catch(e){}}};
}
const _attr=(at,k)=>(at&&at[k]!=null)?at[k].value:undefined;
function _attrNum(at,k){const v=_attr(at,k);const x=Number(Array.isArray(v)||ArrayBuffer.isView(v)?v[0]:v);
 return isFinite(x)?x:NaN;}
function _attrStr(at,k){const v=_attr(at,k);if(v==null)return '';
 return String(Array.isArray(v)||ArrayBuffer.isView(v)?v[0]:v);}

// 時序資料集名稱: MintPy/gmtsar2h5 皆為 'timeseries'; 部分 dolphin 產物用 'displacement'
const TS_KEYS=['timeseries','displacement','disp'];
// 同調性資料集名稱 (averageCoh / avgSpatialCoh / temporalCoherence 等)
const COH_KEYS=['coherence','temporalCoherence','averageCoh','avgSpatialCoh','average_coherence','data'];

// ---- ① 時序 HDF5 → S ----
async function parseTsH5(file){
 const {f,cleanup}=await openH5(file);
 try{
  const keys=f.keys();
  const tsk=TS_KEYS.find(k=>keys.indexOf(k)>=0);
  if(!tsk)throw t('h5_missing_datasets');
  const at=f.attrs;
  const dsT=f.get(tsk),shp=dsT.shape;
  if(!shp||shp.length!==3)throw t('h5_bad_shape',{shp:shp});
  const N=+shp[0],H=+shp[1],W=+shp[2],n=H*W;
  if(N<2)throw t('h5_too_few_epochs',{n:N});
  // 幾何: 優先 GEOTRANSFORM (gmtsar2h5), 否則 MintPy 的 X_FIRST/Y_FIRST/X_STEP/Y_STEP
  let gt=null;
  const gtA=_attr(at,'GEOTRANSFORM');
  if(gtA){
   gt=Array.prototype.slice.call(gtA).map(Number);
   if(gt.length!==6||!gt.every(isFinite))gt=null;
  }
  if(!gt){
   const x0=_attrNum(at,'X_FIRST'),y0=_attrNum(at,'Y_FIRST'),
         dx=_attrNum(at,'X_STEP'),dy=_attrNum(at,'Y_STEP');
   if(![x0,y0,dx,dy].every(isFinite))throw t('h5_missing_geotransform');
   gt=[x0,dx,0,y0,0,dy];                    // MintPy Y_STEP 本身為負
  }
  if(!(Math.abs(gt[1])>0)||!(Math.abs(gt[5])>0))throw t('h5_bad_geotransform');
  // 日期: /date (bytes YYYYMMDD) 或屬性 DATE_LIST; 缺時退回 1..N 的年序
  let dates=null;
  if(keys.indexOf('date')>=0){
   dates=Array.prototype.slice.call(f.get('date').value).map(v=>{
    const s=String(v);
    return s.length===8?s.slice(0,4)+'-'+s.slice(4,6)+'-'+s.slice(6,8):s;});
  }
  if(!dates||dates.length!==N)throw t('h5_date_mismatch',{a:dates?dates.length:0,b:N});
  // 單位: MintPy 用 m, gmtsar2h5 用 mm/cm; 一律轉 mm
  const u=(_attrStr(at,'UNIT')||'m').trim().toLowerCase();
  const uk={'m':1000,'meter':1000,'meters':1000,'cm':10,'mm':1}[u];
  if(!uk)throw t('h5_bad_unit',{u:u});
  appStatus(t('h5_parsing',{n:N,h:H,w:W}));await appTick();
  const raw=dsT.value;
  if(!raw||raw.length!==N*n)throw t('h5_size_mismatch',{a:(raw?raw.length:0),b:`${N}x${H}x${W}`});
  // NO_DATA_VALUE 為 0 時忽略: 時序產品的參考期本來就是全零, 拿 0 當 nodata 會整期抹掉
  const nodRaw=_attrNum(at,'NO_DATA_VALUE');
  const nod=(isFinite(nodRaw)&&nodRaw!==0)?nodRaw:NaN;
  const mm=[];
  for(let e=0;e<N;e++){
   const a=new Float32Array(n),off=e*n;
   for(let i=0;i<n;i++){
    const v=raw[off+i];
    a[i]=(isFinite(v)&&!(isFinite(nod)&&v===nod))?v*uk:NaN;
   }
   mm.push(a);
  }
  return {dates,N,H,W,gt,mm,src:'h5',srcUnit:u};
 }finally{cleanup();}
}

// ---- 同調性遮罩: 讀 coh h5 → 依門檻把時序設 NaN ----
// 尺寸相同時逐格對應; 尺寸不同但遮罩自帶地理屬性時, 以經緯度最近鄰取樣
// (GMTSAR 資料夾模式的格網是即時決定的, 幾乎不會與遮罩同尺寸, 所以重取樣是常態)
async function applyCohMask(S,file,thr){
 const {f,cleanup}=await openH5(file);
 try{
  const keys=f.keys();
  const ck=COH_KEYS.find(k=>keys.indexOf(k)>=0)||keys[0];
  if(!ck)throw t('coh_no_dataset');
  const ds=f.get(ck),shp=ds.shape;
  if(!shp||shp.length!==2)throw t('coh_bad_shape',{a:(shp||[]).join('x')});
  const cH=+shp[0],cW=+shp[1],c=ds.value;
  let cut=0;
  if(cH===S.H&&cW===S.W){
   for(let i=0;i<S.H*S.W;i++){
    if(!(c[i]>=thr)){cut++;for(let e=0;e<S.N;e++)S.mm[e][i]=NaN;}
   }
   return {cut,mode:'direct'};
  }
  // 尺寸不同 → 需要遮罩自己的幾何
  const at=f.attrs;
  let cgt=null;
  const gtA=_attr(at,'GEOTRANSFORM');
  if(gtA){
   const g=Array.prototype.slice.call(gtA).map(Number);
   if(g.length===6&&g.every(isFinite))cgt=g;
  }
  if(!cgt){
   const x0=_attrNum(at,'X_FIRST'),y0=_attrNum(at,'Y_FIRST'),
         dx=_attrNum(at,'X_STEP'),dy=_attrNum(at,'Y_STEP');
   if([x0,y0,dx,dy].every(isFinite))cgt=[x0,dx,0,y0,0,dy];
  }
  if(!cgt)throw t('coh_shape_mismatch',{a:cH+'x'+cW,b:S.H+'x'+S.W});
  const g=S.gt;
  let outside=0;
  for(let r=0;r<S.H;r++){
   const lat=g[3]+(r+0.5)*g[5];
   const cr=Math.floor((lat-cgt[3])/cgt[5]);
   for(let cc=0;cc<S.W;cc++){
    const lon=g[0]+(cc+0.5)*g[1];
    const ccol=Math.floor((lon-cgt[0])/cgt[1]);
    const i=r*S.W+cc;
    let v=NaN;
    if(cr>=0&&cr<cH&&ccol>=0&&ccol<cW)v=c[cr*cW+ccol];
    else outside++;
    if(!(v>=thr)){cut++;for(let e=0;e<S.N;e++)S.mm[e][i]=NaN;}
   }
  }
  return {cut,mode:'resampled',outside};
 }finally{cleanup();}
}

// ---- TWD97 TM2 正/逆解 (GRS80, 121E, k0=0.9999, FE=250000) ----
function wgs84ToTwd97(lon,lat){
 const a=6378137.0,f=1/298.257222101,k0=0.9999,fe=250000,lon0=121*Math.PI/180;
 const e2=f*(2-f),ep2=e2/(1-e2);
 const rl=lat*Math.PI/180,rlon=lon*Math.PI/180;
 const N=a/Math.sqrt(1-e2*Math.sin(rl)**2);
 const T=Math.tan(rl)**2,C=ep2*Math.cos(rl)**2;
 const A=(rlon-lon0)*Math.cos(rl);
 const M=a*((1-e2/4-3*e2*e2/64-5*e2**3/256)*rl
  -(3*e2/8+3*e2*e2/32+45*e2**3/1024)*Math.sin(2*rl)
  +(15*e2*e2/256+45*e2**3/1024)*Math.sin(4*rl)
  -(35*e2**3/3072)*Math.sin(6*rl));
 const x=k0*N*(A+(1-T+C)*A**3/6+(5-18*T+T*T+72*C-58*ep2)*A**5/120)+fe;
 const y=k0*(M+N*Math.tan(rl)*(A*A/2+(5-T+9*C+4*C*C)*A**4/24
  +(61-58*T+T*T+600*C-330*ep2)*A**6/720));
 return [x,y];
}
function appTwd97ToWgs84(x,y){                       // 與 viewer 內 twd97ToWgs84 同式 (app scope 用)
 const a=6378137.0,f=1/298.257222101,e2=f*(2-f),k0=0.9999,fx=250000,fy=0;
 const lon0=121*Math.PI/180;
 const M=(y-fy)/k0;
 const e1=(1-Math.sqrt(1-e2))/(1+Math.sqrt(1-e2));
 const mu=M/(a*(1-e2/4-3*e2*e2/64-5*e2*e2*e2/256));
 const phi1=mu+(3*e1/2-27*Math.pow(e1,3)/32)*Math.sin(2*mu)
   +(21*e1*e1/16-55*Math.pow(e1,4)/32)*Math.sin(4*mu)
   +(151*Math.pow(e1,3)/96)*Math.sin(6*mu)
   +(1097*Math.pow(e1,4)/512)*Math.sin(8*mu);
 const e2p=e2/(1-e2),C1=e2p*Math.cos(phi1)**2,T1=Math.tan(phi1)**2;
 const N1=a/Math.sqrt(1-e2*Math.sin(phi1)**2),R1=a*(1-e2)/Math.pow(1-e2*Math.sin(phi1)**2,1.5);
 const Dd=(x-fx)/(N1*k0);
 const lat=phi1-(N1*Math.tan(phi1)/R1)*(Dd*Dd/2-(5+3*T1+10*C1-4*C1*C1-9*e2p)*Math.pow(Dd,4)/24
   +(61+90*T1+298*C1+45*T1*T1-252*e2p-3*C1*C1)*Math.pow(Dd,6)/720);
 const lon=lon0+(Dd-(1+2*T1+C1)*Math.pow(Dd,3)/6
   +(5-2*C1+28*T1-3*C1*C1+8*e2p+24*T1*T1)*Math.pow(Dd,5)/120)/Math.cos(phi1);
 return [lon*180/Math.PI,lat*180/Math.PI];
}

// ---- ② GMTSAR 散點資料夾 → S ----
// disp_NNN_ll.xy: 三欄 "lon lat 位移(mm)"; data_date.txt: 每列一個 YYYYMMDD, 列數 = 檔數。
// 散點以等經緯格網「格內平均」聚合 (O(N), 秒級); 空格以半徑 rad 內的反距離加權補值,
// 超出半徑留 NaN。GMTSAR 沒有同調性遮罩, 有效範圍就是靠這個緩衝距離由 PS 點分布自己決定
// (等同 gmtsar2h5.py 的 --max-ps-dist-km)。這是快視結果, 與克利金產物不同。
const GMT_DISP_RE=/disp[_-]?(\d+)[_-]?ll\.xy$/i;
async function parseGmtsarDir(files,stepM,rad){
 const all=[...files];
 const disp=all.filter(f=>GMT_DISP_RE.test(f.name))
               .sort((a,b)=>(+GMT_DISP_RE.exec(a.name)[1])-(+GMT_DISP_RE.exec(b.name)[1]));
 if(!disp.length)throw t('gmt_no_disp');
 const dfile=all.find(f=>/^data_date\.txt$/i.test(f.name));
 if(!dfile)throw t('gmt_no_datefile');
 const dtxt=(await dfile.text()).trim().split(/\r?\n/).map(s=>s.trim()).filter(Boolean);
 if(dtxt.length!==disp.length)
  throw t('gmt_date_count_mismatch',{a:dtxt.length,b:disp.length});
 const dates=dtxt.map(s=>s.length===8?s.slice(0,4)+'-'+s.slice(4,6)+'-'+s.slice(6,8):s);
 const N=disp.length;
 // 第一遍: 讀全部散點, 同時求 bbox (逐期檔的點位不一定相同)
 const epochs=[];
 let lo0=Infinity,lo1=-Infinity,la0=Infinity,la1=-Infinity;
 for(let e=0;e<N;e++){
  appStatus(t('gmt_reading',{i:e+1,n:N,name:disp[e].name}));await appTick();
  const txt=await disp[e].text();
  const lines=txt.split('\n');
  const lon=new Float64Array(lines.length),lat=new Float64Array(lines.length),
        val=new Float64Array(lines.length);
  let k=0;
  for(let i=0;i<lines.length;i++){
   const s=lines[i];
   if(!s||s.charCodeAt(0)===35)continue;               // '#' 註解
   const p=s.trim().split(/\s+/);
   if(p.length<3)continue;
   const x=+p[0],y=+p[1],v=+p[2];
   if(!isFinite(x)||!isFinite(y)||!isFinite(v))continue;
   lon[k]=x;lat[k]=y;val[k]=v;k++;
   if(x<lo0)lo0=x;if(x>lo1)lo1=x;if(y<la0)la0=y;if(y>la1)la1=y;
  }
  epochs.push({lon:lon.subarray(0,k),lat:lat.subarray(0,k),val:val.subarray(0,k),k});
 }
 if(!isFinite(lo0)||!(lo1>lo0)||!(la1>la0))throw t('gmt_no_points');
 // 等經緯格網 (以 bbox 中心緯度換算經度間距, 讓格子接近正方)
 const latC=(la0+la1)/2;
 const dLat=stepM/111320,dLon=stepM/(111320*Math.cos(latC*Math.PI/180));
 const W=Math.max(2,Math.ceil((lo1-lo0)/dLon)+1),H=Math.max(2,Math.ceil((la1-la0)/dLat)+1);
 if(W*H>4e6)throw t('gmt_grid_too_big',{w:W,h:H});
 const gt=[lo0-dLon/2,dLon,0,la1+dLat/2,0,-dLat];      // gt 原點 = 左上外緣
 const n=W*H;
 const mm=[];
 for(let e=0;e<N;e++){
  appStatus(t('gmt_gridding',{i:e+1,n:N}));await appTick();
  const s=new Float64Array(n),c=new Float64Array(n);
  const E=epochs[e];
  for(let i=0;i<E.k;i++){
   const cx=Math.floor((E.lon[i]-gt[0])/dLon),ry=Math.floor((gt[3]-E.lat[i])/dLat);
   if(cx<0||cx>=W||ry<0||ry>=H)continue;
   const o=ry*W+cx;s[o]+=E.val[i];c[o]++;
  }
  const a=new Float32Array(n).fill(NaN);
  for(let i=0;i<n;i++)if(c[i]>0)a[i]=s[i]/c[i];
  if(rad>0)idwFill(a,s,c,W,H,rad);
  mm.push(a);
 }
 return {dates,N,H,W,gt,mm,src:'gmtsar'};
}
// 空格補值: 以半徑 rad(格) 內已有格值做反距離加權; 半徑內無資料則保持 NaN
function idwFill(a,s,c,W,H,rad){
 const out=new Float32Array(a.length);out.set(a);
 for(let r=0;r<H;r++)for(let cc=0;cc<W;cc++){
  const o=r*W+cc;
  if(c[o]>0)continue;
  let sw=0,sv=0;
  for(let dr=-rad;dr<=rad;dr++){
   const rr=r+dr;if(rr<0||rr>=H)continue;
   for(let dc=-rad;dc<=rad;dc++){
    const c2=cc+dc;if(c2<0||c2>=W)continue;
    const o2=rr*W+c2;
    if(!(c[o2]>0))continue;
    const d2=dr*dr+dc*dc;
    if(d2>rad*rad)continue;
    const w=1/d2;sw+=w;sv+=w*a[o2];
   }
  }
  if(sw>0)out[o]=sv/sw;
 }
 a.set(out);
}

// ---- 速度場校正: 用外部速度場逐格改寫時序趨勢 ----
// 與 gmtsar2h5.py 的 --vel-geojson 同式: D += (v_cal − v_ols)·t, 校正後每格的
// 回歸速度就等於該處的參考速度。原始時序若帶未校正的軌道/參考點坡面, 這一步能把
// 它對齊到已知的速度場; 取不到參考速度的網格保持原值並回報格數 (不外插)。
// 支援: GeoJSON 點檔 (properties.<欄名>) 與二維 HDF5 速度場 (自帶地理屬性)。

// 點檔查詢索引: 依 cell 大小 binning, 查詢時只掃 3x3 格 (瀏覽器沒有 KD-tree)
function buildPointIndex(lon,lat,val,cell){
 let lo0=Infinity,la0=Infinity,lo1=-Infinity,la1=-Infinity;
 for(let i=0;i<lon.length;i++){
  if(lon[i]<lo0)lo0=lon[i];if(lon[i]>lo1)lo1=lon[i];
  if(lat[i]<la0)la0=lat[i];if(lat[i]>la1)la1=lat[i];
 }
 const W=Math.max(1,Math.ceil((lo1-lo0)/cell)+1),H=Math.max(1,Math.ceil((la1-la0)/cell)+1);
 const bins=new Map();
 for(let i=0;i<lon.length;i++){
  const c=Math.floor((lon[i]-lo0)/cell),r=Math.floor((lat[i]-la0)/cell);
  const k=r*W+c;
  let a=bins.get(k);if(!a){a=[];bins.set(k,a);}
  a.push(i);
 }
 return {lon,lat,val,cell,lo0,la0,W,H,bins};
}
function nearestVal(ix,lon,lat,maxD){
 const c0=Math.floor((lon-ix.lo0)/ix.cell),r0=Math.floor((lat-ix.la0)/ix.cell);
 let best=null,bd=maxD*maxD;
 for(let dr=-1;dr<=1;dr++)for(let dc=-1;dc<=1;dc++){
  const a=ix.bins.get((r0+dr)*ix.W+(c0+dc));
  if(!a)continue;
  for(const i of a){
   const dx=ix.lon[i]-lon,dy=ix.lat[i]-lat,d2=dx*dx+dy*dy;
   if(d2<bd){bd=d2;best=ix.val[i];}
  }
 }
 return best;
}
// 速度場檔 → getV(lon,lat) → 參考速度 (mm/yr) 或 null
async function loadVelSource(file,field,unitK,deramp){
 const ext=(file.name.split('.').pop()||'').toLowerCase();
 const dr=deramp;                                  // [C0,C1,C2,LON0,LAT0] 或 null
 const adj=(v,lon,lat)=>{
  if(v==null||!isFinite(v))return null;
  let x=v*unitK;
  if(dr)x-=(dr[0]+dr[1]*(lon-dr[3])+dr[2]*(lat-dr[4]));
  return x;
 };
 if(ext==='geojson'||ext==='json'){
  const gj=JSON.parse(await file.text());
  const fs=gj.features||[];
  if(!fs.length)throw t('vel_no_features');
  const lon=new Float64Array(fs.length),lat=new Float64Array(fs.length),
        val=new Float64Array(fs.length);
  let k=0;
  for(const f of fs){
   const g=f.geometry,p=f.properties||{};
   if(!g||g.type!=='Point')continue;
   const v=p[field];
   if(v==null||!isFinite(+v))continue;
   lon[k]=g.coordinates[0];lat[k]=g.coordinates[1];val[k]=+v;k++;
  }
  if(!k)throw t('vel_no_field',{f:field});
  const ix=buildPointIndex(lon.subarray(0,k),lat.subarray(0,k),val.subarray(0,k),0.003);
  return {n:k,kind:'points',
          getV:(lon,lat)=>adj(nearestVal(ix,lon,lat,0.003),lon,lat)};
 }
 // HDF5 二維速度場
 const {f,cleanup}=await openH5(file);
 try{
  const keys=f.keys();
  const vk=['velocity','vel','data',field].find(k=>keys.indexOf(k)>=0)||keys[0];
  const ds=f.get(vk),shp=ds.shape;
  if(!shp||shp.length!==2)throw t('vel_bad_shape',{a:(shp||[]).join('x')});
  const H=+shp[0],W=+shp[1],arr=ds.value;
  const at=f.attrs;
  let gt=null;
  const gtA=_attr(at,'GEOTRANSFORM');
  if(gtA){const g=Array.prototype.slice.call(gtA).map(Number);
   if(g.length===6&&g.every(isFinite))gt=g;}
  if(!gt){
   const x0=_attrNum(at,'X_FIRST'),y0=_attrNum(at,'Y_FIRST'),
         dx=_attrNum(at,'X_STEP'),dy=_attrNum(at,'Y_STEP');
   if(![x0,y0,dx,dy].every(isFinite))throw t('vel_no_geo');
   gt=[x0,dx,0,y0,0,dy];
  }
  const cp=new Float32Array(arr.length);cp.set(arr);   // h5 關檔後仍要能用
  return {n:H*W,kind:'grid',
   getV:(lon,lat)=>{
    const c=Math.floor((lon-gt[0])/gt[1]),r=Math.floor((lat-gt[3])/gt[5]);
    if(c<0||c>=W||r<0||r>=H)return null;
    return adj(cp[r*W+c],lon,lat);
   }};
 }finally{cleanup();}
}
// 逐格改寫趨勢; 回傳 {fix, skip}
function retrendS(S,getV,tYr){
 const {N,H,W,gt}=S,n=H*W;
 const vols=olsVel(S.mm,N,n,tYr);
 let fix=0,skip=0;
 for(let r=0;r<H;r++){
  const lat=gt[3]+(r+0.5)*gt[5];
  for(let c=0;c<W;c++){
   const i=r*W+c;
   if(!isFinite(vols[i]))continue;
   const vc=getV(gt[0]+(c+0.5)*gt[1],lat);
   if(vc==null||!isFinite(vc)){skip++;continue;}
   const d=vc-vols[i];
   for(let e=0;e<N;e++){const v=S.mm[e][i];if(isFinite(v))S.mm[e][i]=v+d*tYr[e];}
   fix++;
  }
 }
 return {fix,skip};
}

// ---- deramp: 對每一期擬合平面 z = a + b·Δlon + c·Δlat 並扣除 (最小二乘, 忽略 NaN) ----
// 首期通常為全零參考期, 擬合會得到零平面, 不影響。
function derampS(S){
 const {N,H,W,gt}=S,n=H*W;
 const lon0=gt[0]+gt[1]*W/2,lat0=gt[3]+gt[5]*H/2;
 let nfit=0;
 for(let e=0;e<N;e++){
  const a=S.mm[e];
  // 正規方程 (3x3 對稱)
  let s00=0,s01=0,s02=0,s11=0,s12=0,s22=0,b0=0,b1=0,b2=0,cnt=0;
  for(let r=0;r<H;r++){
   const dy=(gt[3]+(r+0.5)*gt[5])-lat0,ro=r*W;
   for(let c=0;c<W;c++){
    const v=a[ro+c];
    if(!isFinite(v))continue;
    const dx=(gt[0]+(c+0.5)*gt[1])-lon0;
    s00++;s01+=dx;s02+=dy;s11+=dx*dx;s12+=dx*dy;s22+=dy*dy;
    b0+=v;b1+=dx*v;b2+=dy*v;cnt++;
   }
  }
  if(cnt<10)continue;
  const M=[[s00,s01,s02],[s01,s11,s12],[s02,s12,s22]],B=[b0,b1,b2];
  const x=solve3(M,B);
  if(!x)continue;
  for(let r=0;r<H;r++){
   const dy=(gt[3]+(r+0.5)*gt[5])-lat0,ro=r*W;
   for(let c=0;c<W;c++){
    if(!isFinite(a[ro+c]))continue;
    const dx=(gt[0]+(c+0.5)*gt[1])-lon0;
    a[ro+c]-=(x[0]+x[1]*dx+x[2]*dy);
   }
  }
  nfit++;
 }
 return nfit;
}
// 3x3 高斯消去 (部分樞軸); 奇異回傳 null
function solve3(M,B){
 const A=[[M[0][0],M[0][1],M[0][2],B[0]],[M[1][0],M[1][1],M[1][2],B[1]],
          [M[2][0],M[2][1],M[2][2],B[2]]];
 for(let i=0;i<3;i++){
  let p=i;
  for(let r=i+1;r<3;r++)if(Math.abs(A[r][i])>Math.abs(A[p][i]))p=r;
  if(Math.abs(A[p][i])<1e-12)return null;
  const tmp=A[i];A[i]=A[p];A[p]=tmp;
  for(let r=0;r<3;r++){
   if(r===i)continue;
   const f=A[r][i]/A[i][i];
   for(let c=i;c<4;c++)A[r][c]-=f*A[i][c];
  }
 }
 return [A[0][3]/A[0][0],A[1][3]/A[1][1],A[2][3]/A[2][2]];
}

// ---- 不對稱色階 (與 viewer NEG/POS_STOPS 同組) + raster PNG ----
// 必須與 viewer 內的 NEG_STOPS/POS_STOPS (以及 Python 的 _NEG_STOPS/_POS_STOPS) 同值:
// 圖層 raster 由這裡畫、colorbar 漸層與點位顏色由 viewer 的 col() 畫, 不同步就圖例對不上。
const APP_NEG=[[128,255,64],[128,255,64],[160,255,64],[192,255,64],[224,240,64],[255,224,64],[255,192,64],[255,160,64],[255,128,64],[255,96,64],[255,96,128],[255,96,192],[255,128,224],[255,160,255],[255,192,255],[255,224,255],[255,240,255]];
const APP_POS=[[64,255,64],[64,255,64],[64,255,160],[64,255,255],[64,240,255],[64,224,255],[64,208,255],[64,192,255],[32,160,255],[0,128,255],[0,96,255],[0,64,255],[0,56,255],[0,48,255],[0,40,255],[0,32,255],[0,24,255]];
function appLerp(stops,s){const x=s*(stops.length-1),
 i=Math.max(0,Math.min(stops.length-2,Math.floor(x))),f=x-i;
 const a=stops[i],b=stops[i+1];
 return [Math.round(a[0]+(b[0]-a[0])*f),Math.round(a[1]+(b[1]-a[1])*f),Math.round(a[2]+(b[2]-a[2])*f)];}
function rasterPNG(arr,W,H,lims){
 const cv=document.createElement('canvas');cv.width=W;cv.height=H;
 const ctx=cv.getContext('2d');const img=ctx.createImageData(W,H);const d=img.data;
 for(let i=0;i<W*H;i++){
  const v=arr[i];
  if(!isFinite(v))continue;
  const rgb=v>=0?appLerp(APP_POS,Math.min(1,v/lims.pos)):appLerp(APP_NEG,Math.min(1,-v/lims.neg));
  d[i*4]=rgb[0];d[i*4+1]=rgb[1];d[i*4+2]=rgb[2];d[i*4+3]=255;
 }
 ctx.putImageData(img,0,0);
 return cv.toDataURL('image/png');
}
function asymLims(arr){                              // neg=|p0.5|, pos=p99.5 (顯示單位)
 const fin=[];
 for(let i=0;i<arr.length;i++)if(isFinite(arr[i]))fin.push(arr[i]);
 if(!fin.length)return {neg:1,pos:1};
 fin.sort((a,b)=>a-b);
 const q=p=>{const t=p*(fin.length-1),i=Math.floor(t);
  return fin[i]+(fin[Math.min(i+1,fin.length-1)]-fin[i])*(t-i);};
 let neg=-q(0.005),pos=q(0.995);
 if(!(neg>0))neg=Math.max(pos*0.05,1e-6);
 if(!(pos>0))pos=Math.max(neg*0.05,1e-6);
 return {neg:+neg.toFixed(3),pos:+pos.toFixed(3)};
}

// ---- marching squares 等值線 (即時計算 + 快取) ----
let __appF=null;                                     // {vel,cum,W,H,gt} 顯示場 (cm)
const _ctrCache=new Map();
window.computeContours=function(mode,interval){
 if(!__appF)return [];
 const key=mode+'_'+interval;
 if(_ctrCache.has(key))return _ctrCache.get(key);
 const feats=marchingContours(mode==='cum'?__appF.cum:__appF.vel,__appF.W,__appF.H,__appF.gt,interval);
 _ctrCache.set(key,feats);
 return feats;
};
function marchLevel(f,W,H,lev){
 const segs=[];
 for(let r=0;r<H-1;r++){
  const ro=r*W;
  for(let c=0;c<W-1;c++){
   const v00=f[ro+c],v10=f[ro+c+1],v01=f[ro+W+c],v11=f[ro+W+c+1];
   if(!isFinite(v00)||!isFinite(v10)||!isFinite(v01)||!isFinite(v11))continue;
   let idx=0;
   if(v00>=lev)idx|=1;if(v10>=lev)idx|=2;if(v11>=lev)idx|=4;if(v01>=lev)idx|=8;
   if(idx===0||idx===15)continue;
   const T=[c+(lev-v00)/(v10-v00),r],Rt=[c+1,r+(lev-v10)/(v11-v10)],
         B=[c+(lev-v01)/(v11-v01),r+1],Lf=[c,r+(lev-v00)/(v01-v00)];
   switch(idx){
    case 1:case 14:segs.push([Lf,T]);break;
    case 2:case 13:segs.push([T,Rt]);break;
    case 3:case 12:segs.push([Lf,Rt]);break;
    case 4:case 11:segs.push([Rt,B]);break;
    case 5:segs.push([Lf,T],[Rt,B]);break;
    case 6:case 9:segs.push([T,B]);break;
    case 7:case 8:segs.push([Lf,B]);break;
    case 10:segs.push([T,Rt],[B,Lf]);break;
   }
  }
 }
 return segs;
}
function chainSegments(segs){                        // 端點雜湊把 2 點小段串成折線
 const key=p=>p[0].toFixed(3)+','+p[1].toFixed(3);
 const adj=new Map();
 const add=(k,i)=>{if(!adj.has(k))adj.set(k,[]);adj.get(k).push(i);};
 segs.forEach((s,i)=>{add(key(s[0]),i);add(key(s[1]),i);});
 const used=new Array(segs.length).fill(false);
 const lines=[];
 for(let i=0;i<segs.length;i++){
  if(used[i])continue;
  used[i]=true;
  const line=[segs[i][0],segs[i][1]];
  for(const dir of [1,0]){
   for(;;){
    const end=dir?line[line.length-1]:line[0];
    const k=key(end);
    const cands=(adj.get(k)||[]).filter(si=>!used[si]);
    if(!cands.length)break;
    const si=cands[0];used[si]=true;
    const s=segs[si];
    const nxt=key(s[0])===k?s[1]:s[0];
    if(dir)line.push(nxt);else line.unshift(nxt);
   }
  }
  lines.push(line);
 }
 return lines;
}
function marchingContours(f,W,H,gt,interval){
 let vmin=Infinity,vmax=-Infinity;
 for(let i=0;i<W*H;i++){const v=f[i];if(isFinite(v)){if(v<vmin)vmin=v;if(v>vmax)vmax=v;}}
 if(!(vmax>vmin))return [];
 const feats=[];
 for(let lev=Math.ceil(vmin/interval)*interval;lev<=vmax+1e-9;lev+=interval){
  const L=+lev.toFixed(6);
  for(const ln of chainSegments(marchLevel(f,W,H,L))){
   if(ln.length<2)continue;
   feats.push({type:'Feature',properties:{lev:+L.toFixed(3)},
    geometry:{type:'LineString',coordinates:ln.map(p=>[
     +(gt[0]+(p[0]+0.5)*gt[1]).toFixed(5),+(gt[3]+(p[1]+0.5)*gt[5]).toFixed(5)])}});
  }
 }
 return feats;
}

// ---- 按「載入並繪圖」才開始算 (兩種來源擇一) ----
let _busy=false;
// 重算前拆除前一輪的地圖與圖表 (同一頁重複 startViewer 用)
function teardownViewer(){
 try{if(window.__test.tsc)window.__test.tsc.destroy();}catch(e){}
 try{if(window.__test.pc)window.__test.pc.destroy();}catch(e){}
 try{window.__test.map.remove();}catch(e){}
 window._lyrCtrl=null;
 _ctrCache.clear();
}

// ---- 主流程: 使用者選項 → S → D → startViewer ----
// 逐像元 OLS 速度 (顯示單位/yr); 與 gmtsar2h5.py 的 ols_slope 同式
function olsVel(mm,N,n,t){
 const cnt=new Float64Array(n),sT=new Float64Array(n),sY=new Float64Array(n),
       sTY=new Float64Array(n),sTT=new Float64Array(n);
 for(let e=0;e<N;e++){
  const a=mm[e],te=t[e];
  for(let i=0;i<n;i++){
   const y=a[i];
   if(!isFinite(y))continue;
   cnt[i]++;sT[i]+=te;sY[i]+=y;sTY[i]+=te*y;sTT[i]+=te*te;
  }
 }
 const vel=new Float32Array(n).fill(NaN);
 for(let i=0;i<n;i++){
  if(cnt[i]<2)continue;
  const den=sTT[i]-sT[i]*sT[i]/cnt[i];
  if(!(den>1e-12))continue;
  vel[i]=(sTY[i]-sT[i]*sY[i]/cnt[i])/den;
 }
 return vel;
}
// 'YYYY-MM-DD' / 'YYYY' → 相對首期的十進位年
function yearsFrom(dates){
 const ms=dates.map(s=>{
  if(/^\d{4}$/.test(s))return Date.UTC(+s,0,1);
  const p=s.split('-');
  return Date.UTC(+p[0],(+p[1]||1)-1,+p[2]||1);
 });
 return ms.map(v=>(v-ms[0])/(365.25*86400000));
}

async function runInsarLoad(){
 if(_busy)return null;
 _busy=true;
 try{
  if(window.__test)teardownViewer();          // 重新載入: 先清掉舊圖層/地圖
  const fH5=document.getElementById('srcH5').files[0];
  const fCoh=document.getElementById('srcCoh').files[0];
  const dGmt=document.getElementById('srcGmtDir').files;
  const useGmt=(!fH5&&dGmt&&dGmt.length);
  if(!fH5&&!useGmt){appStatus(t('src_pick_first'));return null;}
  const U=document.getElementById('optUnit').value==='cm'?'cm':'mm';
  const uscale=(U==='cm')?0.1:1;
  const vlim=+document.getElementById('optVlim').value||0;
  const doDeramp=document.getElementById('optDeramp').checked;
  const notes=[];

  let S;
  if(fH5){
   appStatus(t('msg_reading_file'));await appTick();
   S=await parseTsH5(fH5);
  }else{
   const stepM=Math.max(10,+document.getElementById('srcGmtStep').value||100);
   const bufM=Math.max(0,+document.getElementById('srcGmtBuf').value||0);
   // 緩衝距離 → 半徑格數 (上限 12 格; 再大 IDW 逐格掃描會拖垮載入)
   const rad=Math.min(12,Math.round(bufM/stepM));
   S=await parseGmtsarDir(dGmt,stepM,rad);
   notes.push(t('gmt_note',{step:stepM,buf:rad*stepM,rad:rad}));
  }
  // 遮罩對兩種來源都適用 (尺寸不符時依經緯度重取樣)
  if(fCoh){
   const thr=+document.getElementById('srcCohMin').value;
   appStatus(t('coh_applying',{thr:thr}));await appTick();
   const r=await applyCohMask(S,fCoh,thr);
   notes.push(t('coh_note',{n:r.cut,thr:thr})+
    (r.mode==='resampled'?t('coh_resampled',{n:r.outside}):''));
  }
  // 速度場校正 (在 deramp 之前: 校正後趨勢已對齊參考速度, 通常就不需要再扣坡面)
  const fVel=document.getElementById('srcVel').files[0];
  if(fVel){
   const field=(document.getElementById('srcVelField').value||'field_3').trim();
   const unitK=document.getElementById('srcVelUnit').value==='m'?1000:1;
   const drTxt=(document.getElementById('srcVelDeramp').value||'').trim();
   let dr=null;
   if(drTxt){
    const a=drTxt.split(',').map(x=>+x.trim());
    if(a.length!==5||!a.every(isFinite))throw t('vel_bad_deramp');
    dr=a;
   }
   appStatus(t('vel_loading'));await appTick();
   const src=await loadVelSource(fVel,field,unitK,dr);
   appStatus(t('vel_retrending',{n:src.n}));await appTick();
   const r=retrendS(S,src.getV,yearsFrom(S.dates));
   notes.push(t('vel_note',{n:src.n,fix:r.fix,skip:r.skip})+(dr?t('vel_note_deramp'):''));
  }
  if(doDeramp){
   appStatus(t('deramp_running'));await appTick();
   const nf=derampS(S);
   notes.push(t('deramp_note',{n:nf}));
  }
  const D=await buildD(S,U,uscale,vlim);
  // 一定要回報實際生效的色階: 自動色階取 p0.5/p99.5, 對「大部分穩定、少數極端」的
  // 資料會壓縮到很窄的範圍, 同一份資料看起來會跟指定上限時差很多 (不是數值變了)
  const L=D.lims.vel;
  notes.push(vlim>0?t('lims_fixed',{a:(-L.neg).toFixed(1),b:L.pos.toFixed(1),U:U})
                   :t('lims_auto',{a:(-L.neg).toFixed(1),b:L.pos.toFixed(1),U:U}));
  appStatus(notes.join('\n'));
  document.getElementById('appwarnText').textContent=notes.join('；');
  document.getElementById('appwarn').style.display=notes.length?'':'none';
  return D;
 }catch(err){
  console.error(err);
  appStatus(t('app_error_prefix')+err);
  throw err;
 }finally{_busy=false;}
}

// S (mm 格網) → D → startViewer(D)
async function buildD(S,U,uscale,vlim){
 const {dates,N,H,W,gt}=S,n=H*W;
 appStatus(t('app_preparing'));await appTick();
 const tYr=yearsFrom(dates);
 if(!tYr.every(isFinite)||!(tYr[N-1]>0))throw t('h5_bad_dates');
 // 顯示場: 速度 (全期 OLS 回歸) 與 總累積 (末期 − 首期)
 const velMM=olsVel(S.mm,N,n,tYr);
 const velF=new Float32Array(n).fill(NaN),cumF=new Float32Array(n).fill(NaN);
 const first=S.mm[0],last=S.mm[N-1];
 for(let i=0;i<n;i++){
  if(isFinite(velMM[i]))velF[i]=velMM[i]*uscale;
  const a=first[i],b=last[i];
  if(isFinite(a)&&isFinite(b))cumF[i]=(b-a)*uscale;
 }
 const validIdx=[];
 for(let i=0;i<n;i++)if(isFinite(velF[i]))validIdx.push(i);
 if(!validIdx.length)throw t('err_no_valid_px_after');
 window.__lastValidCount=validIdx.length;
 const lims=vlim>0?{vel:{neg:vlim,pos:vlim},cum:{neg:vlim,pos:vlim}}
                  :{vel:asymLims(velF),cum:asymLims(cumF)};
 const rasters={vel:rasterPNG(velF,W,H,lims.vel),cum:rasterPNG(cumF,W,H,lims.cum),
  bounds:[[gt[3]+gt[5]*H,gt[0]],[gt[3],gt[0]+gt[1]*W]]};
 // 量化格網 (0.1 mm/step) 供地圖點擊/剖面取樣
 appStatus(t('h5_quantizing'));await appTick();
 const qarr=new Int16Array(N*n).fill(-32768);
 for(let e=0;e<N;e++){
  const a=S.mm[e],off=e*n;
  for(let i=0;i<n;i++){
   const v=a[i];
   if(isFinite(v))qarr[off+i]=Math.max(-32767,Math.min(32767,Math.round(v/0.1)));
  }
 }
 // 代表點 (fitBounds/剖面/匯出用, 上限 20000)
 const stride=Math.max(1,Math.ceil(validIdx.length/20000));
 const nd=(U==='cm')?2:1;
 const points=[];
 for(let k=0;k<validIdx.length;k+=stride){
  const i=validIdx[k],r=Math.floor(i/W),c=i%W;
  const rec={lon:+(gt[0]+(c+0.5)*gt[1]).toFixed(6),lat:+(gt[3]+(r+0.5)*gt[5]).toFixed(6),
   v:+velF[i].toFixed(nd),t:[]};
  for(let e=0;e<N;e++){const v=S.mm[e][i];rec.t.push(isFinite(v)?+(v*uscale).toFixed(nd):null);}
  points.push(rec);
 }
 const gnss=(window.GNSS_ALL||[]).filter(g=>g.lon>=gt[0]&&g.lon<=gt[0]+gt[1]*W&&
  g.lat<=gt[3]&&g.lat>=gt[3]+gt[5]*H);
 __appF={vel:velF,cum:cumF,W,H,gt};
 // velAlgo 必須明寫 'ols': 清除遮罩/按「自動」色階等操作會走 refreshDerived, 把
 // window.DISPF 重設為 null (h5F 只有「載入 HDF5」按鈕那條路徑會設), 之後數值場改由
 // computeFieldFromGridT 現場算 — 它讀 D.velAlgo, 沒給就退成 'endpoint' (首尾期差分),
 // 速度場會在使用者沒察覺的情況下從回歸值跳成差分值 (實測差可達 13 mm/yr)
 const D={dates,unit:U,lims,hasLine:false,n:points.length,buffer:50,line:[],bearing:0,
  velAlgo:'ols',
  points,grid:{arr:qarr,shape:[N,H,W],gt,scale:0.1*uscale,nodata:-32768},gnss,
  rasters,contours:{dynamic:true,base:(U==='cm')?0.1:1}};
 appStatus(t('app_rendering'));await appTick();
 document.getElementById('apppanel').style.display='none';
 document.body.classList.remove('preload');
 startViewer(D);
 // startViewer(D) 內部會把 window.DISPF 重置為 null, 故登錄須排在呼叫之後
 window.DISPF={vel:velF,cum:cumF,src:'app'};
 window.__APPF={vel:velF,cum:cumF};      // refreshDerived 的 fallback 來源, 見該函式註解
 window.__APP_D=D;
 window.__S=S;                                // 供 GNSS 疊圖/重算 deramp 用
 showShareBtn();
 return D;
}
document.getElementById('runLoad').addEventListener('click',()=>{runInsarLoad().catch(()=>{});});
// 選了 HDF5 就清掉資料夾選擇 (反之亦然), 避免兩個來源同時給值時的歧義
document.getElementById('srcH5').addEventListener('change',()=>{
 if(document.getElementById('srcH5').files.length)document.getElementById('srcGmtDir').value='';});
document.getElementById('srcGmtDir').addEventListener('change',()=>{
 if(document.getElementById('srcGmtDir').files.length)document.getElementById('srcH5').value='';});

// ==================== 匯出分享版 HTML (結果烘進檔案, 開啟不需 grd) ====================
async function gzipB64(u8){
 const stream=new Blob([u8]).stream().pipeThrough(new CompressionStream('gzip'));
 const b=new Uint8Array(await new Response(stream).arrayBuffer());
 let s='';
 for(let i=0;i<b.length;i+=32768)s+=String.fromCharCode.apply(null,b.subarray(i,i+32768));
 return btoa(s);
}
async function gunzipB64ToText(b64){
 const bin=atob(b64);
 const u8=new Uint8Array(bin.length);
 for(let i=0;i<bin.length;i++)u8[i]=bin.charCodeAt(i);
 const stream=new Blob([u8]).stream().pipeThrough(new DecompressionStream('gzip'));
 return await new Response(stream).text();
}
// 分享版等值線 bootstrap: 由解壓後 GRID 重建顯示場 → window.computeContours (字串常數, 嵌進匯出檔)
const SHARE_BOOT=[
"const _shCache=new Map();let _shF=null;",
"window.computeContours=function(mode,interval){",
" if(!_shF)return [];",
" const k=mode+'_'+interval;",
" if(!_shCache.has(k))_shCache.set(k,marchingContours(mode==='cum'?_shF.cum:_shF.vel,_shF.W,_shF.H,_shF.gt,interval));",
" return _shCache.get(k);",
"};",
"(async function(){",
" for(let i=0;i<600&&!(window.__test&&window.__test.GRID);i++)await new Promise(r=>setTimeout(r,100));",
" const G=window.__test.GRID;",
" if(!G)return;",
" const span=(+D.dates[D.dates.length-1])-(+D.dates[0])||1;",
" const n=G.W*G.H,off=(G.N-1)*n;",
" const vel=new Float32Array(n).fill(NaN),cum=new Float32Array(n).fill(NaN);",
" for(let i=0;i<n;i++){",
"  const q0=G.arr[i],q1=G.arr[off+i];",
"  if(q0===G.nodata||q1===G.nodata)continue;",
"  vel[i]=(q1-q0)*G.scale/span;cum[i]=q1*G.scale;",
" }",
" _shF={vel,cum,W:G.W,H:G.H,gt:G.gt};",
" window.DISPF={vel:_shF.vel,cum:_shF.cum,src:'share'};",
" window.__test.buildContours(document.getElementById('layerMode').value);",
"})();"].join("\n");
// 純函式: 回傳自足 HTML 字串 (E2E 可直接驗證)
async function buildShareHTML(){
 if(!window.__APP_D)throw t('err_calc_not_done');
 const D=window.__APP_D;
 const tpl=await gunzipB64ToText(APP_SHARE_TPL_B64);
 const g=D.grid;
 const b64=await gzipB64(new Uint8Array(g.arr.buffer,g.arr.byteOffset,g.arr.byteLength));
 // 目前檢視狀態快照: 遮罩後重繪的 raster 與生效中的色階範圍一併帶進分享檔
 const snapR=(window.__test&&window.__test.clientRasters)||null;
 const snapL=window.__test?window.__test.effLimsSnapshot:null;
 // gnss 另包 GNSS 起訖標記 (與 CLI main() 的 _embed_gnss_json 同一約定), 讓分享版
 // HTML 開啟後仍可再次載入 GNSS CSV / 再匯出更新版 (標記需保留才能鏈式運作). 標記常數
 // 刻意用字串相加組成 (理由同主 template 內 buildGnssUpdatedHTML 旁的說明).
 const GNSS_PH='__LEVELING_GNSS_PLACEHOLDER__';
 const _GS='/*__GNSS'+'_S__*/', _GE='/*__GNSS'+'_E__*/';
 const D2=Object.assign({},D,
  {grid:{b64:b64,shape:g.shape,gt:g.gt,scale:g.scale,nodata:g.nodata}},
  snapR?{rasters:Object.assign({},snapR,{bounds:D.rasters.bounds})}:{},
  snapL?{lims:snapL}:{},
  {gnss:GNSS_PH});
 const gnssSeg=_GS+JSON.stringify(D.gnss||[])+_GE;
 const dataStr=JSON.stringify(D2).replace(JSON.stringify(GNSS_PH),gnssSeg);
 let html=tpl.replace('/*__DATA__*/',dataStr);
 // 等值線模組: marching 函式原始碼 + bootstrap (動態計算沿用)
 const mod='<script>\n'+marchLevel.toString()+'\n'+chainSegments.toString()+'\n'
  +marchingContours.toString()+'\n'+SHARE_BOOT+'\n<'+'/script>';
 // 必須錨到「最後一個」</'+'body></html>: 內嵌的 SheetJS 原始碼裡也有這串字面,
 // 用 String.replace 會插進它的字串中間, 整個分享版變成語法錯 (等值線模組失效)
 const k=html.lastIndexOf('</body></html>');
 if(k<0)throw t('err_share_anchor');
 return html.slice(0,k)+mod+html.slice(k);
}
async function exportShare(){
 const btn=document.getElementById('shareBtn');
 try{
  if(btn){btn.disabled=true;btn.textContent=t('btn_exporting');}
  const html=await buildShareHTML();
  const blob=new Blob([html],{type:'text/html'});
  const a=document.createElement('a');
  a.href=URL.createObjectURL(blob);
  a.download='ks_share_'+window.__APP_D.dates[0]+'_'
   +window.__APP_D.dates[window.__APP_D.dates.length-1]+'.html';
  a.click();
  setTimeout(()=>URL.revokeObjectURL(a.href),5000);
 }catch(err){console.error(err);alert(t('alert_export_failed')+err);}
 finally{if(btn){btn.disabled=false;btn.textContent=t('btn_export_share');}}
}
function showShareBtn(){
 if(document.getElementById('shareBtn'))return;
 const b=document.createElement('button');
 b.id='shareBtn';b.className='panel';
 b.textContent=t('btn_export_share');
 b.title=t('title_export_share');
 b.style.cssText='top:10px;right:452px;padding:4px 9px;z-index:1500';
 b.onclick=exportShare;
 document.body.appendChild(b);
}
"""


if __name__ == "__main__":
    main()
